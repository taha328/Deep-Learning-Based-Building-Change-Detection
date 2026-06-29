from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, date, datetime
from contextvars import ContextVar
import csv
from io import BytesIO
import html
import hashlib
import json
import logging
import math
from pathlib import Path
import re
import shutil
import tempfile
import time
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from typing import Any

import geopandas as gpd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from shapely.geometry import GeometryCollection, MultiPolygon, Polygon, box, mapping, shape
from shapely.geometry.base import BaseGeometry
from shapely.geometry.polygon import orient
from shapely.ops import unary_union

from src.config import Settings, get_settings
from src.domain.cache import load_cached_response
from src.schemas import RunResponse, TemporalMilestone, TemporalMilestoneMetrics, TemporalProject
from src.services.temporal_projects import (
    _ensure_temporal_derived_geometry_layers,
    _hydrate_milestone_buffer_layers,
    _hydrate_temporal_layer_artifacts,
    get_temporal_project,
    resolve_temporal_project_artifact_path,
)
from src.utils.geometry import centroid_lonlat, reproject_geometry, utm_epsg_from_lonlat


SURFACE_CRS_MOROCCO = "EPSG:32629"
KML_NS = "http://www.opengis.net/kml/2.2"
GX_NS = "http://www.google.com/kml/ext/2.2"
DEFAULT_GOOGLE_EARTH_LOOKAT_RANGE_METERS = 2000.0
logger = logging.getLogger(__name__)


TEMPORAL_RESULTS_EXPORT_MEDIA_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "kml": "application/vnd.google-earth.kml+xml",
    "geojson": "application/geo+json",
    "topojson": "application/json",
    "json": "application/json",
    "tsv": "text/tab-separated-values; charset=utf-8",
    "shapefile": "application/zip",
}

TEMPORAL_RESULTS_EXPORT_FILENAMES = {
    "xlsx": "results.xlsx",
    "kml": "results.kml",
    "geojson": "results.geojson",
    "topojson": "results.topojson",
    "json": "results.json",
    "tsv": "results_powerbi.tsv",
    "shapefile": "results_shapefile.zip",
}

TEMPORAL_RESULTS_EXPORT_LABELS = {
    "additions": "Building change polygons / additions",
    "cumulative_growth": "Cumulative growth",
    "buffer_10m": "Cumulative Building-change buffer 10 m",
    "buffer_15m": "Cumulative Building-change buffer 15 m",
    "buffer_20m": "Cumulative Building-change buffer 20 m",
    "diagnostics": "Addition candidate diagnostics",
}

TOPOJSON_DEFAULT_QUANTIZATION = 1_000_000
TOPOJSON_EXPORT_VERSION = "clean-quantized-v3"
SHAPEFILE_EXPORT_VERSION = "zone-clipped-mutually-exclusive-qgz-v14-startup-extent"
QGIS_PROJECT_CRS = "EPSG:3857"
QGIS_VECTOR_CRS = "EPSG:4326"
QGIS_PROJECT_EXTENT_PADDING_RATIO = 0.075
TOPOJSON_ALLOWED_LAYERS = ("additions", "buffer_10m")
POWERBI_TSV_COLUMNS = (
    "project_id",
    "project_name",
    "date",
    "layer_type",
    "area_m2",
    "centroid_lon",
    "centroid_lat",
)
POWERBI_TSV_LAYER_ORDER = {
    "additions": 0,
    "buffer_10m": 1,
    "buffer_15m": 2,
    "buffer_20m": 3,
}
TOPOJSON_PROPERTY_KEYS = ("id", "project", "date", "year", "period", "layer", "area_m2", "area_ha")
TOPOJSON_LAYER_ID_SLUGS = {
    "additions": "additions",
    "buffer_10m": "buffer10m",
    "cumulative_growth": "cumulative-growth",
}
TOPOJSON_REMOVED_PROPERTY_KEYS = {
    "run_id",
    "release_identifier",
    "release_t1",
    "release_t2",
    "src_date_t1",
    "src_date_t2",
    "source_backend",
    "feature_index",
    "buffer_id",
    "buffer_part_index",
    "source_change_block_id",
    "source_change_count",
    "block_gap_m",
    "cluster_gap_m",
    "kind",
    "release_date",
    "source_building_count",
    "confidence",
    "status",
}
TOPOJSON_ID_PATTERN = re.compile(r"^[0-9]{4}-[a-z0-9-]+-[0-9]{6}$")
EXPORT_CACHE_VERSION = "temporal-results-file-backed-v2"
EXPORT_ARTIFACT_MANIFEST_VERSION = "temporal-results-export-artifact-manifest-v1"
EXPORT_GEOMETRY_HASH_PRECISION = 7
EXPORT_FORMAT_VERSIONS = {
    "xlsx": "xlsx-v2",
    "kml": "kml-v2",
    "geojson": "geojson-v2",
    "topojson": TOPOJSON_EXPORT_VERSION,
    "json": "json-v2",
    "tsv": "tsv-v2",
    "shapefile": SHAPEFILE_EXPORT_VERSION,
}
EXPORT_RESULT_ARTIFACT_KEYS = {
    "additions": ("additions", "additions_geojson", None),
    "buffer_10m": ("building_change_buffer_10m", "buffer_layers_geojson", "10m"),
    "buffer_15m": ("building_change_buffer_15m", "buffer_layers_geojson", "15m"),
    "buffer_20m": ("building_change_buffer_20m", "buffer_layers_geojson", "20m"),
}


@dataclass(frozen=True)
class KmlMilestone:
    source_index: int
    milestone: TemporalMilestone
    archive_date: date
    archive_date_text: str
    date_note: str
    geometry: BaseGeometry


@dataclass(frozen=True)
class KmlView:
    lon: float
    lat: float
    range_m: float = DEFAULT_GOOGLE_EARTH_LOOKAT_RANGE_METERS


@dataclass(frozen=True)
class TemporalShapefileLayer:
    group_key: str
    group_label: str
    filename: str
    display_name: str
    artifact_key: str
    release_identifier: str
    feature_collection: dict[str, Any]
    is_global: bool = False


@dataclass(frozen=True)
class ExportedLayerValidation:
    path: Path
    display_name: str
    group_name: str
    relative_path: str
    feature_count: int
    crs_authid: str
    xmin: float
    ymin: float
    xmax: float
    ymax: float
    is_valid_for_qgis_project: bool
    reason: str | None = None


@dataclass(frozen=True)
class TemporalQgisRaster:
    release_identifier: str
    period_label: str
    source_path: Path
    relative_path: str
    layer_id: str
    display_name: str
    bounds_3857: Bounds


@dataclass(frozen=True)
class QgisProjectInitialExtent:
    bounds_3857: Bounds
    source_count: int
    used_fallback: bool = False


Bounds = tuple[float, float, float, float]
TEMPORAL_WEB_MILESTONE_COLORS = ("#00B050", "#FFD700", "#0066FF", "#E31A1C", "#00C8C8", "#FF1493", "#7FFF00", "#8B4513")
_EXPORT_CONTEXT: ContextVar[dict[str, Any] | None] = ContextVar("temporal_export_context", default=None)


def _settings(settings: Settings | None) -> Settings:
    return settings or get_settings()


def _export_now() -> datetime:
    return datetime.now(UTC).replace(microsecond=0)


def _load_project(project_id: str, settings: Settings) -> TemporalProject:
    context = _EXPORT_CONTEXT.get()
    cached_project = context.get("loaded_project") if isinstance(context, dict) else None
    if isinstance(cached_project, TemporalProject) and cached_project.project_id == project_id:
        project = cached_project.model_copy(deep=True)
        scope_type = context.get("scope_type") or ("custom_geometry" if context.get("geometry") is not None else "project_aoi")
        logger.info(
            "EXPORT_FULL_PROJECT_LOAD_SKIPPED projectId=%s scopeType=%s reason=reuse_loaded_project",
            project_id,
            scope_type,
        )
        if scope_type == "custom_geometry":
            logger.info(
                "EXPORT_SCOPE_APPLIED projectId=%s scopeType=custom_geometry source=%s wasClippedToProjectAoi=%s",
                project.project_id,
                context.get("source"),
                context.get("was_clipped_to_project_aoi"),
            )
            return _clip_project_to_export_geometry(project, context["geometry"])
        logger.info("EXPORT_SCOPE_APPLIED projectId=%s scopeType=project_aoi", project.project_id)
        return project

    project = get_temporal_project(project_id, settings)
    project = _hydrate_temporal_layer_artifacts(project, settings)
    project = _hydrate_milestone_buffer_layers(project, settings)
    project = _hydrate_export_file_backed_result_artifacts(project, settings)
    project = _ensure_temporal_derived_geometry_layers(project)
    if context and (context.get("scope_type") == "custom_geometry" or context.get("geometry") is not None):
        logger.info(
            "EXPORT_SCOPE_APPLIED projectId=%s scopeType=custom_geometry source=%s wasClippedToProjectAoi=%s",
            project.project_id,
            context.get("source"),
            context.get("was_clipped_to_project_aoi"),
        )
        return _clip_project_to_export_geometry(project, context["geometry"])
    logger.info("EXPORT_SCOPE_APPLIED projectId=%s scopeType=project_aoi", project.project_id)
    return project


def _polygonal_geometry(value: BaseGeometry) -> BaseGeometry:
    if value.geom_type in {"Polygon", "MultiPolygon"}:
        return value.buffer(0)
    if value.geom_type == "GeometryCollection":
        polygons = [part for part in value.geoms if part.geom_type in {"Polygon", "MultiPolygon"}]
        return unary_union(polygons).buffer(0) if polygons else GeometryCollection()
    return GeometryCollection()


def _oriented_polygonal_geometry(geometry: BaseGeometry) -> BaseGeometry:
    geometry = _polygonal_geometry(geometry)
    if isinstance(geometry, Polygon):
        return orient(geometry, sign=1.0)
    if isinstance(geometry, MultiPolygon):
        return MultiPolygon([orient(part, sign=1.0) for part in geometry.geoms if not part.is_empty])
    return geometry


def _canonicalize_geojson_coordinates(value: Any) -> Any:
    if isinstance(value, (int, float)):
        rounded = round(float(value), EXPORT_GEOMETRY_HASH_PRECISION)
        return 0.0 if rounded == -0.0 else rounded
    if isinstance(value, (list, tuple)):
        return [_canonicalize_geojson_coordinates(item) for item in value]
    return value


def _canonical_export_geometry_payload(geometry_payload: dict[str, Any]) -> dict[str, Any]:
    geometry = _oriented_polygonal_geometry(shape(geometry_payload))
    if geometry.is_empty:
        raise ValueError("Export perimeter geometry is empty.")
    payload = mapping(geometry)
    return {
        "type": payload.get("type"),
        "coordinates": _canonicalize_geojson_coordinates(payload.get("coordinates")),
    }


def _stable_custom_geometry_hash(perimeter: dict[str, Any] | None) -> str | None:
    if _is_project_aoi_export_scope(perimeter):
        return None
    geometry_payload = perimeter.get("geometry") if isinstance(perimeter, dict) else None
    if not isinstance(geometry_payload, dict):
        return None
    canonical = {
        "crs": "EPSG:4326",
        "geometry": _canonical_export_geometry_payload(geometry_payload),
    }
    return hashlib.sha256(json.dumps(canonical, sort_keys=True, separators=(",", ":")).encode("utf-8")).hexdigest()


def resolve_export_perimeter(project: TemporalProject, perimeter: dict[str, Any] | None) -> dict[str, Any] | None:
    if not perimeter or perimeter.get("mode") == "project_aoi":
        return None
    if not project.aoi_geojson:
        raise ValueError("Project AOI is unavailable.")
    geometry_payload = perimeter.get("geometry")
    try:
        requested = _polygonal_geometry(shape(geometry_payload))
        project_aoi = _polygonal_geometry(shape(project.aoi_geojson))
    except Exception as exc:
        raise ValueError(f"Invalid export perimeter geometry: {exc}") from exc
    if requested.is_empty:
        raise ValueError("Export perimeter geometry is empty.")
    clipped = _polygonal_geometry(requested.intersection(project_aoi))
    if clipped.is_empty:
        raise ValueError("La zone sélectionnée est hors de l’AOI du projet.")
    return {
        "scope_type": "custom_geometry",
        "mode": "custom_geometry",
        "source": perimeter.get("source"),
        "geometry": clipped,
        "geometry_hash": _stable_custom_geometry_hash(perimeter),
        "was_clipped_to_project_aoi": not clipped.equals(requested),
    }


def _clip_feature_collection(payload: dict[str, Any] | None, mask: BaseGeometry) -> dict[str, Any] | None:
    if not isinstance(payload, dict):
        return payload
    clipped_features: list[dict[str, Any]] = []
    input_count = 0
    bbox_skipped = 0
    mask_bounds = mask.bounds
    for feature in _features(payload):
        input_count += 1
        geometry_payload = feature.get("geometry") if isinstance(feature, dict) else None
        if not isinstance(geometry_payload, dict):
            continue
        try:
            source_geometry = _polygonal_geometry(shape(geometry_payload))
        except Exception:
            continue
        if source_geometry.is_empty:
            continue
        minx, miny, maxx, maxy = source_geometry.bounds
        if maxx < mask_bounds[0] or minx > mask_bounds[2] or maxy < mask_bounds[1] or miny > mask_bounds[3]:
            bbox_skipped += 1
            continue
        geometry = _polygonal_geometry(source_geometry.intersection(mask))
        if geometry.is_empty:
            continue
        clipped = dict(feature)
        clipped["geometry"] = mapping(geometry)
        clipped_features.append(clipped)
    logger.info(
        "EXPORT_SPATIAL_PREFILTER_DONE inputFeatures=%s outputFeatures=%s bboxSkipped=%s",
        input_count,
        len(clipped_features),
        bbox_skipped,
    )
    return {**payload, "features": clipped_features}


def _clip_project_to_export_geometry(project: TemporalProject, mask: BaseGeometry) -> TemporalProject:
    clipped = project.model_copy(deep=True)
    clipped.aoi_geojson = mapping(mask)
    geojson_fields = (
        "automated_additions_geojson",
        "automated_candidate_footprint_geojson",
        "automated_building_blocks_geojson",
        "manual_override_geojson",
        "additions_geojson",
        "effective_building_blocks_geojson",
        "effective_footprint_geojson",
        "cumulative_union_geojson",
        "cumulative_growth_blocks_geojson",
        "cumulative_growth_envelope_geojson",
    )
    for milestone in clipped.milestones:
        for field_name in geojson_fields:
            setattr(milestone, field_name, _clip_feature_collection(getattr(milestone, field_name), mask))
        milestone.buffer_layers_geojson = {
            key: value
            for key, payload in milestone.buffer_layers_geojson.items()
            if (value := _clip_feature_collection(payload, mask)) is not None
        }
    return clipped


def _export_metadata() -> dict[str, Any]:
    context = _EXPORT_CONTEXT.get()
    if not context or context.get("scope_type") == "project_aoi":
        return {"perimeter_mode": "project_aoi"}
    return {
        "perimeter_mode": "custom_geometry",
        "perimeter_source": context.get("source"),
        "was_clipped_to_project_aoi": context.get("was_clipped_to_project_aoi"),
        "perimeter_geometry": mapping(context["geometry"]),
    }


def _metric_crs(project: TemporalProject) -> str:
    if project.aoi_geojson:
        try:
            geometry = shape(project.aoi_geojson)
            lon, lat = centroid_lonlat(geometry)
            if -13.5 <= lon <= -0.5 and 20.0 <= lat <= 36.5:
                return SURFACE_CRS_MOROCCO
            return f"EPSG:{utm_epsg_from_lonlat(lon, lat)}"
        except Exception:
            return SURFACE_CRS_MOROCCO
    return SURFACE_CRS_MOROCCO


def _features(payload: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    features = payload.get("features")
    return features if isinstance(features, list) else []


def _set_result_layer_payload(milestone: TemporalMilestone, layer_type: str, payload: dict[str, Any]) -> None:
    if layer_type == "additions":
        milestone.additions_geojson = payload
        return
    _artifact_key, field_name, distance_key = EXPORT_RESULT_ARTIFACT_KEYS[layer_type]
    if field_name == "buffer_layers_geojson" and distance_key is not None:
        milestone.buffer_layers_geojson[distance_key] = payload


def _existing_result_layer_payload(milestone: TemporalMilestone, layer_type: str) -> dict[str, Any] | None:
    if layer_type == "additions":
        return milestone.additions_geojson
    _artifact_key, field_name, distance_key = EXPORT_RESULT_ARTIFACT_KEYS[layer_type]
    if field_name == "buffer_layers_geojson" and distance_key is not None:
        return milestone.buffer_layers_geojson.get(distance_key) or milestone.buffer_layers_geojson.get(distance_key.removesuffix("m"))
    return None


def _load_export_geojson_artifact(
    *,
    project_id: str,
    release_identifier: str,
    artifact_key: str,
    settings: Settings,
) -> tuple[dict[str, Any] | None, Path | None]:
    try:
        path, media_type = resolve_temporal_project_artifact_path(
            project_id=project_id,
            release_identifier=release_identifier,
            artifact_key=artifact_key,
            settings=settings,
            access_mode="export_results",
        )
    except FileNotFoundError:
        return None, None
    if media_type != "application/geo+json":
        logger.info(
            "EXPORT_LAYER_EMPTY_SKIPPED projectId=%s release=%s artifact_key=%s reason=unsupported_media_type mediaType=%s",
            project_id,
            release_identifier,
            artifact_key,
            media_type,
        )
        return None, path
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning(
            "EXPORT_LAYER_EMPTY_SKIPPED projectId=%s release=%s artifact_key=%s path=%s reason=invalid_geojson error=%s",
            project_id,
            release_identifier,
            artifact_key,
            path,
            exc.__class__.__name__,
        )
        return None, path
    if not isinstance(payload, dict) or payload.get("type") != "FeatureCollection":
        logger.info(
            "EXPORT_LAYER_EMPTY_SKIPPED projectId=%s release=%s artifact_key=%s path=%s reason=not_feature_collection",
            project_id,
            release_identifier,
            artifact_key,
            path,
        )
        return None, path
    feature_count = len(_features(payload))
    if feature_count <= 0:
        logger.info(
            "EXPORT_LAYER_EMPTY_SKIPPED projectId=%s release=%s artifact_key=%s path=%s reason=empty_feature_collection",
            project_id,
            release_identifier,
            artifact_key,
            path,
        )
        return payload, path
    logger.info(
        "EXPORT_FILE_BACKED_LAYER_SELECTED projectId=%s release=%s artifact_key=%s path=%s featureCount=%s bytes=%s",
        project_id,
        release_identifier,
        artifact_key,
        path,
        feature_count,
        path.stat().st_size,
    )
    logger.info(
        "EXPORT_LAYER_NON_EMPTY projectId=%s release=%s artifact_key=%s source=file_backed featureCount=%s",
        project_id,
        release_identifier,
        artifact_key,
        feature_count,
    )
    return payload, path


def _hydrate_export_file_backed_result_artifacts(project: TemporalProject, settings: Settings) -> TemporalProject:
    for milestone in project.milestones:
        if milestone.status != "complete":
            continue
        for layer_type, (artifact_key, _field_name, _distance_key) in EXPORT_RESULT_ARTIFACT_KEYS.items():
            existing_payload = _existing_result_layer_payload(milestone, layer_type)
            existing_count = len(_features(existing_payload))
            if existing_count > 0:
                logger.info(
                    "EXPORT_LAYER_NON_EMPTY projectId=%s release=%s artifact_key=%s source=project_payload featureCount=%s",
                    project.project_id,
                    milestone.release_identifier,
                    artifact_key,
                    existing_count,
                )
                continue
            payload, path = _load_export_geojson_artifact(
                project_id=project.project_id,
                release_identifier=milestone.release_identifier,
                artifact_key=artifact_key,
                settings=settings,
            )
            if payload is None:
                logger.info(
                    "EXPORT_LAYER_EMPTY_SKIPPED projectId=%s release=%s artifact_key=%s reason=no_resolved_artifact",
                    project.project_id,
                    milestone.release_identifier,
                    artifact_key,
                )
                continue
            if len(_features(payload)) <= 0:
                continue
            _set_result_layer_payload(milestone, layer_type, payload)
            logger.info(
                "EXPORT_LAYER_NON_EMPTY projectId=%s release=%s artifact_key=%s source=file_backed_hydrated path=%s featureCount=%s",
                project.project_id,
                milestone.release_identifier,
                artifact_key,
                path,
                len(_features(payload)),
            )
    return project


def _geometry_from_geojson(payload: dict[str, Any] | None) -> BaseGeometry:
    geometries: list[BaseGeometry] = []
    for feature in _features(payload):
        if not isinstance(feature, dict):
            continue
        geometry_payload = feature.get("geometry")
        if not isinstance(geometry_payload, dict):
            continue
        try:
            geometry = shape(geometry_payload).buffer(0)
        except Exception:
            continue
        if geometry.is_empty or geometry.geom_type not in {"Polygon", "MultiPolygon"}:
            continue
        geometries.append(geometry)
    if not geometries:
        return GeometryCollection()
    return unary_union(geometries).buffer(0)


def _area_m2(payload: dict[str, Any] | None, crs: str) -> float | None:
    geometry = _geometry_from_geojson(payload)
    if geometry.is_empty:
        return 0.0
    try:
        return float(reproject_geometry(geometry, "EPSG:4326", crs).area)
    except Exception:
        return None


def _powerbi_tsv_area_m2(payload: dict[str, Any] | None, crs: str) -> float | str:
    geometry = _geometry_from_geojson(payload)
    if geometry.is_empty:
        return ""
    try:
        return float(reproject_geometry(geometry, "EPSG:4326", crs).area)
    except Exception:
        return ""


def _float(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return None
    return None


def _percent(numerator: float | None, denominator: float | None) -> float | None:
    if numerator is None or denominator is None or denominator <= 0:
        return None
    return numerator / denominator


def _date_string(value: str | None) -> str | None:
    if not value:
        return None
    match = re.search(r"\d{4}-\d{2}-\d{2}", value)
    if match:
        return match.group(0)
    match = re.search(r"(20\d{2})", value)
    return f"{match.group(1)}-01-01" if match else None


def _parse_date(value: str | None) -> date | None:
    parsed = _date_string(value)
    if not parsed:
        return None
    try:
        return date.fromisoformat(parsed)
    except ValueError:
        return None


def _date_cell(value: str | None) -> date | str | None:
    parsed = _date_string(value)
    if not parsed:
        return value
    try:
        return date.fromisoformat(parsed)
    except ValueError:
        return value


def _run_response_for_milestone(project: TemporalProject, milestone: TemporalMilestone, settings: Settings) -> tuple[RunResponse | None, str | None]:
    if milestone.pair_request_hash:
        return load_cached_response(settings, milestone.pair_request_hash), "t2"
    milestones = project.milestones
    index = milestones.index(milestone)
    if index + 1 < len(milestones):
        next_hash = milestones[index + 1].pair_request_hash
        if next_hash:
            return load_cached_response(settings, next_hash), "t1"
    return None, None


def _imagery_source(project: TemporalProject, milestone: TemporalMilestone) -> str:
    return "ESRI Wayback"


def _archive_date(
    project: TemporalProject,
    milestone: TemporalMilestone,
    settings: Settings,
    export_now: datetime,
) -> tuple[str | None, str]:
    run_response, side = _run_response_for_milestone(project, milestone, settings)
    summary = run_response.summary if run_response else None

    if summary is not None and side == "t2":
        for value in (summary.dominant_src_date_t2, summary.release_date_t2):
            parsed = _date_string(value)
            if parsed:
                note = "" if value == summary.dominant_src_date_t2 else "Date utilisée: date de publication, date d'acquisition indisponible"
                return parsed, note
    if summary is not None and side == "t1":
        for value in (summary.dominant_src_date_t1, summary.release_date_t1):
            parsed = _date_string(value)
            if parsed:
                note = "" if value == summary.dominant_src_date_t1 else "Date utilisée: date de publication, date d'acquisition indisponible"
                return parsed, note

    parsed = _date_string(milestone.release_date)
    if parsed:
        return parsed, "Date utilisée: date de publication, date d'acquisition indisponible"

    label_year = re.search(r"(20\d{2})", milestone.release_identifier)
    if label_year:
        return f"{label_year.group(1)}-01-01", "Date utilisée: année du libellé, date d'acquisition indisponible"

    return None, "Date d'archive indisponible"


def _metrics(milestone: TemporalMilestone) -> TemporalMilestoneMetrics:
    return milestone.metrics or TemporalMilestoneMetrics()


def _milestone_rows(project: TemporalProject, settings: Settings, export_now: datetime) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    completed = [milestone for milestone in project.milestones if milestone.status == "complete"]
    for index, milestone in enumerate(completed):
        metrics = _metrics(milestone)
        previous = completed[index - 1] if index > 0 else None
        previous_total = previous.metrics.total_area_m2 if previous and previous.metrics else None
        footprint_growth = metrics.total_area_m2 - previous_total if previous_total is not None else None
        archive_date, _note = _archive_date(project, milestone, settings, export_now)
        rows.append(
            {
                "Date d'archive": _date_cell(archive_date),
                "Source d'imagerie": _imagery_source(project, milestone),
                "Surface ajoutée (m²)": metrics.added_area_m2,
                "Emprise bâtie actuelle (m²)": metrics.total_area_m2,
                "Nombre d'ajouts détectés": metrics.additions_feature_count,
                "Nombre de blocs ajoutés": metrics.added_block_count,
                "Densité de croissance (%)": _percent(metrics.total_area_m2, metrics.growth_envelope_area_m2),
                "Ajouté / actuel (%)": _percent(metrics.added_area_m2, metrics.total_area_m2),
                "Emprise / enveloppe (%)": _percent(metrics.total_area_m2, metrics.growth_envelope_area_m2),
                "Surface blocs ajoutés (m²)": metrics.added_block_area_m2,
                "Surface cumulée (m²)": metrics.cumulative_block_area_m2,
                "Surface enveloppe (m²)": metrics.growth_envelope_area_m2,
                "Comparé avec": previous.release_identifier if previous else "",
                "Croissance de l'emprise (m²)": footprint_growth,
                "Croissance en %": _percent(footprint_growth, previous_total),
                "Statut": milestone.status,
            }
        )
    return rows


def _block_rows(project: TemporalProject, settings: Settings, export_now: datetime, crs: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for milestone in project.milestones:
        if milestone.status != "complete":
            continue
        archive_date, _note = _archive_date(project, milestone, settings, export_now)
        for index, feature in enumerate(_features(milestone.effective_building_blocks_geojson), start=1):
            geometry_payload = feature.get("geometry") if isinstance(feature, dict) else None
            properties = feature.get("properties") if isinstance(feature, dict) else None
            try:
                geometry = shape(geometry_payload).buffer(0) if isinstance(geometry_payload, dict) else GeometryCollection()
            except Exception:
                geometry = GeometryCollection()
            metric_area = _float(properties.get("area_m2")) if isinstance(properties, dict) else None
            if metric_area is None and not geometry.is_empty:
                try:
                    metric_area = float(reproject_geometry(geometry, "EPSG:4326", crs).area)
                except Exception:
                    metric_area = None
            centroid = geometry.centroid if not geometry.is_empty else None
            block_id = properties.get("block_id") if isinstance(properties, dict) else None
            rows.append(
                {
                    "Date d'archive": _date_cell(archive_date),
                    "Identifiant bloc": block_id or f"{milestone.release_identifier}-{index}",
                    "Surface (m²)": metric_area,
                    "Type géométrie": geometry.geom_type if not geometry.is_empty else "",
                    "Longitude centroïde": float(centroid.x) if centroid is not None else None,
                    "Latitude centroïde": float(centroid.y) if centroid is not None else None,
                }
            )
    return rows


def _append_rows(sheet, rows: list[dict[str, Any]], headers: list[str]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    _format_sheet(sheet)


def _set_workbook_view_zoom(workbook: Workbook, zoom_scale: int = 150) -> None:
    for sheet in workbook.worksheets:
        if sheet.sheet_state == "visible":
            sheet.sheet_view.zoomScale = zoom_scale
            sheet.sheet_view.zoomScaleNormal = zoom_scale


def _format_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))
            if isinstance(value, date):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(value, (int, float)):
                header = str(sheet.cell(row=1, column=cell.column).value or "")
                if "%" in header:
                    cell.number_format = "0.0%"
                elif "m²" in header or "Surface" in header:
                    cell.number_format = '#,##0'
                elif "Longitude" in header or "Latitude" in header:
                    cell.number_format = "0.000000"
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def build_temporal_results_workbook(project_id: str, settings: Settings | None = None) -> bytes:
    resolved_settings = _settings(settings)
    project = _load_project(project_id, resolved_settings)
    export_now = _export_now()
    crs = _metric_crs(project)
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Synthèse"
    summary_rows = [
        {"Champ": "Identifiant du projet", "Valeur": project.project_id},
        {"Champ": "Nom du projet", "Valeur": project.name},
        {"Champ": "Date d'export", "Valeur": export_now.date()},
        {"Champ": "Nombre de jalons", "Valeur": len(project.milestones)},
        {"Champ": "Système de coordonnées utilisé pour les surfaces", "Valeur": crs},
    ]
    _append_rows(summary, summary_rows, ["Champ", "Valeur"])

    milestones = workbook.create_sheet("Jalons")
    _append_rows(
        milestones,
        _milestone_rows(project, resolved_settings, export_now),
        [
            "Date d'archive",
            "Source d'imagerie",
            "Surface ajoutée (m²)",
            "Emprise bâtie actuelle (m²)",
            "Nombre d'ajouts détectés",
            "Nombre de blocs ajoutés",
            "Densité de croissance (%)",
            "Ajouté / actuel (%)",
            "Emprise / enveloppe (%)",
            "Surface blocs ajoutés (m²)",
            "Surface cumulée (m²)",
            "Surface enveloppe (m²)",
            "Comparé avec",
            "Croissance de l'emprise (m²)",
            "Croissance en %",
            "Statut",
        ],
    )

    blocks = workbook.create_sheet("Détails blocs")
    _append_rows(
        blocks,
        _block_rows(project, resolved_settings, export_now, crs),
        [
            "Date d'archive",
            "Identifiant bloc",
            "Surface (m²)",
            "Type géométrie",
            "Longitude centroïde",
            "Latitude centroïde",
        ],
    )

    _set_workbook_view_zoom(workbook, 150)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _kml_text(name: str, text: str | None = None) -> ET.Element:
    element = ET.Element(name)
    if text is not None:
        element.text = text
    return element


def _kml_child(parent: ET.Element, name: str, text: str | None = None) -> ET.Element:
    child = ET.SubElement(parent, name)
    if text is not None:
        child.text = text
    return child


def _gx_child(parent: ET.Element, name: str, text: str | None = None) -> ET.Element:
    child = ET.SubElement(parent, f"{{{GX_NS}}}{name}")
    if text is not None:
        child.text = text
    return child


def _coords(points: Any) -> str:
    return " ".join(f"{float(lon):.8f},{float(lat):.8f},0" for lon, lat, *_ in points)


def _append_polygon(parent: ET.Element, polygon: Polygon) -> None:
    polygon_el = _kml_child(parent, "Polygon")
    _kml_child(polygon_el, "tessellate", "1")
    outer = _kml_child(polygon_el, "outerBoundaryIs")
    outer_ring = _kml_child(outer, "LinearRing")
    _kml_child(outer_ring, "coordinates", _coords(polygon.exterior.coords))
    for interior in polygon.interiors:
        inner = _kml_child(polygon_el, "innerBoundaryIs")
        inner_ring = _kml_child(inner, "LinearRing")
        _kml_child(inner_ring, "coordinates", _coords(interior.coords))


def _append_geometry(parent: ET.Element, geometry: BaseGeometry) -> bool:
    if geometry.is_empty:
        return False
    if isinstance(geometry, Polygon):
        _append_polygon(parent, geometry)
        return True
    if isinstance(geometry, MultiPolygon):
        multi = _kml_child(parent, "MultiGeometry")
        wrote = False
        for polygon in geometry.geoms:
            if not polygon.is_empty:
                _append_polygon(multi, polygon)
                wrote = True
        return wrote
    repaired = geometry.buffer(0)
    if repaired.is_empty:
        return False
    return _append_geometry(parent, repaired)


def _project_view(project: TemporalProject, fallback_geometry: BaseGeometry) -> KmlView:
    geometry: BaseGeometry = GeometryCollection()
    if project.aoi_geojson:
        try:
            geometry = shape(project.aoi_geojson).buffer(0)
        except Exception:
            geometry = GeometryCollection()
    if geometry.is_empty:
        geometry = fallback_geometry
    if geometry.is_empty:
        return KmlView(lon=0.0, lat=0.0)
    point = geometry.representative_point()
    return KmlView(lon=float(point.x), lat=float(point.y))


def _append_look_at(parent: ET.Element, view: KmlView, archive_date: str | None = None) -> ET.Element:
    look_at = _kml_child(parent, "LookAt")
    _kml_child(look_at, "longitude", f"{view.lon:.8f}")
    _kml_child(look_at, "latitude", f"{view.lat:.8f}")
    _kml_child(look_at, "altitude", "0")
    _kml_child(look_at, "heading", "0")
    _kml_child(look_at, "tilt", "0")
    _kml_child(look_at, "range", f"{view.range_m:.2f}")
    _kml_child(look_at, "altitudeMode", "relativeToGround")
    if archive_date:
        timestamp = _gx_child(look_at, "TimeStamp")
        _kml_child(timestamp, "when", archive_date)
    return look_at


def _append_timespan(parent: ET.Element, begin: str, end: str | None) -> None:
    timespan = _kml_child(parent, "TimeSpan")
    _kml_child(timespan, "begin", begin)
    if end:
        _kml_child(timespan, "end", end)


def _all_view_geometry(project: TemporalProject, milestones: list[KmlMilestone]) -> BaseGeometry:
    geometries: list[BaseGeometry] = [entry.geometry for entry in milestones if not entry.geometry.is_empty]
    if project.aoi_geojson:
        try:
            aoi = shape(project.aoi_geojson).buffer(0)
            if not aoi.is_empty:
                geometries.append(aoi)
        except Exception:
            pass
    if not geometries:
        return GeometryCollection()
    return unary_union(geometries).buffer(0)


def _kml_milestones(project: TemporalProject, settings: Settings, export_now: datetime) -> tuple[list[KmlMilestone], list[str]]:
    milestones: list[KmlMilestone] = []
    skipped: list[str] = []
    for source_index, milestone in enumerate(project.milestones):
        if milestone.status != "complete":
            continue
        archive_date_text, date_note = _archive_date(project, milestone, settings, export_now)
        archive_date = _parse_date(archive_date_text or milestone.release_date or milestone.release_identifier)
        if archive_date is None:
            skipped.append(milestone.release_identifier)
            continue
        milestones.append(
            KmlMilestone(
                source_index=source_index,
                milestone=milestone,
                archive_date=archive_date,
                archive_date_text=archive_date.isoformat(),
                date_note=date_note,
                geometry=_cumulative_buffer_geometry(project, source_index),
            )
        )
    return sorted(milestones, key=lambda item: (item.archive_date, item.source_index)), skipped


def _append_tour(document: ET.Element, milestones: list[KmlMilestone], view: KmlView) -> None:
    if not milestones:
        return
    tour = _gx_child(document, "Tour")
    _kml_child(tour, "name", "Chronological building growth")
    playlist = _gx_child(tour, "Playlist")
    for entry in milestones:
        fly_to = _gx_child(playlist, "FlyTo")
        _gx_child(fly_to, "duration", "2.5")
        _gx_child(fly_to, "flyToMode", "smooth")
        _append_look_at(fly_to, view, entry.archive_date_text)
        wait = _gx_child(playlist, "Wait")
        _gx_child(wait, "duration", "1.0")


def _description(project: TemporalProject, milestone: TemporalMilestone, date_note: str) -> str:
    metrics = _metrics(milestone)
    rows = [
        ("Source d'imagerie", _imagery_source(project, milestone)),
        ("Surface ajoutée", f"{metrics.added_area_m2:.0f} m²"),
        ("Emprise bâtie actuelle", f"{metrics.total_area_m2:.0f} m²"),
        ("Ajouts détectés", str(metrics.additions_feature_count)),
        ("Blocs ajoutés", str(metrics.added_block_count)),
        ("Densité de croissance", f"{(_percent(metrics.total_area_m2, metrics.growth_envelope_area_m2) or 0) * 100:.1f}%"),
        ("Ajouté / actuel", f"{(_percent(metrics.added_area_m2, metrics.total_area_m2) or 0) * 100:.1f}%"),
        ("Emprise / enveloppe", f"{(_percent(metrics.total_area_m2, metrics.growth_envelope_area_m2) or 0) * 100:.1f}%"),
    ]
    if date_note:
        rows.append(("Note", date_note))
    html_rows = "".join(f"<tr><th>{html.escape(label)}</th><td>{html.escape(value)}</td></tr>" for label, value in rows)
    return f"<table>{html_rows}</table>"


def _cumulative_buffer_geometry(project: TemporalProject, milestone_index: int) -> BaseGeometry:
    geometries = [
        _geometry_from_geojson(milestone.buffer_layers_geojson.get("10m") or milestone.buffer_layers_geojson.get("10"))
        for milestone in project.milestones[: milestone_index + 1]
    ]
    geometries = [geometry for geometry in geometries if not geometry.is_empty]
    if not geometries:
        return GeometryCollection()
    return unary_union(geometries).buffer(0)


def _building_name(milestone: TemporalMilestone, feature: dict[str, Any], index: int) -> str:
    properties = feature.get("properties") if isinstance(feature, dict) else None
    block_id = properties.get("block_id") if isinstance(properties, dict) else None
    return f"Building block {block_id or index} - {milestone.release_identifier}"


def _building_description(milestone: TemporalMilestone, feature: dict[str, Any], first_seen: str) -> str:
    properties = feature.get("properties") if isinstance(feature, dict) else None
    rows = [("Jalon première détection", milestone.release_identifier), ("Date première détection", first_seen)]
    if isinstance(properties, dict):
        for key in ("block_id", "source_building_count", "area_m2"):
            value = properties.get(key)
            if value is not None:
                rows.append((str(key), str(value)))
    html_rows = "".join(f"<tr><th>{html.escape(label)}</th><td>{html.escape(value)}</td></tr>" for label, value in rows)
    return f"<table>{html_rows}</table>"


def _append_building_placemarks(folder: ET.Element, entry: KmlMilestone, final_date: str, *, is_baseline: bool = False) -> None:
    features = _features(entry.milestone.effective_building_blocks_geojson)
    if is_baseline and not features:
        features = _features(entry.milestone.cumulative_union_geojson)
    end_date = None if entry.archive_date_text == final_date else final_date
    for feature_index, feature in enumerate(features, start=1):
        if not isinstance(feature, dict):
            continue
        geometry_payload = feature.get("geometry")
        if not isinstance(geometry_payload, dict):
            continue
        try:
            geometry = shape(geometry_payload).buffer(0)
        except Exception:
            continue
        if geometry.is_empty or geometry.geom_type not in {"Polygon", "MultiPolygon"}:
            continue
        placemark = _kml_child(folder, "Placemark")
        _kml_child(placemark, "name", _building_name(entry.milestone, feature, feature_index))
        _kml_child(placemark, "styleUrl", "#buffer-rouge-transparent")
        _kml_child(placemark, "description", _building_description(entry.milestone, feature, entry.archive_date_text))
        _append_timespan(placemark, entry.archive_date_text, end_date)
        _append_geometry(placemark, geometry)


def build_temporal_results_kml(project_id: str, settings: Settings | None = None) -> bytes:
    resolved_settings = _settings(settings)
    project = _load_project(project_id, resolved_settings)
    export_now = _export_now()

    ET.register_namespace("", KML_NS)
    ET.register_namespace("gx", GX_NS)
    kml = _kml_text(f"{{{KML_NS}}}kml")
    document = _kml_child(kml, "Document")
    _kml_child(document, "name", f"Projet: {project.name or project.project_id}")
    kml_milestones, skipped_milestones = _kml_milestones(project, resolved_settings, export_now)
    if skipped_milestones:
        _kml_child(
            document,
            "description",
            "Jalons exclus du tour chronologique faute de date valide: " + ", ".join(skipped_milestones),
        )
    style = _kml_child(document, "Style")
    style.set("id", "buffer-rouge-transparent")
    line_style = _kml_child(style, "LineStyle")
    _kml_child(line_style, "color", "cc0000ff")
    _kml_child(line_style, "width", "1.5")
    poly_style = _kml_child(style, "PolyStyle")
    _kml_child(poly_style, "color", "660000ff")
    _kml_child(poly_style, "fill", "1")
    _kml_child(poly_style, "outline", "1")

    view_geometry = _all_view_geometry(project, kml_milestones)
    project_view = _project_view(project, view_geometry)
    final_date = kml_milestones[-1].archive_date_text if kml_milestones else None
    _append_look_at(document, project_view)
    for sorted_index, entry in enumerate(kml_milestones):
        milestone = entry.milestone
        folder = _kml_child(document, "Folder")
        label = entry.archive_date_text
        _kml_child(folder, "name", f"Jalon {milestone.release_identifier} - {label}")
        _append_look_at(folder, project_view, entry.archive_date_text)
        geometry = entry.geometry
        if geometry.is_empty:
            _append_building_placemarks(folder, entry, final_date or entry.archive_date_text, is_baseline=sorted_index == 0)
            continue
        placemark = _kml_child(folder, "Placemark")
        _kml_child(placemark, "name", "Buffer cumulatif changement bâtiment 10 m")
        _kml_child(placemark, "styleUrl", "#buffer-rouge-transparent")
        _kml_child(placemark, "description", _description(project, milestone, entry.date_note))
        _append_timespan(placemark, entry.archive_date_text, None if entry.archive_date_text == final_date else final_date)
        _append_geometry(placemark, geometry)
        _append_building_placemarks(folder, entry, final_date or entry.archive_date_text, is_baseline=sorted_index == 0)

    _append_tour(document, kml_milestones, project_view)

    return ET.tostring(kml, encoding="utf-8", xml_declaration=True)


def _source_backend(project: TemporalProject) -> str | None:
    return project.execution_config.inference_backend if project.execution_config is not None else None


def _feature_area_m2(feature: dict[str, Any], crs: str) -> float | None:
    properties = feature.get("properties") if isinstance(feature, dict) else None
    if isinstance(properties, dict):
        for key in ("area_m2", "added_area_m2", "surface_m2"):
            value = _float(properties.get(key))
            if value is not None:
                return value
    geometry_payload = feature.get("geometry") if isinstance(feature, dict) else None
    if not isinstance(geometry_payload, dict):
        return None
    try:
        geometry = shape(geometry_payload).buffer(0)
    except Exception:
        return None
    if geometry.is_empty:
        return 0.0
    try:
        return float(reproject_geometry(geometry, "EPSG:4326", crs).area)
    except Exception:
        return None


def _result_layer_payloads(milestone: TemporalMilestone) -> list[tuple[str, dict[str, Any] | None]]:
    return [
        ("additions", milestone.additions_geojson),
        ("buffer_10m", milestone.buffer_layers_geojson.get("10m") or milestone.buffer_layers_geojson.get("10")),
        ("buffer_15m", milestone.buffer_layers_geojson.get("15m") or milestone.buffer_layers_geojson.get("15")),
        ("buffer_20m", milestone.buffer_layers_geojson.get("20m") or milestone.buffer_layers_geojson.get("20")),
    ]


def _temporal_result_features(project: TemporalProject, settings: Settings, export_now: datetime) -> list[dict[str, Any]]:
    crs = _metric_crs(project)
    backend = _source_backend(project)
    features: list[dict[str, Any]] = []
    for milestone in project.milestones:
        if milestone.status != "complete":
            continue
        archive_date, _note = _archive_date(project, milestone, settings, export_now)
        for layer_type, payload in _result_layer_payloads(milestone):
            for index, feature in enumerate(_features(payload), start=1):
                if not isinstance(feature, dict) or not isinstance(feature.get("geometry"), dict):
                    continue
                original_properties = feature.get("properties") if isinstance(feature.get("properties"), dict) else {}
                area_m2 = _feature_area_m2(feature, crs)
                properties = {
                    **original_properties,
                    "project_id": project.project_id,
                    "run_id": milestone.pair_request_hash,
                    "release_identifier": milestone.release_identifier,
                    "date": archive_date or milestone.release_date,
                    "layer_type": layer_type,
                    "layer_label": TEMPORAL_RESULTS_EXPORT_LABELS[layer_type],
                    "feature_index": index,
                    "area_m2": area_m2,
                    "source_backend": backend,
                }
                features.append({"type": "Feature", "properties": properties, "geometry": feature["geometry"]})
    return features


def build_temporal_results_geojson(project_id: str, settings: Settings | None = None) -> bytes:
    resolved_settings = _settings(settings)
    project = _load_project(project_id, resolved_settings)
    export_now = _export_now()
    payload = {
        "type": "FeatureCollection",
        "name": f"{project.project_id}_temporal_results",
        "export_metadata": _export_metadata(),
        "features": _temporal_result_features(project, resolved_settings, export_now),
    }
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")


def _clean_project_display_name(project: TemporalProject) -> str:
    name = (project.name or "").strip()
    if name and not name.lower().startswith("temporal-"):
        return name.title() if name.isupper() else name
    candidate = re.sub(r"^(temporal|project|qgis)[-_]+", "", project.project_id, flags=re.IGNORECASE)
    words: list[str] = []
    for token in re.split(r"[-_\s]+", candidate):
        if not token:
            continue
        if re.fullmatch(r"mp[a-z0-9]+", token, flags=re.IGNORECASE):
            break
        if re.fullmatch(r"[a-z0-9]{6,}", token, flags=re.IGNORECASE) and any(char.isdigit() for char in token):
            break
        words.append(token)
    return " ".join(word.capitalize() for word in words) or project.project_id


def _topojson_result_layer_payloads(milestone: TemporalMilestone) -> list[tuple[str, dict[str, Any] | None]]:
    return [
        ("additions", milestone.additions_geojson),
        ("buffer_10m", milestone.buffer_layers_geojson.get("10m") or milestone.buffer_layers_geojson.get("10")),
        ("buffer_15m", milestone.buffer_layers_geojson.get("15m") or milestone.buffer_layers_geojson.get("15")),
        ("buffer_20m", milestone.buffer_layers_geojson.get("20m") or milestone.buffer_layers_geojson.get("20")),
    ]


def _milestone_archive_dates(project: TemporalProject, settings: Settings, export_now: datetime) -> dict[str, str]:
    archive_dates: dict[str, str] = {}
    for milestone in project.milestones:
        archive_date, _note = _archive_date(project, milestone, settings, export_now)
        fallback = _date_string(milestone.release_date) or _date_string(milestone.release_identifier)
        if archive_date or fallback:
            archive_dates[milestone.release_identifier] = archive_date or fallback or ""
    return archive_dates


def _topojson_period(
    layer_type: str,
    milestone_index: int,
    archive_dates: dict[str, str],
    project: TemporalProject,
    current_year: int,
) -> str:
    milestone_years = [
        int(date_text[:4])
        for milestone in project.milestones
        if (date_text := archive_dates.get(milestone.release_identifier))
    ]
    baseline_year = milestone_years[0] if milestone_years else current_year
    if layer_type == "cumulative_growth":
        return f"{baseline_year}-{current_year}"
    previous_year = baseline_year
    for previous in reversed(project.milestones[:milestone_index]):
        previous_date = archive_dates.get(previous.release_identifier)
        if previous_date:
            previous_year = int(previous_date[:4])
            break
    return f"{previous_year}-{current_year}"


def _topojson_clean_features(project: TemporalProject, settings: Settings, export_now: datetime) -> list[dict[str, Any]]:
    crs = _metric_crs(project)
    project_name = _clean_project_display_name(project)
    archive_dates = _milestone_archive_dates(project, settings, export_now)
    sequence_by_layer_year: dict[tuple[int, str], int] = {}
    features: list[dict[str, Any]] = []
    raw_feature_count = 0
    filtered_feature_count = 0

    for milestone_index, milestone in enumerate(project.milestones):
        if milestone.status != "complete":
            continue
        date_text = archive_dates.get(milestone.release_identifier)
        if not date_text:
            continue
        year = int(date_text[:4])
        for layer_type, payload in _topojson_result_layer_payloads(milestone):
            layer_features = _features(payload)
            raw_feature_count += len(layer_features)
            if layer_type not in TOPOJSON_ALLOWED_LAYERS:
                continue
            for feature in layer_features:
                if not isinstance(feature, dict) or not isinstance(feature.get("geometry"), dict):
                    continue
                filtered_feature_count += 1
                sequence_key = (year, layer_type)
                sequence_by_layer_year[sequence_key] = sequence_by_layer_year.get(sequence_key, 0) + 1
                area_m2 = _feature_area_m2(feature, crs)
                rounded_area_m2 = round(float(area_m2 or 0.0), 2)
                slug = TOPOJSON_LAYER_ID_SLUGS[layer_type]
                properties = {
                    "id": f"{year}-{slug}-{sequence_by_layer_year[sequence_key]:06d}",
                    "project": project_name,
                    "date": date_text,
                    "year": year,
                    "period": _topojson_period(layer_type, milestone_index, archive_dates, project, year),
                    "layer": layer_type,
                    "area_m2": rounded_area_m2,
                    "area_ha": round(rounded_area_m2 / 10000, 4),
                }
                features.append({"type": "Feature", "properties": properties, "geometry": feature["geometry"]})

    logger.info("TOPOJSON_EXPORT_FEATURES_COLLECTED count=%s", raw_feature_count)
    logger.info(
        "TOPOJSON_EXPORT_FILTERED layers=%s count=%s",
        ",".join(TOPOJSON_ALLOWED_LAYERS),
        filtered_feature_count,
    )
    logger.info("TOPOJSON_EXPORT_PROPERTIES_NORMALIZED allowedKeys=%s", ",".join(TOPOJSON_PROPERTY_KEYS))
    return features


def _topojson_bbox_from_project(project: TemporalProject) -> list[float]:
    if project.aoi_geojson:
        try:
            bounds = shape(project.aoi_geojson).bounds
            if len(bounds) == 4:
                return [round(float(value), 6) for value in bounds]
        except Exception:
            pass
    return [0.0, 0.0, 0.0, 0.0]


def _topojson_bbox(features: list[dict[str, Any]], project: TemporalProject) -> list[float]:
    geometries: list[BaseGeometry] = []
    for feature in features:
        geometry_payload = feature.get("geometry")
        if not isinstance(geometry_payload, dict):
            continue
        try:
            geometry = shape(geometry_payload)
        except Exception:
            continue
        if not geometry.is_empty:
            geometries.append(geometry)
    if not geometries:
        return _topojson_bbox_from_project(project)
    return [round(float(value), 6) for value in unary_union(geometries).bounds]


def _topojson_transform(bbox: list[float], quantization: int) -> dict[str, list[float]]:
    min_lon, min_lat, max_lon, max_lat = bbox
    denominator = max(quantization - 1, 1)
    lon_scale = (max_lon - min_lon) / denominator if max_lon > min_lon else 1.0 / denominator
    lat_scale = (max_lat - min_lat) / denominator if max_lat > min_lat else 1.0 / denominator
    return {
        "scale": [lon_scale, lat_scale],
        "translate": [min_lon, min_lat],
    }


def _quantize_point(point: Any, transform: dict[str, list[float]], quantization: int) -> list[int] | None:
    if not isinstance(point, (list, tuple)) or len(point) < 2:
        return None
    try:
        lon = float(point[0])
        lat = float(point[1])
    except (TypeError, ValueError):
        return None
    lon_scale, lat_scale = transform["scale"]
    min_lon, min_lat = transform["translate"]
    qx = 0 if lon_scale == 0 else int(round((lon - min_lon) / lon_scale))
    qy = 0 if lat_scale == 0 else int(round((lat - min_lat) / lat_scale))
    return [max(0, min(quantization - 1, qx)), max(0, min(quantization - 1, qy))]


def _delta_encode_ring(ring: Any, transform: dict[str, list[float]], quantization: int) -> list[list[int]] | None:
    if not isinstance(ring, list) or len(ring) < 4:
        return None
    quantized: list[list[int]] = []
    for point in ring:
        quantized_point = _quantize_point(point, transform, quantization)
        if quantized_point is None:
            continue
        if not quantized or quantized[-1] != quantized_point:
            quantized.append(quantized_point)
    if len(quantized) < 3:
        return None
    if quantized[0] != quantized[-1]:
        quantized.append(quantized[0])
    if len(quantized) < 4:
        return None
    encoded: list[list[int]] = []
    previous = [0, 0]
    for point in quantized:
        encoded.append([point[0] - previous[0], point[1] - previous[1]])
        previous = point
    return encoded


def _topojson_geometry(
    geometry_payload: dict[str, Any],
    arcs: list[Any],
    transform: dict[str, list[float]],
    quantization: int,
) -> dict[str, Any] | None:
    geometry_type = geometry_payload.get("type")
    coordinates = geometry_payload.get("coordinates")
    if geometry_type == "Polygon" and isinstance(coordinates, list):
        polygon_arcs: list[list[int]] = []
        for ring in coordinates:
            encoded_ring = _delta_encode_ring(ring, transform, quantization)
            if encoded_ring is None:
                continue
            arcs.append(encoded_ring)
            polygon_arcs.append([len(arcs) - 1])
        return {"type": "Polygon", "arcs": polygon_arcs} if polygon_arcs else None
    if geometry_type == "MultiPolygon" and isinstance(coordinates, list):
        multipolygon_arcs: list[list[list[int]]] = []
        for polygon in coordinates:
            if not isinstance(polygon, list):
                continue
            polygon_arcs = []
            for ring in polygon:
                encoded_ring = _delta_encode_ring(ring, transform, quantization)
                if encoded_ring is None:
                    continue
                arcs.append(encoded_ring)
                polygon_arcs.append([len(arcs) - 1])
            if polygon_arcs:
                multipolygon_arcs.append(polygon_arcs)
        return {"type": "MultiPolygon", "arcs": multipolygon_arcs} if multipolygon_arcs else None
    return None


def _validate_topojson_payload(payload: dict[str, Any]) -> None:
    if payload.get("type") != "Topology":
        raise ValueError("invalid_topojson_type")
    if not isinstance(payload.get("bbox"), list) or len(payload["bbox"]) != 4:
        raise ValueError("missing_topojson_bbox")
    if not isinstance(payload.get("transform"), dict):
        raise ValueError("missing_topojson_transform")
    objects = payload.get("objects")
    if not isinstance(objects, dict) or not isinstance(objects.get("results"), dict):
        raise ValueError("missing_topojson_results_object")
    results = objects["results"]
    if results.get("type") != "GeometryCollection":
        raise ValueError("invalid_topojson_results_type")
    if not isinstance(payload.get("arcs"), list):
        raise ValueError("missing_topojson_arcs")
    allowed_keys = set(TOPOJSON_PROPERTY_KEYS)
    for geometry in results.get("geometries", []):
        properties = geometry.get("properties")
        if not isinstance(properties, dict):
            raise ValueError("missing_topojson_properties")
        if set(properties) != allowed_keys:
            raise ValueError(f"invalid_topojson_property_keys:{sorted(set(properties) - allowed_keys)}")
        if set(properties) & TOPOJSON_REMOVED_PROPERTY_KEYS:
            raise ValueError("topojson_internal_properties_present")
        if properties["layer"] not in TOPOJSON_ALLOWED_LAYERS:
            raise ValueError("invalid_topojson_layer")
        if not TOPOJSON_ID_PATTERN.fullmatch(str(properties["id"])):
            raise ValueError("invalid_topojson_id")
        expected_ha = round(float(properties["area_m2"]) / 10000, 4)
        if abs(float(properties["area_ha"]) - expected_ha) > 0.0001:
            raise ValueError("invalid_topojson_area_ha")


def build_temporal_results_topojson(project_id: str, settings: Settings | None = None) -> bytes:
    logger.info("TOPOJSON_EXPORT_START projectId=%s", project_id)
    resolved_settings = _settings(settings)
    project = _load_project(project_id, resolved_settings)
    export_now = _export_now()
    features = _topojson_clean_features(project, resolved_settings, export_now)
    bbox = _topojson_bbox(features, project)
    transform = _topojson_transform(bbox, TOPOJSON_DEFAULT_QUANTIZATION)
    logger.info("TOPOJSON_EXPORT_QUANTIZED quantization=%s", TOPOJSON_DEFAULT_QUANTIZATION)
    logger.info("TOPOJSON_EXPORT_BBOX bbox=%s", bbox)
    arcs: list[Any] = []
    geometries: list[dict[str, Any]] = []
    for feature in features:
        geometry_payload = feature.get("geometry")
        if not isinstance(geometry_payload, dict):
            continue
        geometry = _topojson_geometry(geometry_payload, arcs, transform, TOPOJSON_DEFAULT_QUANTIZATION)
        if geometry is None:
            continue
        geometry["properties"] = feature.get("properties") or {}
        geometries.append(geometry)
    payload = {
        "type": "Topology",
        "bbox": bbox,
        "transform": transform,
        "objects": {"results": {"type": "GeometryCollection", "geometries": geometries}},
        "arcs": arcs,
    }
    try:
        _validate_topojson_payload(payload)
    except ValueError as exc:
        logger.error("TOPOJSON_EXPORT_VALIDATION_FAILED projectId=%s reason=%s", project_id, exc)
        raise
    result = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    logger.info(
        "TOPOJSON_EXPORT_DONE projectId=%s sizeBytes=%s features=%s arcs=%s hasTransform=%s hasBbox=%s",
        project_id,
        len(result),
        len(geometries),
        len(arcs),
        bool(payload.get("transform")),
        bool(payload.get("bbox")),
    )
    return result


def build_temporal_results_json(project_id: str, settings: Settings | None = None) -> bytes:
    resolved_settings = _settings(settings)
    project = _load_project(project_id, resolved_settings)
    export_now = _export_now()
    milestones = []
    for milestone in project.milestones:
        metrics = milestone.metrics.model_dump(mode="json") if milestone.metrics is not None else {}
        archive_date, date_note = _archive_date(project, milestone, resolved_settings, export_now)
        layer_counts = {
            layer_type: len(_features(payload))
            for layer_type, payload in _result_layer_payloads(milestone)
        }
        milestones.append(
            {
                "release_identifier": milestone.release_identifier,
                "date": archive_date or milestone.release_date,
                "date_note": date_note,
                "status": milestone.status,
                "source_mode": milestone.source_mode,
                "run_id": milestone.pair_request_hash,
                "metrics": metrics,
                "layer_feature_counts": layer_counts,
                "artifacts": [artifact.model_dump(mode="json") for artifact in milestone.artifacts],
            }
        )
    payload = {
        "project": {
            "project_id": project.project_id,
            "name": project.name,
            "semantics": project.semantics,
            "created_at": project.created_at,
            "updated_at": project.updated_at,
        },
        "run": {
            "source_backend": _source_backend(project),
            "exported_at": export_now.isoformat().replace("+00:00", "Z"),
            "export_perimeter": _export_metadata(),
        },
        "milestones": milestones,
        "artifacts": {
            "download_bundle_path": project.download_bundle_path,
        },
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def build_temporal_results_tsv(project_id: str, settings: Settings | None = None) -> bytes:
    resolved_settings = _settings(settings)
    project = _load_project(project_id, resolved_settings)
    export_now = _export_now()
    crs = _metric_crs(project)
    rows: list[dict[str, Any]] = []
    for milestone in project.milestones:
        if milestone.status != "complete":
            continue
        archive_date, _note = _archive_date(project, milestone, resolved_settings, export_now)
        for layer_type, payload in _result_layer_payloads(milestone):
            if len(_features(payload)) == 0:
                continue
            geometry = _geometry_from_geojson(payload)
            centroid = geometry.centroid if not geometry.is_empty else None
            rows.append(
                {
                    "project_id": project.project_id,
                    "project_name": project.name,
                    "date": _date_string(archive_date or milestone.release_date) or "",
                    "layer_type": layer_type,
                    "area_m2": _powerbi_tsv_area_m2(payload, crs),
                    "centroid_lon": float(centroid.x) if centroid is not None else "",
                    "centroid_lat": float(centroid.y) if centroid is not None else "",
                }
            )
    rows.sort(
        key=lambda row: (
            row["date"] or "9999-99-99",
            POWERBI_TSV_LAYER_ORDER.get(str(row["layer_type"]), len(POWERBI_TSV_LAYER_ORDER)),
            str(row["layer_type"]),
        )
    )
    stream = BytesIO()
    writer = csv.DictWriter(TextIOBytesWriter(stream), fieldnames=POWERBI_TSV_COLUMNS, delimiter="\t", lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return stream.getvalue()


class TextIOBytesWriter:
    def __init__(self, stream: BytesIO) -> None:
        self._stream = stream

    def write(self, value: str) -> int:
        data = value.encode("utf-8")
        self._stream.write(data)
        return len(value)


def _filesystem_safe_label(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^A-Za-z0-9]+", "_", normalized).strip("_")
    return normalized or "resultats"


def _milestone_quarter_label(milestone: TemporalMilestone) -> str:
    parsed = _parse_date(milestone.release_date)
    if parsed is not None:
        return f"{parsed.year} Q{((parsed.month - 1) // 3) + 1}"
    match = re.search(r"(20\d{2})", milestone.release_identifier)
    return f"{match.group(1)} Q1" if match else _filesystem_safe_label(milestone.release_identifier).replace("_", " ")


def _feature_collection_with_export_metadata(
    payload: dict[str, Any],
    *,
    release_identifier: str,
    period_label: str,
) -> dict[str, Any]:
    features: list[dict[str, Any]] = []
    for feature in _features(payload):
        if not isinstance(feature, dict):
            continue
        copied = dict(feature)
        properties = dict(feature.get("properties")) if isinstance(feature.get("properties"), dict) else {}
        properties["release_id"] = release_identifier
        properties["period"] = period_label
        copied["properties"] = properties
        features.append(copied)
    return {"type": "FeatureCollection", "features": features}


def _dissolved_feature_collection_with_export_metadata(
    payload: dict[str, Any],
    *,
    release_identifier: str,
    period_label: str,
) -> dict[str, Any]:
    geometries: list[BaseGeometry] = []
    area_m2 = 0.0
    score_values: list[float] = []
    for feature in _features(payload):
        if not isinstance(feature, dict):
            continue
        geometry_payload = feature.get("geometry")
        if not isinstance(geometry_payload, dict):
            continue
        try:
            geometry = _polygonal_geometry(shape(geometry_payload))
        except Exception:
            continue
        if geometry.is_empty:
            continue
        geometries.append(geometry)
        properties = feature.get("properties") if isinstance(feature.get("properties"), dict) else {}
        area_value = _float(properties.get("area_m2"))
        if area_value is not None:
            area_m2 += area_value
        score_value = _float(properties.get("score"))
        if score_value is not None:
            score_values.append(score_value)
    if not geometries:
        return {"type": "FeatureCollection", "features": []}
    dissolved = _polygonal_geometry(unary_union(geometries))
    if dissolved.is_empty:
        return {"type": "FeatureCollection", "features": []}
    properties: dict[str, Any] = {
        "release_id": release_identifier,
        "period": period_label,
        "source_feature_count": len(geometries),
    }
    if area_m2 > 0:
        properties["area_m2"] = area_m2
    if score_values:
        properties["score"] = sum(score_values) / len(score_values)
    return {
        "type": "FeatureCollection",
        "features": [{
            "type": "Feature",
            "properties": properties,
            "geometry": mapping(dissolved),
        }],
    }


def _merged_feature_collection(
    milestones: list[TemporalMilestone],
    *,
    payload_getter,
    dissolve_each_milestone: bool = False,
) -> dict[str, Any]:
    features: list[dict[str, Any]] = []
    for milestone in milestones:
        payload = payload_getter(milestone)
        if not _features(payload):
            continue
        period_label = _milestone_quarter_label(milestone)
        builder = _dissolved_feature_collection_with_export_metadata if dissolve_each_milestone else _feature_collection_with_export_metadata
        enriched = builder(
            payload,
            release_identifier=milestone.release_identifier,
            period_label=period_label,
        )
        features.extend(enriched["features"])
    return {"type": "FeatureCollection", "features": features}


def _temporal_shapefile_export_layers(project: TemporalProject) -> list[TemporalShapefileLayer]:
    completed = [milestone for milestone in project.milestones if milestone.status == "complete"]
    if not completed:
        return []
    baseline = completed[0]
    non_baseline = completed[1:]
    newest_first = list(reversed(non_baseline))
    start_label = _milestone_quarter_label(baseline)
    end_label = _milestone_quarter_label(completed[-1])
    range_file_label = f"{_filesystem_safe_label(start_label)}_{_filesystem_safe_label(end_label)}"
    range_display_label = f"{start_label} \u2192 {end_label}"
    layers: list[TemporalShapefileLayer] = []

    cumulative_payload = _merged_feature_collection(
        non_baseline,
        payload_getter=lambda milestone: milestone.additions_geojson,
    )
    if _features(cumulative_payload):
        layers.append(
            TemporalShapefileLayer(
                group_key="tous_les_nouveaux_batiments",
                group_label="Tous les nouveaux bâtiments",
                filename=f"tous_les_nouveaux_batiments_{range_file_label}",
                display_name=f"Tous les nouveaux bâtiments {range_display_label}",
                artifact_key="additions",
                release_identifier=f"{baseline.release_identifier}:{completed[-1].release_identifier}",
                feature_collection=cumulative_payload,
                is_global=True,
            )
        )

    for milestone in newest_first:
        period = _milestone_quarter_label(milestone)
        if _features(milestone.additions_geojson):
            layers.append(
                TemporalShapefileLayer(
                    group_key="batiments_ajoutes_par_date",
                    group_label="Bâtiments ajoutés par date",
                    filename=f"batiments_ajoutes_{_filesystem_safe_label(period)}",
                    display_name=f"Bâtiments ajoutés {period}",
                    artifact_key="additions",
                    release_identifier=milestone.release_identifier,
                    feature_collection=_feature_collection_with_export_metadata(
                        milestone.additions_geojson or {"type": "FeatureCollection", "features": []},
                        release_identifier=milestone.release_identifier,
                        period_label=period,
                    ),
                )
            )

    for distance in ("10m", "15m", "20m"):
        group_key = f"buffer_{distance}"
        group_label = f"Buffer {distance}"
        for milestone in newest_first:
            payload = milestone.buffer_layers_geojson.get(distance) or milestone.buffer_layers_geojson.get(distance.removesuffix("m"))
            if not _features(payload):
                continue
            period = _milestone_quarter_label(milestone)
            layers.append(
                TemporalShapefileLayer(
                    group_key=group_key,
                    group_label=group_label,
                    filename=f"buffer_{distance}_{_filesystem_safe_label(period)}",
                    display_name=f"Buffer {distance} {period}",
                    artifact_key=f"building_change_buffer_{distance}",
                    release_identifier=milestone.release_identifier,
                    feature_collection=_dissolved_feature_collection_with_export_metadata(
                        payload or {"type": "FeatureCollection", "features": []},
                        release_identifier=milestone.release_identifier,
                        period_label=period,
                    ),
                )
            )
        merged = _merged_feature_collection(
            non_baseline,
            payload_getter=lambda milestone, key=distance: milestone.buffer_layers_geojson.get(key)
            or milestone.buffer_layers_geojson.get(key.removesuffix("m")),
            dissolve_each_milestone=True,
        )
        if _features(merged):
            layers.append(
                TemporalShapefileLayer(
                    group_key=group_key,
                    group_label=group_label,
                    filename=f"buffer_{distance}_{range_file_label}",
                    display_name=f"Buffer {distance} {range_display_label}",
                    artifact_key=f"building_change_buffer_{distance}",
                    release_identifier=f"{baseline.release_identifier}:{completed[-1].release_identifier}",
                    feature_collection=merged,
                    is_global=True,
                )
            )
    context = _EXPORT_CONTEXT.get()
    if context and context.get("scope_type") == "custom_geometry":
        layers.append(
            TemporalShapefileLayer(
                group_key="zone_export",
                group_label="Zone d’export",
                filename="zone_export",
                display_name="Zone d’export",
                artifact_key="export_perimeter",
                release_identifier="export_perimeter",
                feature_collection={
                    "type": "FeatureCollection",
                    "features": [{
                        "type": "Feature",
                        "properties": {"release_id": "export_perimeter", "period": context["source"]},
                        "geometry": mapping(context["geometry"]),
                    }],
                },
                is_global=True,
            )
        )
    return layers


def _temporal_milestone_color_map(project: TemporalProject) -> dict[str, str]:
    completed = sorted(
        [milestone for milestone in project.milestones if milestone.status == "complete"],
        key=lambda milestone: (_parse_date(milestone.release_date) or date.min, milestone.release_identifier),
    )
    return {
        milestone.release_identifier: TEMPORAL_WEB_MILESTONE_COLORS[index - 1]
        for index, milestone in enumerate(completed)
        if index > 0
    }


def _hex_rgba(color: str, alpha: int) -> str:
    red, green, blue = (int(color[index:index + 2], 16) for index in (1, 3, 5))
    return f"{red},{green},{blue},{alpha}"


def _qgis_layer_style(project: TemporalProject, layer: TemporalShapefileLayer) -> tuple[str, str, str]:
    if layer.group_key == "zone_export":
        return _hex_rgba("#F59E0B", 36), _hex_rgba("#D97706", 255), "0.5"
    if layer.is_global and layer.group_key == "tous_les_nouveaux_batiments":
        return _hex_rgba("#00C8C8", 128), _hex_rgba("#0E7490", 245), "0.4"
    if layer.is_global and not layer.group_key.startswith("buffer_"):
        return _hex_rgba("#F59E0B", 112), _hex_rgba("#C2410C", 245), "0.35"
    color = _temporal_milestone_color_map(project).get(layer.release_identifier, "#64748B")
    alpha = 72 if layer.group_key.startswith("buffer_") else 150
    return _hex_rgba(color, alpha), _hex_rgba(color, 245), "0.35"


def _create_shapefile_spatial_index(shp_path: Path) -> bool:
    try:
        from osgeo import ogr
    except Exception as exc:
        logger.info(
            "EXPORT_SHAPEFILE_SPATIAL_INDEX_SKIPPED path=%s reason=osgeo_unavailable error=%s",
            shp_path,
            exc.__class__.__name__,
        )
        return False
    datasource = ogr.Open(str(shp_path), update=1)
    if datasource is None:
        logger.info("EXPORT_SHAPEFILE_SPATIAL_INDEX_SKIPPED path=%s reason=open_failed", shp_path)
        return False
    try:
        layer = datasource.GetLayer(0)
        layer_name = layer.GetName() if layer is not None else shp_path.stem
        datasource.ExecuteSQL(f"CREATE SPATIAL INDEX ON {layer_name}")
    except Exception as exc:
        logger.info(
            "EXPORT_SHAPEFILE_SPATIAL_INDEX_SKIPPED path=%s reason=create_failed error=%s",
            shp_path,
            exc.__class__.__name__,
        )
        return False
    finally:
        datasource = None
    created = shp_path.with_suffix(".qix").is_file()
    if created:
        logger.info("EXPORT_SHAPEFILE_SPATIAL_INDEX_CREATED path=%s index=%s", shp_path, shp_path.with_suffix(".qix"))
    else:
        logger.info("EXPORT_SHAPEFILE_SPATIAL_INDEX_SKIPPED path=%s reason=index_not_created", shp_path)
    return created


def _write_temporal_shapefile(layer: TemporalShapefileLayer, output_dir: Path) -> Path | None:
    records: list[dict[str, Any]] = []
    geometries: list[BaseGeometry] = []
    logger.info(
        "EXPORT_SHAPEFILE_STREAM_WRITE_START group=%s release=%s artifact_key=%s inputFeatures=%s",
        layer.group_key,
        layer.release_identifier,
        layer.artifact_key,
        len(_features(layer.feature_collection)),
    )
    for feature in _features(layer.feature_collection):
        geometry_payload = feature.get("geometry") if isinstance(feature, dict) else None
        if not isinstance(geometry_payload, dict):
            continue
        try:
            geometry = shape(geometry_payload).buffer(0)
        except Exception:
            continue
        if geometry.is_empty:
            continue
        properties = feature.get("properties") if isinstance(feature.get("properties"), dict) else {}
        records.append(
            {
                "release_id": str(properties.get("release_id") or layer.release_identifier)[:254],
                "period": str(properties.get("period") or "")[:254],
                "area_m2": _float(properties.get("area_m2")),
                "score": _float(properties.get("score")),
            }
        )
        geometries.append(geometry)
    if not records:
        return None
    output_dir.mkdir(parents=True, exist_ok=True)
    shp_path = output_dir / f"{layer.filename}.shp"
    gdf = gpd.GeoDataFrame(records, geometry=geometries, crs="EPSG:4326")
    gdf.to_file(shp_path, driver="ESRI Shapefile", engine="pyogrio", encoding="UTF-8")
    cpg_path = shp_path.with_suffix(".cpg")
    if not cpg_path.exists():
        cpg_path.write_text("UTF-8", encoding="ascii")
    _create_shapefile_spatial_index(shp_path)
    logger.info(
        "EXPORT_SHAPEFILE_LAYER_WRITE_DONE path=%s feature_count=%s bytes=%s",
        shp_path,
        len(records),
        shp_path.stat().st_size,
    )
    return shp_path


def _invalid_layer_validation(
    path: Path,
    display_name: str,
    group_name: str,
    *,
    reason: str,
) -> ExportedLayerValidation:
    return ExportedLayerValidation(
        path=path,
        display_name=display_name,
        group_name=group_name,
        relative_path="",
        feature_count=0,
        crs_authid="",
        xmin=0.0,
        ymin=0.0,
        xmax=0.0,
        ymax=0.0,
        is_valid_for_qgis_project=False,
        reason=reason,
    )


def validate_exported_vector_layer(
    path: Path,
    display_name: str,
    group_name: str,
    root_dir: Path,
) -> ExportedLayerValidation:
    required = [path, path.with_suffix(".shx"), path.with_suffix(".dbf")]
    missing = [candidate.name for candidate in required if not candidate.is_file()]
    if missing:
        return _invalid_layer_validation(path, display_name, group_name, reason=f"missing sidecars: {missing}")
    try:
        relative_path = path.resolve().relative_to(root_dir.resolve()).as_posix()
    except ValueError:
        return _invalid_layer_validation(path, display_name, group_name, reason="datasource path escapes export root")
    try:
        gdf = gpd.read_file(path, engine="pyogrio")
    except Exception as exc:
        return _invalid_layer_validation(path, display_name, group_name, reason=f"unable to reopen shapefile: {exc}")
    if gdf.empty:
        return _invalid_layer_validation(path, display_name, group_name, reason="feature_count is zero")
    if gdf.geometry.name not in gdf.columns:
        return _invalid_layer_validation(path, display_name, group_name, reason="geometry column is missing")
    geometry = gdf.geometry
    if geometry.isna().all():
        return _invalid_layer_validation(path, display_name, group_name, reason="all geometries are null")
    non_null = geometry.dropna()
    if non_null.empty or non_null.is_empty.all():
        return _invalid_layer_validation(path, display_name, group_name, reason="all geometries are empty")
    if gdf.crs is None:
        gdf = gdf.set_crs("EPSG:4326", allow_override=True)
        gdf.to_file(path, driver="ESRI Shapefile", engine="pyogrio", encoding="UTF-8")
    crs_authid = gdf.crs.to_string() if gdf.crs is not None else ""
    if crs_authid.upper() not in {"EPSG:4326", "OGC:CRS84"}:
        return _invalid_layer_validation(path, display_name, group_name, reason=f"unsupported CRS: {crs_authid}")
    bounds = tuple(float(value) for value in gdf.total_bounds)
    if len(bounds) != 4 or not all(math.isfinite(value) for value in bounds):
        return _invalid_layer_validation(path, display_name, group_name, reason=f"non-finite bounds: {bounds}")
    xmin, ymin, xmax, ymax = bounds
    if xmin > xmax or ymin > ymax:
        return _invalid_layer_validation(path, display_name, group_name, reason=f"invalid bounds: {bounds}")
    return ExportedLayerValidation(
        path=path,
        display_name=display_name,
        group_name=group_name,
        relative_path=f"./{relative_path}",
        feature_count=len(gdf),
        crs_authid="EPSG:4326",
        xmin=xmin,
        ymin=ymin,
        xmax=xmax,
        ymax=ymax,
        is_valid_for_qgis_project=True,
    )


def _union_bounds(bounds: list[Bounds]) -> Bounds:
    if not bounds:
        raise ValueError("Cannot compute an extent from zero layers.")
    return (
        min(item[0] for item in bounds),
        min(item[1] for item in bounds),
        max(item[2] for item in bounds),
        max(item[3] for item in bounds),
    )


def _validate_bounds(bounds: Bounds) -> Bounds:
    xmin, ymin, xmax, ymax = (float(value) for value in bounds)
    if not all(math.isfinite(value) for value in (xmin, ymin, xmax, ymax)):
        raise ValueError(f"QGIS extent contains non-finite values: {bounds}")
    if xmin >= xmax or ymin >= ymax:
        raise ValueError(f"QGIS extent is empty or unordered: {bounds}")
    return xmin, ymin, xmax, ymax


def _pad_bounds(bounds: Bounds, ratio: float = QGIS_PROJECT_EXTENT_PADDING_RATIO) -> Bounds:
    xmin, ymin, xmax, ymax = _validate_bounds(bounds)
    width = max(xmax - xmin, 1.0)
    height = max(ymax - ymin, 1.0)
    x_padding = width * ratio
    y_padding = height * ratio
    return xmin - x_padding, ymin - y_padding, xmax + x_padding, ymax + y_padding


def _transform_qgis_extent(bounds: Bounds, source_crs: str, target_crs: str = QGIS_PROJECT_CRS) -> Bounds:
    bounds = _validate_bounds(bounds)
    if source_crs.upper() == target_crs.upper():
        return bounds
    transformed = reproject_geometry(
        box(*bounds),
        source_crs,
        target_crs,
    )
    return _validate_bounds(tuple(float(value) for value in transformed.bounds))


def _qgis_srs_xml(authid: str = QGIS_VECTOR_CRS, indent: str = "      ") -> list[str]:
    if authid == QGIS_PROJECT_CRS:
        return [
            f"{indent}<srs>",
            f"{indent}  <spatialrefsys nativeFormat=\"Wkt\">",
            f"{indent}    <wkt>PROJCRS[&quot;WGS 84 / Pseudo-Mercator&quot;,BASEGEOGCRS[&quot;WGS 84&quot;,ENSEMBLE[&quot;World Geodetic System 1984 ensemble&quot;,MEMBER[&quot;World Geodetic System 1984 (Transit)&quot;],MEMBER[&quot;World Geodetic System 1984 (G730)&quot;],MEMBER[&quot;World Geodetic System 1984 (G873)&quot;],MEMBER[&quot;World Geodetic System 1984 (G1150)&quot;],MEMBER[&quot;World Geodetic System 1984 (G1674)&quot;],MEMBER[&quot;World Geodetic System 1984 (G1762)&quot;],MEMBER[&quot;World Geodetic System 1984 (G2139)&quot;],MEMBER[&quot;World Geodetic System 1984 (G2296)&quot;],ELLIPSOID[&quot;WGS 84&quot;,6378137,298.257223563,LENGTHUNIT[&quot;metre&quot;,1]],ENSEMBLEACCURACY[2.0]],PRIMEM[&quot;Greenwich&quot;,0,ANGLEUNIT[&quot;degree&quot;,0.0174532925199433]],ID[&quot;EPSG&quot;,4326]],CONVERSION[&quot;Popular Visualisation Pseudo-Mercator&quot;,METHOD[&quot;Popular Visualisation Pseudo Mercator&quot;,ID[&quot;EPSG&quot;,1024]],PARAMETER[&quot;Latitude of natural origin&quot;,0,ANGLEUNIT[&quot;degree&quot;,0.0174532925199433],ID[&quot;EPSG&quot;,8801]],PARAMETER[&quot;Longitude of natural origin&quot;,0,ANGLEUNIT[&quot;degree&quot;,0.0174532925199433],ID[&quot;EPSG&quot;,8802]],PARAMETER[&quot;False easting&quot;,0,LENGTHUNIT[&quot;metre&quot;,1],ID[&quot;EPSG&quot;,8806]],PARAMETER[&quot;False northing&quot;,0,LENGTHUNIT[&quot;metre&quot;,1],ID[&quot;EPSG&quot;,8807]]],CS[Cartesian,2],AXIS[&quot;easting (X)&quot;,east,ORDER[1],LENGTHUNIT[&quot;metre&quot;,1]],AXIS[&quot;northing (Y)&quot;,north,ORDER[2],LENGTHUNIT[&quot;metre&quot;,1]],USAGE[SCOPE[&quot;Web mapping and visualisation.&quot;],AREA[&quot;World between 85.06°S and 85.06°N.&quot;],BBOX[-85.06,-180,85.06,180]],ID[&quot;EPSG&quot;,3857]]</wkt>",
            f"{indent}    <proj4>+proj=merc +a=6378137 +b=6378137 +lat_ts=0 +lon_0=0 +x_0=0 +y_0=0 +k=1 +units=m +nadgrids=@null +wktext +no_defs</proj4>",
            f"{indent}    <srsid>3857</srsid>",
            f"{indent}    <srid>3857</srid>",
            f"{indent}    <authid>EPSG:3857</authid>",
            f"{indent}    <description>WGS 84 / Pseudo-Mercator</description>",
            f"{indent}    <projectionacronym>merc</projectionacronym>",
            f"{indent}    <ellipsoidacronym>EPSG:7030</ellipsoidacronym>",
            f"{indent}    <geographicflag>false</geographicflag>",
            f"{indent}  </spatialrefsys>",
            f"{indent}</srs>",
        ]
    return [
        f"{indent}<srs>",
        f"{indent}  <spatialrefsys nativeFormat=\"Wkt\">",
        f"{indent}    <wkt>GEOGCRS[&quot;WGS 84&quot;,DATUM[&quot;World Geodetic System 1984&quot;,ELLIPSOID[&quot;WGS 84&quot;,6378137,298.257223563]],CS[ellipsoidal,2],AXIS[&quot;geodetic latitude (Lat)&quot;,north],AXIS[&quot;geodetic longitude (Lon)&quot;,east],ANGLEUNIT[&quot;degree&quot;,0.0174532925199433],ID[&quot;EPSG&quot;,4326]]</wkt>",
        f"{indent}    <proj4>+proj=longlat +datum=WGS84 +no_defs</proj4>",
        f"{indent}    <srsid>3452</srsid>",
        f"{indent}    <srid>4326</srid>",
        f"{indent}    <authid>EPSG:4326</authid>",
        f"{indent}    <description>WGS 84</description>",
        f"{indent}    <projectionacronym>longlat</projectionacronym>",
        f"{indent}    <ellipsoidacronym>EPSG:7030</ellipsoidacronym>",
        f"{indent}    <geographicflag>true</geographicflag>",
        f"{indent}  </spatialrefsys>",
        f"{indent}</srs>",
    ]


def _extent_xml(bounds: Bounds, indent: str = "      ") -> list[str]:
    xmin, ymin, xmax, ymax = bounds
    return [
        f"{indent}<extent>",
        f"{indent}  <xmin>{xmin:.15g}</xmin>",
        f"{indent}  <ymin>{ymin:.15g}</ymin>",
        f"{indent}  <xmax>{xmax:.15g}</xmax>",
        f"{indent}  <ymax>{ymax:.15g}</ymax>",
        f"{indent}</extent>",
    ]


def _qgis_project_extent(vector_extent_wgs84: Bounds) -> Bounds:
    return _pad_bounds(_transform_qgis_extent(vector_extent_wgs84, QGIS_VECTOR_CRS, QGIS_PROJECT_CRS))


def compute_qgis_project_initial_extent(
    project_id: str,
    validations: list[ExportedLayerValidation],
    rasters: list[TemporalQgisRaster],
    *,
    fallback_extent_wgs84: Bounds | None = None,
) -> QgisProjectInitialExtent:
    logger.info(
        "QGIS_PROJECT_EXTENT_COMPUTE_START projectId=%s vectorLayerCount=%s rasterLayerCount=%s projectCrs=%s",
        project_id,
        len(validations),
        len(rasters),
        QGIS_PROJECT_CRS,
    )
    project_crs_bounds: list[Bounds] = []
    for validation in validations:
        if not validation.is_valid_for_qgis_project or validation.feature_count <= 0:
            logger.info(
                "QGIS_PROJECT_LAYER_EXTENT_SKIPPED projectId=%s layer=%s reason=%s",
                project_id,
                validation.display_name,
                validation.reason or "invalid_or_empty",
            )
            continue
        layer_bounds = (validation.xmin, validation.ymin, validation.xmax, validation.ymax)
        logger.info(
            "QGIS_PROJECT_LAYER_EXTENT projectId=%s layer=%s source=vector crs=%s xmin=%s ymin=%s xmax=%s ymax=%s",
            project_id,
            validation.display_name,
            validation.crs_authid,
            *layer_bounds,
        )
        transformed = _transform_qgis_extent(layer_bounds, validation.crs_authid, QGIS_PROJECT_CRS)
        logger.info(
            "QGIS_PROJECT_EXTENT_TRANSFORMED projectId=%s layer=%s source=vector sourceCrs=%s targetCrs=%s xmin=%s ymin=%s xmax=%s ymax=%s",
            project_id,
            validation.display_name,
            validation.crs_authid,
            QGIS_PROJECT_CRS,
            *transformed,
        )
        project_crs_bounds.append(transformed)

    for raster in rasters:
        try:
            raster_bounds = _validate_bounds(raster.bounds_3857)
        except ValueError as exc:
            logger.info(
                "QGIS_PROJECT_LAYER_EXTENT_SKIPPED projectId=%s layer=%s source=raster reason=%s",
                project_id,
                raster.display_name,
                exc,
            )
            continue
        logger.info(
            "QGIS_PROJECT_LAYER_EXTENT projectId=%s layer=%s source=raster crs=%s xmin=%s ymin=%s xmax=%s ymax=%s",
            project_id,
            raster.display_name,
            QGIS_PROJECT_CRS,
            *raster_bounds,
        )
        project_crs_bounds.append(raster_bounds)

    if project_crs_bounds:
        combined = _union_bounds(project_crs_bounds)
        padded = _pad_bounds(combined)
        logger.info(
            "QGIS_PROJECT_INITIAL_EXTENT_SET projectId=%s crs=%s sourceCount=%s paddingRatio=%s xmin=%s ymin=%s xmax=%s ymax=%s",
            project_id,
            QGIS_PROJECT_CRS,
            len(project_crs_bounds),
            QGIS_PROJECT_EXTENT_PADDING_RATIO,
            *padded,
        )
        return QgisProjectInitialExtent(bounds_3857=padded, source_count=len(project_crs_bounds))

    if fallback_extent_wgs84 is None:
        raise ValueError("Cannot compute QGIS initial extent from empty layers and no fallback extent.")
    fallback = _qgis_project_extent(fallback_extent_wgs84)
    logger.warning(
        "QGIS_PROJECT_INITIAL_EXTENT_FALLBACK projectId=%s crs=%s reason=no_non_empty_export_layers xmin=%s ymin=%s xmax=%s ymax=%s",
        project_id,
        QGIS_PROJECT_CRS,
        *fallback,
    )
    return QgisProjectInitialExtent(bounds_3857=fallback, source_count=0, used_fallback=True)


def _default_visible_layer_names(layers: list[TemporalShapefileLayer]) -> set[str]:
    visible: set[str] = set()
    newest_additions = next((layer for layer in layers if layer.group_key == "batiments_ajoutes_par_date"), None)
    synthesis_additions = next((layer for layer in layers if layer.is_global and layer.group_key == "tous_les_nouveaux_batiments"), None)
    synthesis_buffer_10m = next((layer for layer in layers if layer.is_global and layer.group_key == "buffer_10m"), None)
    zone = next((layer for layer in layers if layer.group_key == "zone_export"), None)
    for layer in (newest_additions, synthesis_additions, synthesis_buffer_10m, zone):
        if layer is not None:
            visible.add(layer.display_name)
    return visible


def _reference_raster_source(project_dir: Path, milestone: TemporalMilestone) -> Path | None:
    candidates = []
    if milestone.reference_imagery:
        for value in (
            milestone.reference_imagery.canonical_cog_path,
            milestone.reference_imagery.cog_path,
            milestone.reference_imagery.image_path,
        ):
            if value:
                candidates.append(Path(value).expanduser())
    candidates.append(project_dir / "milestones" / milestone.release_identifier / "reference_imagery_cog.tif")
    return next((path.resolve() for path in candidates if path.is_file()), None)


def _copy_qgis_reference_rasters(
    project: TemporalProject,
    settings: Settings,
    export_root: Path,
) -> list[TemporalQgisRaster]:
    from rasterio import open as rasterio_open
    from rasterio.mask import mask as rasterio_mask
    from rasterio.transform import array_bounds
    from rasterio.warp import transform_bounds, transform_geom

    project_dir = settings.temporal_projects_dir / project.project_id
    context = _EXPORT_CONTEXT.get()
    custom_context = context if context and context.get("scope_type") == "custom_geometry" else None
    rasters: list[TemporalQgisRaster] = []
    for milestone in [item for item in project.milestones if item.status == "complete"]:
        source = _reference_raster_source(project_dir, milestone)
        if source is None:
            logger.warning(
                "EXPORT_RESULTS_QGIS_RASTER_MISSING projectId=%s release=%s expected=%s",
                project.project_id,
                milestone.release_identifier,
                project_dir / "milestones" / milestone.release_identifier / "reference_imagery_cog.tif",
            )
            continue
        relative_path = Path("rasters") / _filesystem_safe_label(milestone.release_identifier) / (
            "reference_imagery_export_zone.tif" if custom_context else "reference_imagery_cog.tif"
        )
        target = export_root / relative_path
        target.parent.mkdir(parents=True, exist_ok=True)
        with rasterio_open(source) as source_dataset:
            if source_dataset.crs is None:
                raise ValueError(f"Reference raster has no CRS: {source}")
            source_bounds = tuple(float(value) for value in source_dataset.bounds)
            if custom_context:
                logger.info(
                    "EXPORT_RESULTS_QGIS_RASTER_CLIP_START projectId=%s release=%s",
                    project.project_id,
                    milestone.release_identifier,
                )
                raster_geometry = transform_geom(QGIS_VECTOR_CRS, source_dataset.crs, mapping(custom_context["geometry"]))
                clipped, clipped_transform = rasterio_mask(source_dataset, [raster_geometry], crop=True)
                profile = source_dataset.profile.copy()
                profile.update(
                    driver="GTiff",
                    height=clipped.shape[1],
                    width=clipped.shape[2],
                    transform=clipped_transform,
                    compress="deflate",
                    tiled=True,
                )
                with rasterio_open(target, "w", **profile) as output_dataset:
                    output_dataset.write(clipped)
                output_bounds = tuple(
                    float(value)
                    for value in array_bounds(clipped.shape[1], clipped.shape[2], clipped_transform)
                )
                logger.info(
                    "EXPORT_RESULTS_QGIS_RASTER_CLIPPED projectId=%s release=%s source=%s output=%s srcBounds=%s outBounds=%s bytes=%s",
                    project.project_id,
                    milestone.release_identifier,
                    source,
                    relative_path,
                    source_bounds,
                    output_bounds,
                    target.stat().st_size,
                )
            else:
                shutil.copy2(source, target)
        with rasterio_open(target) as dataset:
            if dataset.crs is None:
                raise ValueError(f"Reference raster has no CRS: {source}")
            bounds = tuple(float(value) for value in dataset.bounds)
            if dataset.crs.to_string().upper() != QGIS_PROJECT_CRS:
                bounds = transform_bounds(dataset.crs, QGIS_PROJECT_CRS, *bounds)
        raster = TemporalQgisRaster(
            release_identifier=milestone.release_identifier,
            period_label=_milestone_quarter_label(milestone),
            source_path=target,
            relative_path=f"./{relative_path.as_posix()}",
            layer_id=f"reference_imagery_{_filesystem_safe_label(milestone.release_identifier)}",
            display_name=f"Imagerie de référence – {_milestone_quarter_label(milestone)}",
            bounds_3857=bounds,
        )
        rasters.append(raster)
        logger.info(
            "EXPORT_RESULTS_QGIS_RASTER_INCLUDED projectId=%s release=%s relativePath=%s bytes=%s",
            project.project_id,
            milestone.release_identifier,
            relative_path,
            target.stat().st_size,
        )
    return rasters


def _categorized_renderer_xml(project: TemporalProject, indent: str = "      ", *, fill_alpha: int = 150) -> list[str]:
    colors = _temporal_milestone_color_map(project)
    categories: list[str] = []
    symbols: list[str] = []
    for index, milestone in enumerate([item for item in project.milestones[1:] if item.status == "complete"]):
        color = colors.get(milestone.release_identifier, "#64748B")
        categories.append(
            f'{indent}    <category value="{html.escape(milestone.release_identifier)}" symbol="{index}" label="{html.escape(_milestone_quarter_label(milestone))}" render="true"/>'
        )
        symbols.append(
            f'{indent}    <symbol type="fill" name="{index}" alpha="1" clip_to_extent="1" force_rhr="0"><layer class="SimpleFill" enabled="1" pass="0" locked="0"><Option type="Map"><Option name="color" value="{_hex_rgba(color, fill_alpha)}" type="QString"/><Option name="outline_color" value="{_hex_rgba(color, 245)}" type="QString"/><Option name="outline_width" value="0.35" type="QString"/><Option name="outline_width_unit" value="Pixel" type="QString"/><Option name="style" value="solid" type="QString"/></Option></layer></symbol>'
        )
    return [
        f'{indent}<renderer-v2 type="categorizedSymbol" attr="release_id" symbollevels="0" forceraster="0" enableorderby="0" referencescale="-1">',
        f"{indent}  <categories>",
        *categories,
        f"{indent}  </categories>",
        f"{indent}  <symbols>",
        *symbols,
        f"{indent}  </symbols>",
        f"{indent}</renderer-v2>",
    ]


def _temporal_shapefile_qgs_xml(
    project: TemporalProject,
    layers: list[TemporalShapefileLayer],
    validations: list[ExportedLayerValidation],
    project_extent_3857: Bounds,
    rasters: list[TemporalQgisRaster],
) -> str:
    project_extent_3857 = _validate_bounds(project_extent_3857)
    layer_ids = {layer.filename: f"{layer.filename}_{index}" for index, layer in enumerate(layers, start=1)}
    validation_by_name = {validation.display_name: validation for validation in validations}
    visible_names = _default_visible_layer_names(layers)
    completed = [milestone for milestone in project.milestones if milestone.status == "complete"]
    latest_release = completed[-1].release_identifier if completed else None
    raster_by_release = {raster.release_identifier: raster for raster in rasters}
    latest_raster = raster_by_release.get(latest_release) if latest_release else None
    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<qgis projectname="" version="3.40.0">',
        f"  <title>{html.escape(project.name)}</title>",
        '  <homePath path=""/>',
        '  <projectionsEnabled>1</projectionsEnabled>',
        '  <autotransaction active="0"/>',
        '  <evaluateDefaultValues active="0"/>',
        '  <trust active="0"/>',
    ]
    lines.extend(["  <projectCrs>"])
    lines.extend(_qgis_srs_xml(QGIS_PROJECT_CRS, "    ")[1:-1])
    lines.extend(["  </projectCrs>", '  <layer-tree-group name="Résultats temporels" expanded="1">'])
    lines.append(
        '    <layer-tree-group name="Bâtiments ajoutés par date" checked="Qt::Checked" expanded="1" mutually-exclusive="1" mutually-exclusive-child="0">'
    )
    for milestone in reversed(completed):
        is_latest = milestone.release_identifier == latest_release
        checked = "Qt::Checked" if is_latest else "Qt::Unchecked"
        lines.append(
            f'      <layer-tree-group name="{html.escape(_milestone_quarter_label(milestone))}" checked="{checked}" expanded="{1 if is_latest else 0}">'
        )
        raster = raster_by_release.get(milestone.release_identifier)
        for layer in [item for item in layers if not item.is_global and item.release_identifier == milestone.release_identifier]:
            layer_checked = "Qt::Checked" if layer.group_key in {"batiments_ajoutes_par_date", "buffer_10m"} else "Qt::Unchecked"
            lines.append(
                f'        <layer-tree-layer id="{layer_ids[layer.filename]}" name="{html.escape(layer.display_name)}" checked="{layer_checked}" expanded="0" providerKey="ogr" source="{html.escape(validation_by_name[layer.display_name].relative_path)}"/>'
            )
        if raster is not None and raster.release_identifier != latest_release:
            lines.append(
                f'        <layer-tree-layer id="{raster.layer_id}" name="{html.escape(raster.display_name)}" checked="Qt::Checked" expanded="0" providerKey="gdal" source="{html.escape(raster.relative_path)}"/>'
            )
        lines.append("      </layer-tree-group>")
    lines.append("    </layer-tree-group>")
    lines.append('    <layer-tree-group name="Synthèse" expanded="1">')
    for layer in [item for item in layers if item.is_global]:
        checked = "Qt::Checked" if layer.display_name in visible_names else "Qt::Unchecked"
        lines.append(
            f'      <layer-tree-layer id="{layer_ids[layer.filename]}" name="{html.escape(layer.display_name)}" checked="{checked}" expanded="0" providerKey="ogr" source="{html.escape(validation_by_name[layer.display_name].relative_path)}"/>'
        )
    if latest_raster is not None:
        lines.append(
            f'      <layer-tree-layer id="{latest_raster.layer_id}" name="{html.escape(latest_raster.display_name)}" checked="Qt::Checked" expanded="0" providerKey="gdal" source="{html.escape(latest_raster.relative_path)}"/>'
        )
    lines.append("    </layer-tree-group>")
    lines.extend(["  </layer-tree-group>", "  <layerorder>"])
    for layer in layers:
        lines.append(f'    <layer id="{layer_ids[layer.filename]}"/>')
    for raster in reversed(rasters):
        lines.append(f'    <layer id="{raster.layer_id}"/>')
    lines.extend(["  </layerorder>", "  <mapcanvas name=\"theMapCanvas\" annotationsVisible=\"1\" destinationCrs=\"EPSG:3857\">"])
    lines.extend(_extent_xml(project_extent_3857, "    "))
    lines.extend(_qgis_srs_xml(QGIS_PROJECT_CRS, "    "))
    lines.extend(["    <rotation>0</rotation>", "    <renderMapTile>0</renderMapTile>", "  </mapcanvas>", "  <projectViewSettings>"])
    lines.extend(["    <mapCanvasExtent>"])
    lines.extend(_extent_xml(project_extent_3857, "      ")[1:-1])
    lines.extend(_qgis_srs_xml(QGIS_PROJECT_CRS, "      "))
    lines.extend(["    </mapCanvasExtent>"])
    lines.extend(["    <defaultViewExtent>"])
    lines.extend(_extent_xml(project_extent_3857, "      ")[1:-1])
    lines.extend(_qgis_srs_xml(QGIS_PROJECT_CRS, "      "))
    lines.extend(["    </defaultViewExtent>", "  </projectViewSettings>", "  <projectlayers>"])
    for layer in layers:
        validation = validation_by_name[layer.display_name]
        source = validation.relative_path
        layer_bounds = (validation.xmin, validation.ymin, validation.xmax, validation.ymax)
        fill_color, outline_color, outline_width = _qgis_layer_style(project, layer)
        lines.extend(
            [
                f'    <maplayer type="vector" geometry="Polygon" simplifyDrawingHints="1" simplifyDrawingTol="1" simplifyLocal="1" simplifyMaxScale="1" readOnly="0" hasScaleBasedVisibilityFlag="0" autoRefreshMode="Disabled" autoRefreshTime="0" styleCategories="AllStyleCategories" name="{html.escape(layer.display_name)}" id="{layer_ids[layer.filename]}">',
                f"      <id>{layer_ids[layer.filename]}</id>",
                f"      <layername>{html.escape(layer.display_name)}</layername>",
                f"      <datasource>{html.escape(source)}</datasource>",
                "      <provider>ogr</provider>",
            ]
        )
        lines.extend(_extent_xml(layer_bounds))
        lines.extend(_qgis_srs_xml(QGIS_VECTOR_CRS))
        if layer.group_key == "tous_les_nouveaux_batiments":
            lines.extend(_categorized_renderer_xml(project))
        elif layer.is_global and layer.group_key.startswith("buffer_"):
            lines.extend(_categorized_renderer_xml(project, fill_alpha=72))
        else:
            lines.extend(
                [
                    '      <renderer-v2 type="singleSymbol" symbollevels="0" forceraster="0" enableorderby="0" referencescale="-1">',
                    '        <symbols><symbol type="fill" name="0" alpha="1" clip_to_extent="1" force_rhr="0"><layer class="SimpleFill" enabled="1" pass="0" locked="0">',
                    f'          <Option type="Map"><Option name="color" value="{fill_color}" type="QString"/><Option name="outline_color" value="{outline_color}" type="QString"/><Option name="outline_width" value="{outline_width}" type="QString"/><Option name="outline_width_unit" value="Pixel" type="QString"/><Option name="style" value="solid" type="QString"/></Option>',
                    "        </layer></symbol></symbols>",
                    "      </renderer-v2>",
                ]
            )
        lines.extend(["      <labeling type=\"simple\"/>", "      <customproperties/>", "    </maplayer>"])
    for raster in rasters:
        logger.info(
            "EXPORT_RESULTS_QGIS_RASTER_LAYER_ADDED projectId=%s release=%s layer=%s source=%s",
            project.project_id,
            raster.release_identifier,
            raster.display_name,
            raster.relative_path,
        )
        lines.extend(
            [
                f'    <maplayer type="raster" hasScaleBasedVisibilityFlag="0" autoRefreshMode="Disabled" autoRefreshTime="0" name="{html.escape(raster.display_name)}" id="{raster.layer_id}">',
                f"      <id>{raster.layer_id}</id>",
                f"      <layername>{html.escape(raster.display_name)}</layername>",
                f"      <datasource>{html.escape(raster.relative_path)}</datasource>",
                "      <provider>gdal</provider>",
            ]
        )
        lines.extend(_extent_xml(raster.bounds_3857))
        lines.extend(_qgis_srs_xml(QGIS_PROJECT_CRS))
        lines.extend(["    </maplayer>"])
    lines.extend(
        [
            "  </projectlayers>",
            "  <properties>",
            "    <SpatialRefSys>",
            '      <ProjectionsEnabled type="int">1</ProjectionsEnabled>',
            "    </SpatialRefSys>",
            "  </properties>",
            "</qgis>",
        ]
    )
    return "\n".join(lines)


def _generate_qgz_with_qgis_api(
    qgz_path: Path,
    project: TemporalProject,
    layers: list[TemporalShapefileLayer],
    validations: list[ExportedLayerValidation],
) -> bool:
    try:
        from qgis.core import QgsCoordinateReferenceSystem, QgsProject, QgsVectorLayer
    except Exception:
        return False
    qgis_project = QgsProject()
    qgis_project.clear()
    qgis_project.setTitle(project.name)
    qgis_project.setCrs(QgsCoordinateReferenceSystem("EPSG:4326"))
    qgis_project.setFilePathStorage(QgsProject.FilePathType.Relative)
    root = qgis_project.layerTreeRoot()
    validation_by_name = {validation.display_name: validation for validation in validations}
    groups: dict[str, Any] = {}
    visible_names = _default_visible_layer_names(layers)
    for layer in layers:
        group = groups.get(layer.group_label)
        if group is None:
            group = root.addGroup(layer.group_label)
            groups[layer.group_label] = group
        vector = QgsVectorLayer(str(validation_by_name[layer.display_name].path), layer.display_name, "ogr")
        if not vector.isValid():
            return False
        qgis_project.addMapLayer(vector, False)
        node = group.addLayer(vector)
        node.setItemVisibilityChecked(layer.display_name in visible_names)
    return bool(qgis_project.write(str(qgz_path)) and qgz_path.is_file())


def _write_temporal_shapefile_qgz(
    path: Path,
    project: TemporalProject,
    layers: list[TemporalShapefileLayer],
    validations: list[ExportedLayerValidation],
    project_extent_3857: Bounds,
    rasters: list[TemporalQgisRaster],
) -> str:
    logger.info("EXPORT_QGZ_GENERATE_START projectId=%s backend=styled_xml", project.project_id)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED, allowZip64=True) as archive:
        archive.writestr(
            f"{path.stem}.qgs",
            _temporal_shapefile_qgs_xml(project, layers, validations, project_extent_3857, rasters),
        )
    logger.info(
        "QGIS_PROJECT_QGZ_WRITTEN projectId=%s path=%s bytes=%s crs=%s",
        project.project_id,
        path,
        path.stat().st_size if path.exists() else 0,
        QGIS_PROJECT_CRS,
    )
    return "styled_xml"


def build_temporal_results_shapefile_zip(project_id: str, settings: Settings | None = None) -> bytes:
    resolved_settings = _settings(settings)
    project = _load_project(project_id, resolved_settings)
    started_at = time.perf_counter()
    completed = [milestone for milestone in project.milestones if milestone.status == "complete"]
    if completed:
        logger.info(
            "EXPORT_LAYER_SKIPPED_BASELINE projectId=%s release=%s",
            project_id,
            completed[0].release_identifier,
        )
    for milestone in completed[1:]:
        expected_payloads = {
            "additions": milestone.additions_geojson,
            "building_change_buffer_10m": milestone.buffer_layers_geojson.get("10m"),
            "building_change_buffer_15m": milestone.buffer_layers_geojson.get("15m"),
            "building_change_buffer_20m": milestone.buffer_layers_geojson.get("20m"),
        }
        for artifact_key, payload in expected_payloads.items():
            if _features(payload):
                logger.info(
                    "EXPORT_LAYER_NON_EMPTY projectId=%s group=%s release=%s artifact_key=%s featureCount=%s",
                    project_id,
                    "batiments_ajoutes_par_date" if artifact_key == "additions" else artifact_key.replace("building_change_", ""),
                    milestone.release_identifier,
                    artifact_key,
                    len(_features(payload)),
                )
                continue
            logger.info(
                "EXPORT_LAYER_EMPTY_SKIPPED projectId=%s group=%s release=%s artifact_key=%s reason=no_features_after_resolution",
                project_id,
                "batiments_ajoutes_par_date" if artifact_key == "additions" else artifact_key.replace("building_change_", ""),
                milestone.release_identifier,
                artifact_key,
            )
    layers = _temporal_shapefile_export_layers(project)
    if not layers:
        raise ValueError(f"No non-empty temporal result layers are available for project {project_id}.")
    group_keys = [
        "tous_les_nouveaux_batiments",
        "batiments_ajoutes_par_date",
        "buffer_10m",
        "buffer_15m",
        "buffer_20m",
    ]
    if any(layer.group_key == "zone_export" for layer in layers):
        group_keys.append("zone_export")
    zip_stream = BytesIO()
    with tempfile.TemporaryDirectory(prefix="temporal-shapefile-export-") as tmp_name:
        tmp_dir = Path(tmp_name)
        written_layers: list[TemporalShapefileLayer] = []
        validated_layers: list[ExportedLayerValidation] = []
        for group_key in group_keys:
            logger.info("EXPORT_GROUP_START projectId=%s group=%s", project_id, group_key)
            (tmp_dir / group_key).mkdir(parents=True, exist_ok=True)
            for layer in [candidate for candidate in layers if candidate.group_key == group_key]:
                shp_path = _write_temporal_shapefile(layer, tmp_dir / group_key)
                if shp_path is None:
                    logger.info(
                        "EXPORT_LAYER_SKIPPED_EMPTY projectId=%s group=%s release=%s artifact_key=%s path=%s",
                        project_id,
                        group_key,
                        layer.release_identifier,
                        layer.artifact_key,
                        tmp_dir / group_key / f"{layer.filename}.shp",
                    )
                    continue
                written_layers.append(layer)
                validation = validate_exported_vector_layer(
                    shp_path,
                    layer.display_name,
                    layer.group_label,
                    tmp_dir,
                )
                if not validation.is_valid_for_qgis_project:
                    written_layers.pop()
                    logger.warning(
                        "EXPORT_QGZ_LAYER_SKIPPED_INVALID projectId=%s path=%s reason=%s",
                        project_id,
                        shp_path,
                        validation.reason,
                    )
                    continue
                validated_layers.append(validation)
                logger.info(
                    "EXPORT_QGZ_LAYER_VALIDATED projectId=%s path=%s feature_count=%s crs=%s xmin=%s ymin=%s xmax=%s ymax=%s",
                    project_id,
                    shp_path,
                    validation.feature_count,
                    validation.crs_authid,
                    validation.xmin,
                    validation.ymin,
                    validation.xmax,
                    validation.ymax,
                )
                event = "EXPORT_GLOBAL_LAYER_WRITTEN" if layer.is_global else "EXPORT_LAYER_WRITTEN"
                logger.info(
                    "%s projectId=%s group=%s release=%s artifact_key=%s feature_count=%s path=%s",
                    event,
                    project_id,
                    group_key,
                    layer.release_identifier,
                    layer.artifact_key,
                    len(_features(layer.feature_collection)),
                    shp_path,
                )
        if not written_layers:
            raise ValueError(f"No valid temporal result geometries are available for project {project_id}.")
        rasters = _copy_qgis_reference_rasters(project, resolved_settings, tmp_dir)
        if not rasters:
            logger.warning("EXPORT_RESULTS_QGIS_RASTER_ALL_MISSING projectId=%s", project_id)
        validation_by_name = {validation.display_name: validation for validation in validated_layers}
        for group_key in group_keys:
            group_layers = [layer for layer in written_layers if layer.group_key == group_key]
            if not group_layers:
                continue
            group_extent = _union_bounds(
                [
                    (
                        validation_by_name[layer.display_name].xmin,
                        validation_by_name[layer.display_name].ymin,
                        validation_by_name[layer.display_name].xmax,
                        validation_by_name[layer.display_name].ymax,
                    )
                    for layer in group_layers
                ]
            )
            logger.info(
                "EXPORT_QGZ_GROUP_EXTENT projectId=%s group=%s xmin=%s ymin=%s xmax=%s ymax=%s layer_count=%s",
                project_id,
                group_key,
                *group_extent,
                len(group_layers),
            )
        vector_extent = _union_bounds(
            [(item.xmin, item.ymin, item.xmax, item.ymax) for item in validated_layers]
        )
        initial_extent = compute_qgis_project_initial_extent(
            project_id,
            validated_layers,
            rasters,
            fallback_extent_wgs84=tuple(shape(project.aoi_geojson).bounds) if project.aoi_geojson else None,
        )
        logger.info(
            "EXPORT_QGZ_SOURCE_VECTOR_EXTENT projectId=%s crs=%s xmin=%s ymin=%s xmax=%s ymax=%s",
            project_id,
            QGIS_VECTOR_CRS,
            *vector_extent,
        )
        logger.info(
            "EXPORT_QGZ_PROJECT_EXTENT projectId=%s crs=%s paddingRatio=%s xmin=%s ymin=%s xmax=%s ymax=%s",
            project_id,
            QGIS_PROJECT_CRS,
            QGIS_PROJECT_EXTENT_PADDING_RATIO,
            *initial_extent.bounds_3857,
        )
        logger.info(
            "EXPORT_QGZ_LAYER_ORDER projectId=%s order=%s",
            project_id,
            ",".join([*[layer.display_name for layer in written_layers], *[raster.display_name for raster in rasters]]),
        )
        logger.info(
            'EXPORT_RESULTS_QGIS_LAYER_TREE projectId=%s rootGroup="Bâtiments ajoutés par date" mutuallyExclusive=true groups=Synthèse',
            project_id,
        )
        logger.info(
            "EXPORT_RESULTS_QGIS_DEFAULT_VISIBILITY projectId=%s activeDate=%s raster=true additions=true buffer10=true buffer15=false buffer20=false",
            project_id,
            _milestone_quarter_label(completed[-1]) if completed else "",
        )
        logger.info("EXPORT_RESULTS_QGIS_ONLINE_BASEMAP_REMOVED projectId=%s", project_id)
        qgz_path = tmp_dir / f"resultats_{_filesystem_safe_label(project.project_id)}.qgz"
        qgz_backend = _write_temporal_shapefile_qgz(
            qgz_path,
            project,
            written_layers,
            validated_layers,
            initial_extent.bounds_3857,
            rasters,
        )
        if not qgz_path.is_file() or qgz_path.stat().st_size <= 1024:
            logger.error(
                "EXPORT_QGZ_VALIDATION_FAILED projectId=%s path=%s reason=missing_or_unrealistically_small",
                project_id,
                qgz_path,
            )
            raise ValueError(f"Generated QGZ is missing or unrealistically small: {qgz_path}")
        logger.info(
            "EXPORT_RESULTS_QGIS_WRITTEN projectId=%s qgz=%s vectorLayers=%s rasterLayers=%s backend=%s projectCrs=%s bytes=%s",
            project_id,
            qgz_path,
            len(written_layers),
            len(rasters),
            qgz_backend,
            QGIS_PROJECT_CRS,
            qgz_path.stat().st_size,
        )
        with zipfile.ZipFile(zip_stream, "w", compression=zipfile.ZIP_DEFLATED, allowZip64=True) as archive:
            for group_key in group_keys:
                archive.writestr(f"{group_key}/", b"")
            archive.writestr(
                "README.txt",
                "\n".join(
                    [
                        "Ouvrez le fichier .qgz dans QGIS pour utiliser le projet style.",
                        "Le projet s'ouvre directement sur l'etendue des resultats exportes.",
                        "Les dossiers contiennent les composants Shapefile bruts et les rasters de reference.",
                        "Si QGIS demande de reparer les chemins, conservez la structure des dossiers extraite du ZIP.",
                        "",
                    ]
                ),
            )
            archive.write(qgz_path, arcname=qgz_path.name)
            for path in sorted(tmp_dir.rglob("*")):
                if path.is_file() and path != qgz_path:
                    archive.write(path, arcname=str(path.relative_to(tmp_dir)))
        logger.info(
            "EXPORT_SHAPEFILE_ZIP_VALIDATE_DONE projectId=%s layerCount=%s rasterCount=%s qgzBytes=%s",
            project_id,
            len(written_layers),
            len(rasters),
            qgz_path.stat().st_size,
        )
    payload = zip_stream.getvalue()
    logger.info(
        "EXPORT_ZIP_WRITTEN projectId=%s path=results_shapefile.zip bytes=%s durationMs=%s",
        project_id,
        len(payload),
        round((time.perf_counter() - started_at) * 1000, 2),
    )
    return payload


def _project_json_path(project_id: str, settings: Settings) -> Path:
    return settings.temporal_projects_dir / project_id / "project.json"


def _topojson_export_metadata_path(path: Path) -> Path:
    return path.with_name(f"{path.name}.metadata.json")


def _shapefile_export_metadata_path(path: Path) -> Path:
    return path.with_name(f"{path.name}.metadata.json")


def _topojson_cache_version_is_valid(path: Path) -> bool:
    metadata_path = _topojson_export_metadata_path(path)
    if not metadata_path.is_file():
        return False
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    except Exception:
        return False
    return metadata.get("version") == TOPOJSON_EXPORT_VERSION


def _shapefile_cache_version_is_valid(path: Path) -> bool:
    metadata_path = _shapefile_export_metadata_path(path)
    if not metadata_path.is_file():
        return False
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    except Exception:
        return False
    return metadata.get("version") == SHAPEFILE_EXPORT_VERSION


def _write_topojson_export_metadata(path: Path, project_id: str) -> None:
    metadata = {
        "version": TOPOJSON_EXPORT_VERSION,
        "project_id": project_id,
        "quantization": TOPOJSON_DEFAULT_QUANTIZATION,
        "layers": list(TOPOJSON_ALLOWED_LAYERS),
        "property_keys": list(TOPOJSON_PROPERTY_KEYS),
        "updated_at": _export_now().isoformat().replace("+00:00", "Z"),
    }
    _atomic_write_bytes(_topojson_export_metadata_path(path), json.dumps(metadata, separators=(",", ":")).encode("utf-8"))


def _write_shapefile_export_metadata(path: Path, project_id: str) -> None:
    metadata = {
        "version": SHAPEFILE_EXPORT_VERSION,
        "project_id": project_id,
        "updated_at": _export_now().isoformat().replace("+00:00", "Z"),
    }
    _atomic_write_bytes(_shapefile_export_metadata_path(path), json.dumps(metadata, separators=(",", ":")).encode("utf-8"))


def _export_metadata_path(path: Path) -> Path:
    return path.with_name(f"{path.name}.metadata.json")


def _export_artifact_manifest_path(settings: Settings, project_id: str) -> Path:
    return settings.temporal_projects_dir / project_id / "exports" / "export_artifact_manifest.json"


def _atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f"{path.name}.partial")
    try:
        tmp_path.write_text(json.dumps(payload, sort_keys=True, separators=(",", ":")), encoding="utf-8")
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def _feature_collection_bounds(payload: dict[str, Any] | None) -> list[float] | None:
    bounds: list[Bounds] = []
    for feature in _features(payload):
        geometry_payload = feature.get("geometry") if isinstance(feature, dict) else None
        if not isinstance(geometry_payload, dict):
            continue
        try:
            geometry = _polygonal_geometry(shape(geometry_payload))
        except Exception:
            continue
        if geometry.is_empty:
            continue
        bounds.append(tuple(float(value) for value in geometry.bounds))
    if not bounds:
        return None
    union = _union_bounds(bounds)
    return [union[0], union[1], union[2], union[3]]


def _feature_collection_geometry_types(payload: dict[str, Any] | None) -> list[str]:
    geometry_types: set[str] = set()
    for feature in _features(payload):
        geometry_payload = feature.get("geometry") if isinstance(feature, dict) else None
        if isinstance(geometry_payload, dict) and isinstance(geometry_payload.get("type"), str):
            geometry_types.add(geometry_payload["type"])
    return sorted(geometry_types)


def _artifact_manifest_entry(
    *,
    project: TemporalProject,
    milestone: TemporalMilestone,
    layer_type: str,
    artifact_key: str,
    settings: Settings,
) -> dict[str, Any]:
    artifact = next((item for item in milestone.artifacts if item.key == artifact_key), None)
    payload = _existing_result_layer_payload(milestone, layer_type)
    try:
        path, media_type = resolve_temporal_project_artifact_path(
            project_id=project.project_id,
            release_identifier=milestone.release_identifier,
            artifact_key=artifact_key,
            settings=settings,
            access_mode="export_artifact_manifest",
        )
    except FileNotFoundError:
        return {
            "release_identifier": milestone.release_identifier,
            "layer_type": layer_type,
            "artifact_key": artifact_key,
            "exists": False,
            "feature_count": len(_features(payload)),
            "bbox": _feature_collection_bounds(payload),
            "geometry_types": _feature_collection_geometry_types(payload),
        }
    stat = path.stat()
    bbox = artifact.bbox if artifact is not None and artifact.bbox is not None else _feature_collection_bounds(payload)
    feature_count = artifact.feature_count if artifact is not None and artifact.feature_count is not None else len(_features(payload))
    return {
        "release_identifier": milestone.release_identifier,
        "layer_type": layer_type,
        "artifact_key": artifact_key,
        "exists": True,
        "path": str(path),
        "media_type": media_type,
        "feature_count": feature_count,
        "bbox": bbox,
        "geometry_types": _feature_collection_geometry_types(payload),
        "size_bytes": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
        "sha256": artifact.sha256 if artifact is not None else None,
        "source_mtime_ns": artifact.source_mtime_ns if artifact is not None else None,
    }


def _build_or_refresh_export_artifact_manifest(project: TemporalProject, settings: Settings) -> dict[str, Any]:
    manifest_path = _export_artifact_manifest_path(settings, project.project_id)
    entries: list[dict[str, Any]] = []
    for milestone in project.milestones:
        if milestone.status != "complete":
            continue
        for layer_type, (artifact_key, _field_name, _distance_key) in EXPORT_RESULT_ARTIFACT_KEYS.items():
            entries.append(
                _artifact_manifest_entry(
                    project=project,
                    milestone=milestone,
                    layer_type=layer_type,
                    artifact_key=artifact_key,
                    settings=settings,
                )
            )
    entries_fingerprint = hashlib.sha256(json.dumps(entries, sort_keys=True, separators=(",", ":")).encode("utf-8")).hexdigest()
    manifest = {
        "version": EXPORT_ARTIFACT_MANIFEST_VERSION,
        "project_id": project.project_id,
        "entries": entries,
        "entry_count": len(entries),
        "manifest_fingerprint": entries_fingerprint,
        "updated_at": _export_now().isoformat().replace("+00:00", "Z"),
    }
    _atomic_write_json(manifest_path, manifest)
    logger.info(
        "EXPORT_LIGHTWEIGHT_MANIFEST_UPDATED projectId=%s path=%s entries=%s fingerprint=%s",
        project.project_id,
        manifest_path,
        len(entries),
        entries_fingerprint,
    )
    return manifest


def _load_export_artifact_manifest(settings: Settings, project_id: str) -> dict[str, Any] | None:
    manifest_path = _export_artifact_manifest_path(settings, project_id)
    logger.info("EXPORT_LIGHTWEIGHT_MANIFEST_LOAD_START projectId=%s path=%s", project_id, manifest_path)
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        logger.info("EXPORT_LIGHTWEIGHT_MANIFEST_LOAD_MISS projectId=%s reason=missing_manifest", project_id)
        return None
    except Exception as exc:
        logger.info(
            "EXPORT_LIGHTWEIGHT_MANIFEST_LOAD_MISS projectId=%s reason=invalid_manifest error=%s",
            project_id,
            exc.__class__.__name__,
        )
        return None
    if not isinstance(manifest, dict) or manifest.get("version") != EXPORT_ARTIFACT_MANIFEST_VERSION:
        logger.info("EXPORT_LIGHTWEIGHT_MANIFEST_LOAD_MISS projectId=%s reason=version", project_id)
        return None
    entries = manifest.get("entries")
    if not isinstance(entries, list):
        logger.info("EXPORT_LIGHTWEIGHT_MANIFEST_LOAD_MISS projectId=%s reason=entries", project_id)
        return None
    for entry in entries:
        if not isinstance(entry, dict):
            logger.info("EXPORT_LIGHTWEIGHT_MANIFEST_LOAD_MISS projectId=%s reason=entry_type", project_id)
            return None
        if entry.get("exists") is False:
            try:
                resolve_temporal_project_artifact_path(
                    project_id=project_id,
                    release_identifier=str(entry.get("release_identifier") or ""),
                    artifact_key=str(entry.get("artifact_key") or ""),
                    settings=settings,
                    access_mode="export_artifact_manifest_validate",
                )
            except FileNotFoundError:
                pass
            else:
                logger.info(
                    "EXPORT_CACHE_INVALIDATED projectId=%s reason=artifact_manifest_entry_created release=%s artifact_key=%s",
                    project_id,
                    entry.get("release_identifier"),
                    entry.get("artifact_key"),
                )
                return None
        if entry.get("exists") is True and not _metadata_fingerprint_matches(entry, require_existing_path=True):
            logger.info(
                "EXPORT_CACHE_INVALIDATED projectId=%s reason=artifact_manifest_entry_changed release=%s artifact_key=%s",
                project_id,
                entry.get("release_identifier"),
                entry.get("artifact_key"),
            )
            return None
    logger.info(
        "EXPORT_LIGHTWEIGHT_MANIFEST_LOAD_DONE projectId=%s path=%s entries=%s fingerprint=%s",
        project_id,
        manifest_path,
        len(entries),
        manifest.get("manifest_fingerprint"),
    )
    return manifest


def _export_scope_fingerprint(export_context: dict[str, Any] | None) -> dict[str, Any]:
    if not export_context or export_context.get("scope_type") == "project_aoi":
        return {"scope_type": "project_aoi", "scope_source": None, "geometry_hash": None}
    geometry = export_context["geometry"]
    return {
        "scope_type": "custom_geometry",
        "scope_source": export_context.get("source"),
        "geometry_hash": export_context.get("geometry_hash") or hashlib.sha256(geometry.wkb).hexdigest(),
        "clipped_geometry_hash": hashlib.sha256(geometry.wkb).hexdigest(),
        "was_clipped_to_project_aoi": bool(export_context.get("was_clipped_to_project_aoi")),
    }


def _temporal_result_artifact_fingerprints(
    project: TemporalProject,
    settings: Settings,
) -> list[dict[str, Any]]:
    fingerprints: list[dict[str, Any]] = []
    for milestone in project.milestones:
        if milestone.status != "complete":
            continue
        for artifact_key, _field_name, _distance_key in EXPORT_RESULT_ARTIFACT_KEYS.values():
            try:
                path, media_type = resolve_temporal_project_artifact_path(
                    project_id=project.project_id,
                    release_identifier=milestone.release_identifier,
                    artifact_key=artifact_key,
                    settings=settings,
                    access_mode="export_cache_fingerprint",
                )
            except FileNotFoundError:
                fingerprints.append(
                    {
                        "release_identifier": milestone.release_identifier,
                        "artifact_key": artifact_key,
                        "exists": False,
                    }
                )
                continue
            stat = path.stat()
            artifact = next((item for item in milestone.artifacts if item.key == artifact_key), None)
            fingerprints.append(
                {
                    "release_identifier": milestone.release_identifier,
                    "artifact_key": artifact_key,
                    "exists": True,
                    "path": str(path),
                    "media_type": media_type,
                    "size_bytes": stat.st_size,
                    "mtime_ns": stat.st_mtime_ns,
                    "sha256": artifact.sha256 if artifact is not None else None,
                }
            )
    return fingerprints


def _project_fingerprint(project_id: str, settings: Settings) -> dict[str, Any]:
    project_path = _project_json_path(project_id, settings)
    if not project_path.is_file():
        return {"exists": False}
    stat = project_path.stat()
    return {
        "exists": True,
        "path": str(project_path),
        "size_bytes": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
    }


def _export_cache_metadata(
    *,
    project: TemporalProject,
    settings: Settings,
    export_format: str,
    export_context: dict[str, Any] | None,
) -> dict[str, Any]:
    scope = _export_scope_fingerprint(export_context)
    artifact_manifest = _build_or_refresh_export_artifact_manifest(project, settings)
    payload = {
        "version": EXPORT_FORMAT_VERSIONS[export_format],
        "cache_version": EXPORT_CACHE_VERSION,
        "exporter_version": EXPORT_FORMAT_VERSIONS[export_format],
        "project_id": project.project_id,
        "format": export_format,
        "scope": scope,
        "layers": list(EXPORT_RESULT_ARTIFACT_KEYS),
        "groups": [
            "tous_les_nouveaux_batiments",
            "batiments_ajoutes_par_date",
            "buffer_10m",
            "buffer_15m",
            "buffer_20m",
            *([] if scope["scope_type"] == "project_aoi" else ["zone_export"]),
        ],
        "options": {
            "topojson_quantization": TOPOJSON_DEFAULT_QUANTIZATION if export_format == "topojson" else None,
            "qgis_project_crs": QGIS_PROJECT_CRS if export_format == "shapefile" else None,
        },
        "project_fingerprint": _project_fingerprint(project.project_id, settings),
        "artifact_fingerprints": _temporal_result_artifact_fingerprints(project, settings),
        "artifact_manifest": {
            "version": artifact_manifest.get("version"),
            "path": str(_export_artifact_manifest_path(settings, project.project_id)),
            "fingerprint": artifact_manifest.get("manifest_fingerprint"),
            "entry_count": artifact_manifest.get("entry_count"),
        },
    }
    cache_key = hashlib.sha256(json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")).hexdigest()
    return {
        **payload,
        "cache_key": cache_key,
        "updated_at": _export_now().isoformat().replace("+00:00", "Z"),
    }


def _write_export_cache_metadata(path: Path, metadata: dict[str, Any]) -> None:
    try:
        stat = path.stat()
    except OSError:
        output = {
            "path": str(path),
            "exists": False,
        }
    else:
        output = {
            "path": str(path),
            "exists": True,
            "size_bytes": stat.st_size,
            "mtime_ns": stat.st_mtime_ns,
            "created_at": metadata.get("updated_at"),
        }
    payload = {**metadata, "output": output}
    _atomic_write_bytes(_export_metadata_path(path), json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8"))
    logger.info(
        "EXPORT_CACHE_MANIFEST_UPDATED projectId=%s format=%s path=%s cacheKey=%s outputBytes=%s",
        payload.get("project_id"),
        payload.get("format"),
        _export_metadata_path(path),
        payload.get("cache_key"),
        output.get("size_bytes"),
    )


def _read_export_cache_metadata(path: Path) -> dict[str, Any] | None:
    metadata_path = _export_metadata_path(path)
    if not metadata_path.is_file():
        return None
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(metadata, dict):
        return None
    logger.info(
        "EXPORT_CACHE_MANIFEST_LOADED projectId=%s format=%s path=%s cacheKey=%s",
        metadata.get("project_id"),
        metadata.get("format"),
        metadata_path,
        metadata.get("cache_key"),
    )
    return metadata


def _export_cache_file_invalid(project_id: str, export_format: str, path: Path, reason: str) -> None:
    logger.info(
        "EXPORT_CACHE_FILE_INVALID projectId=%s format=%s path=%s reason=%s",
        project_id,
        export_format,
        path,
        reason,
    )


def _metadata_fingerprint_matches(fingerprint: dict[str, Any], *, require_existing_path: bool) -> bool:
    exists = fingerprint.get("exists")
    if exists is False:
        return not require_existing_path
    path_value = fingerprint.get("path")
    if not isinstance(path_value, str) or not path_value:
        return False
    path = Path(path_value)
    try:
        stat = path.stat()
    except OSError:
        return False
    if fingerprint.get("size_bytes") is not None and int(fingerprint["size_bytes"]) != stat.st_size:
        return False
    if fingerprint.get("mtime_ns") is not None and int(fingerprint["mtime_ns"]) != stat.st_mtime_ns:
        return False
    return True


def _artifact_fingerprints_match(fingerprints: Any) -> bool:
    if not isinstance(fingerprints, list):
        return False
    for fingerprint in fingerprints:
        if not isinstance(fingerprint, dict):
            return False
        if not _metadata_fingerprint_matches(fingerprint, require_existing_path=False):
            return False
    return True


def _is_project_aoi_export_scope(perimeter: dict[str, Any] | None) -> bool:
    if perimeter is None:
        return True
    return perimeter.get("mode") == "project_aoi" and perimeter.get("geometry") is None and perimeter.get("source") is None


def _cached_export_path(settings: Settings, project_id: str, export_format: str, suffix: str = "") -> Path:
    return settings.temporal_projects_dir / project_id / "exports" / (TEMPORAL_RESULTS_EXPORT_FILENAMES[export_format] + suffix)


def _fast_cached_temporal_results_export_file(
    project_id: str,
    export_format: str,
    settings: Settings,
    perimeter: dict[str, Any] | None,
) -> Path | None:
    is_project_aoi_scope = _is_project_aoi_export_scope(perimeter)
    geometry_hash: str | None = None
    custom_suffix = ""
    if not is_project_aoi_scope:
        try:
            geometry_hash = _stable_custom_geometry_hash(perimeter)
        except Exception as exc:
            logger.info(
                "EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=invalid_custom_geometry error=%s",
                project_id,
                export_format,
                exc.__class__.__name__,
            )
            return None
        if not geometry_hash:
            logger.info("EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=missing_custom_geometry", project_id, export_format)
            return None
        custom_suffix = f".custom-{geometry_hash[:12]}"
    logger.info(
        "EXPORT_FAST_CACHE_CHECK_START projectId=%s format=%s scopeType=%s geometryHash=%s",
        project_id,
        export_format,
        "project_aoi" if is_project_aoi_scope else "custom_geometry",
        geometry_hash,
    )

    cache_path = _cached_export_path(settings, project_id, export_format, custom_suffix)
    if cache_path.name.endswith(".partial"):
        _export_cache_file_invalid(project_id, export_format, cache_path, "partial_file")
        return None
    try:
        cache_stat = cache_path.stat()
    except OSError:
        logger.info(
            "EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=missing_file path=%s",
            project_id,
            export_format,
            cache_path,
        )
        return None
    if cache_stat.st_size <= 0:
        _export_cache_file_invalid(project_id, export_format, cache_path, "empty_file")
        return None
    if export_format == "topojson" and not _topojson_cache_version_is_valid(cache_path):
        _export_cache_file_invalid(project_id, export_format, cache_path, "topojson_cache_version")
        return None
    if export_format == "shapefile" and not _shapefile_cache_version_is_valid(cache_path):
        _export_cache_file_invalid(project_id, export_format, cache_path, "shapefile_cache_version")
        return None

    metadata = _read_export_cache_metadata(cache_path)
    if metadata is None:
        logger.info("EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=missing_metadata", project_id, export_format)
        return None
    if metadata.get("cache_version") != EXPORT_CACHE_VERSION or metadata.get("format") != export_format:
        logger.info("EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=metadata_version_or_format", project_id, export_format)
        return None
    scope = metadata.get("scope")
    expected_scope_type = "project_aoi" if is_project_aoi_scope else "custom_geometry"
    if not isinstance(scope, dict) or scope.get("scope_type") != expected_scope_type:
        logger.info("EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=metadata_scope", project_id, export_format)
        return None
    if not is_project_aoi_scope and scope.get("geometry_hash") != geometry_hash:
        logger.info("EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=metadata_geometry_hash", project_id, export_format)
        return None
    project_fingerprint = metadata.get("project_fingerprint")
    if not isinstance(project_fingerprint, dict) or not _metadata_fingerprint_matches(project_fingerprint, require_existing_path=True):
        logger.info("EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=project_fingerprint_changed", project_id, export_format)
        return None
    if not _artifact_fingerprints_match(metadata.get("artifact_fingerprints")):
        logger.info("EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=artifact_fingerprint_changed", project_id, export_format)
        return None
    manifest_info = metadata.get("artifact_manifest")
    manifest = _load_export_artifact_manifest(settings, project_id)
    if not isinstance(manifest_info, dict) or manifest is None:
        logger.info("EXPORT_FAST_CACHE_MISS projectId=%s format=%s reason=artifact_manifest_missing", project_id, export_format)
        return None
    if manifest_info.get("fingerprint") != manifest.get("manifest_fingerprint"):
        logger.info("EXPORT_CACHE_INVALIDATED projectId=%s format=%s reason=artifact_manifest_fingerprint_changed", project_id, export_format)
        return None

    logger.info("EXPORT_CACHE_FILE_VALIDATED projectId=%s format=%s path=%s bytes=%s", project_id, export_format, cache_path, cache_stat.st_size)
    logger.info(
        "EXPORT_FULL_PROJECT_LOAD_SKIPPED projectId=%s scopeType=%s reason=validated_export_cache",
        project_id,
        expected_scope_type,
    )
    logger.info(
        "EXPORT_FAST_CACHE_HIT projectId=%s format=%s scopeType=%s path=%s cacheKey=%s bytes=%s",
        project_id,
        export_format,
        expected_scope_type,
        cache_path,
        metadata.get("cache_key"),
        cache_stat.st_size,
    )
    return cache_path


def _export_cache_is_valid(
    path: Path,
    project_id: str,
    settings: Settings,
    export_format: str | None = None,
    expected_metadata: dict[str, Any] | None = None,
) -> bool:
    if not path.is_file():
        return False
    if export_format == "topojson" and not _topojson_cache_version_is_valid(path):
        return False
    if export_format == "shapefile" and not _shapefile_cache_version_is_valid(path):
        return False
    if expected_metadata is not None:
        metadata = _read_export_cache_metadata(path)
        if metadata is None:
            logger.info("EXPORT_CACHE_STALE projectId=%s format=%s reason=missing_metadata path=%s", project_id, export_format, path)
            return False
        if metadata.get("cache_key") != expected_metadata.get("cache_key"):
            logger.info(
                "EXPORT_CACHE_STALE projectId=%s format=%s reason=fingerprint_changed path=%s",
                project_id,
                export_format,
                path,
            )
            return False
        return True
    project_path = _project_json_path(project_id, settings)
    if not project_path.is_file():
        return True
    return path.stat().st_mtime_ns >= project_path.stat().st_mtime_ns


def _atomic_write_bytes(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f".{path.name}.{datetime.now(UTC).timestamp():.6f}.tmp")
    try:
        tmp_path.write_bytes(payload)
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def build_temporal_results_export_file(
    project_id: str,
    export_format: str,
    settings: Settings | None = None,
    perimeter: dict[str, Any] | None = None,
) -> Path:
    resolved_settings = _settings(settings)
    normalized_format = export_format.lower()
    if normalized_format == "zip":
        normalized_format = "shapefile"
    if normalized_format not in TEMPORAL_RESULTS_EXPORT_FILENAMES:
        raise ValueError(f"Unsupported temporal results export format: {export_format}")

    started_at = datetime.now(UTC)
    fast_cache_path = _fast_cached_temporal_results_export_file(project_id, normalized_format, resolved_settings, perimeter)
    if fast_cache_path is not None:
        duration_ms = round((datetime.now(UTC) - started_at).total_seconds() * 1000, 2)
        fast_scope_type = "project_aoi" if _is_project_aoi_export_scope(perimeter) else "custom_geometry"
        logger.info(
            "EXPORT_DOWNLOAD_TOTAL_MS projectId=%s format=%s scopeType=%s source=fast_cache durationMs=%s",
            project_id,
            normalized_format,
            fast_scope_type,
            duration_ms,
        )
        return fast_cache_path

    project = _load_project(project_id, resolved_settings)
    export_context = resolve_export_perimeter(project, perimeter)
    is_custom_export = export_context is not None
    custom_suffix = ""
    if is_custom_export:
        geometry_hash = export_context.get("geometry_hash") or hashlib.sha256(export_context["geometry"].wkb).hexdigest()
        export_context["geometry_hash"] = geometry_hash
        custom_suffix = f".custom-{geometry_hash[:12]}"
    cache_path = _cached_export_path(resolved_settings, project_id, normalized_format, custom_suffix)
    cache_metadata = _export_cache_metadata(
        project=project,
        settings=resolved_settings,
        export_format=normalized_format,
        export_context=export_context,
    )
    logger.info(
        "EXPORT_REQUEST projectId=%s format=%s scopeType=%s cacheKey=%s",
        project_id,
        normalized_format,
        cache_metadata["scope"]["scope_type"],
        cache_metadata["cache_key"],
    )
    if _export_cache_is_valid(cache_path, project_id, resolved_settings, normalized_format, cache_metadata):
        logger.info(
            "EXPORT_CACHE_HIT projectId=%s format=%s scopeType=%s path=%s cacheKey=%s",
            project_id,
            normalized_format,
            cache_metadata["scope"]["scope_type"],
            cache_path,
            cache_metadata["cache_key"],
        )
        duration_ms = round((datetime.now(UTC) - started_at).total_seconds() * 1000, 2)
        logger.info(
            "EXPORT_DOWNLOAD_TOTAL_MS projectId=%s format=%s scopeType=%s source=validated_cache durationMs=%s",
            project_id,
            normalized_format,
            cache_metadata["scope"]["scope_type"],
            duration_ms,
        )
        return cache_path

    logger.info(
        "EXPORT_GENERATE_START projectId=%s format=%s scopeType=%s cacheKey=%s",
        project_id,
        normalized_format,
        cache_metadata["scope"]["scope_type"],
        cache_metadata["cache_key"],
    )
    try:
        builders = {
            "xlsx": build_temporal_results_workbook,
            "kml": build_temporal_results_kml,
            "geojson": build_temporal_results_geojson,
            "topojson": build_temporal_results_topojson,
            "json": build_temporal_results_json,
            "tsv": build_temporal_results_tsv,
            "shapefile": build_temporal_results_shapefile_zip,
        }
        generation_context = (
            {**export_context, "loaded_project": project}
            if export_context is not None
            else {"scope_type": "project_aoi", "loaded_project": project}
        )
        context_token = _EXPORT_CONTEXT.set(generation_context)
        write_started_at = time.perf_counter()
        logger.info(
            "EXPORT_FORMAT_WRITE_START projectId=%s format=%s path=%s",
            project_id,
            normalized_format,
            cache_path,
        )
        payload = builders[normalized_format](project_id, settings=resolved_settings)
        _atomic_write_bytes(cache_path, payload)
        _write_export_cache_metadata(cache_path, cache_metadata)
        logger.info(
            "EXPORT_FORMAT_WRITE_DONE projectId=%s format=%s path=%s bytes=%s durationMs=%s",
            project_id,
            normalized_format,
            cache_path,
            len(payload),
            round((time.perf_counter() - write_started_at) * 1000, 2),
        )
    except Exception as exc:
        logger.exception("EXPORT_GENERATE_FAILED projectId=%s format=%s error=%s", project_id, normalized_format, exc)
        raise
    finally:
        if "context_token" in locals():
            _EXPORT_CONTEXT.reset(context_token)
    duration_ms = round((datetime.now(UTC) - started_at).total_seconds() * 1000, 2)
    logger.info(
        "EXPORT_GENERATE_DONE projectId=%s format=%s bytes=%s durationMs=%s",
        project_id,
        normalized_format,
        cache_path.stat().st_size,
        duration_ms,
    )
    logger.info(
        "EXPORT_DOWNLOAD_TOTAL_MS projectId=%s format=%s scopeType=%s source=generated durationMs=%s",
        project_id,
        normalized_format,
        cache_metadata["scope"]["scope_type"],
        duration_ms,
    )
    return cache_path
