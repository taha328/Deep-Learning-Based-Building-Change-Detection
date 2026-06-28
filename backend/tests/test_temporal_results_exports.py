from __future__ import annotations

import json
import logging
import math
from io import BytesIO
from pathlib import Path
import re
import xml.etree.ElementTree as ET
import zipfile

from fastapi.testclient import TestClient
import geopandas as gpd
import numpy as np
from openpyxl import load_workbook
import pytest
import rasterio
from rasterio.transform import from_bounds
from rasterio.warp import transform_bounds
from shapely.geometry import shape

from src.api.deps import get_app_settings
from src.api.main import app
from src.config import Settings
from src.schemas import TemporalMilestone, TemporalMilestoneMetrics, TemporalProject
from src.services.temporal_exports import (
    _filesystem_safe_label,
    _qgis_layer_style,
    _qgis_project_extent,
    _temporal_milestone_color_map,
    _temporal_shapefile_export_layers,
    build_temporal_results_export_file,
)
from src.services.temporal_projects import save_temporal_project


def _feature_collection() -> dict:
    return {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "properties": {
                    "area_m2": 123.4,
                    "score": 0.91,
                    "run_id": "debug-run",
                    "release_identifier": "WB_DEBUG",
                    "source_backend": "debug-backend",
                    "feature_index": 99,
                    "confidence": 0.91,
                    "status": "accepted",
                },
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [[
                        [-7.0, 33.0],
                        [-6.999, 33.0],
                        [-6.999, 33.001],
                        [-7.0, 33.001],
                        [-7.0, 33.0],
                    ]],
                },
            }
        ],
    }


def _feature_collection_from_bounds(*bounds_items: tuple[float, float, float, float]) -> dict:
    return {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "properties": {"area_m2": 100.0 + index, "score": 0.9},
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [[
                        [bounds[0], bounds[1]],
                        [bounds[2], bounds[1]],
                        [bounds[2], bounds[3]],
                        [bounds[0], bounds[3]],
                        [bounds[0], bounds[1]],
                    ]],
                },
            }
            for index, bounds in enumerate(bounds_items)
        ],
    }


def _settings(tmp_path: Path) -> Settings:
    return Settings(runtime_cache_dir=tmp_path)


def _project() -> TemporalProject:
    fc = _feature_collection()
    return TemporalProject(
        project_id="temporal-export-formats-test",
        name="Export Formats",
        aoi_geojson=fc["features"][0]["geometry"],
        milestones=[
            TemporalMilestone(
                release_identifier="WB_2022_R03",
                release_date="2022-03-16",
                status="complete",
                metrics=TemporalMilestoneMetrics(total_area_m2=1000.0),
            ),
            TemporalMilestone(
                release_identifier="WB_2024_R03",
                release_date="2024-03-28",
                status="complete",
                pair_request_hash="run-2024",
                additions_geojson=fc,
                cumulative_growth_blocks_geojson=fc,
                cumulative_growth_envelope_geojson=fc,
                automated_candidate_footprint_geojson=fc,
                buffer_layers_geojson={"10m": fc, "15m": fc, "20m": fc},
                metrics=TemporalMilestoneMetrics(
                    added_area_m2=123.4,
                    total_area_m2=1123.4,
                    additions_feature_count=1,
                    added_block_count=1,
                    cumulative_block_area_m2=123.4,
                ),
            ),
        ],
        created_at="2026-05-19T00:00:00Z",
        updated_at="2026-05-19T00:00:00Z",
    )


def _save_project(tmp_path: Path) -> Settings:
    settings = _settings(tmp_path)
    project = _project()
    _save_project_payload(settings, project)
    return settings


def _save_project_payload(settings: Settings, project: TemporalProject) -> None:
    save_temporal_project(project, settings)
    for index, milestone in enumerate(project.milestones):
        raster_path = (
            settings.temporal_projects_dir
            / project.project_id
            / "milestones"
            / milestone.release_identifier
            / "reference_imagery_cog.tif"
        )
        raster_path.parent.mkdir(parents=True, exist_ok=True)
        with rasterio.open(
            raster_path,
            "w",
            driver="GTiff",
            width=16,
            height=16,
            count=3,
            dtype="uint8",
            crs="EPSG:3857",
            transform=from_bounds(-779300, 3895300, -779100, 3895500, 16, 16),
        ) as dataset:
            dataset.write(np.full((3, 16, 16), 80 + index * 40, dtype=np.uint8))


def _three_milestone_project() -> TemporalProject:
    additions_2020 = _feature_collection_from_bounds((-7.0000, 33.0000, -6.9998, 33.0002))
    additions_2022 = _feature_collection_from_bounds((-6.9996, 33.0000, -6.9994, 33.0002))
    buffers_2020 = _feature_collection_from_bounds(
        (-7.00000, 33.00000, -6.99970, 33.00030),
        (-6.99985, 33.00000, -6.99955, 33.00030),
    )
    buffers_2022 = _feature_collection_from_bounds(
        (-6.99960, 33.00000, -6.99930, 33.00030),
        (-6.99945, 33.00000, -6.99915, 33.00030),
    )
    return TemporalProject(
        project_id="temporal-export-three-milestones",
        name="Three Milestones",
        aoi_geojson=_feature_collection_from_bounds((-7.001, 32.999, -6.998, 33.002))["features"][0]["geometry"],
        milestones=[
            TemporalMilestone(
                release_identifier="WB_2018_R01",
                release_date="2018-01-15",
                status="complete",
                metrics=TemporalMilestoneMetrics(total_area_m2=1000.0),
            ),
            TemporalMilestone(
                release_identifier="WB_2020_R01",
                release_date="2020-01-15",
                status="complete",
                additions_geojson=additions_2020,
                buffer_layers_geojson={"10m": buffers_2020, "15m": buffers_2020, "20m": buffers_2020},
                metrics=TemporalMilestoneMetrics(added_area_m2=100.0, total_area_m2=1100.0, additions_feature_count=1),
            ),
            TemporalMilestone(
                release_identifier="WB_2022_R01",
                release_date="2022-01-15",
                status="complete",
                additions_geojson=additions_2022,
                buffer_layers_geojson={"10m": buffers_2022, "15m": buffers_2022, "20m": buffers_2022},
                metrics=TemporalMilestoneMetrics(added_area_m2=100.0, total_area_m2=1200.0, additions_feature_count=1),
            ),
        ],
        created_at="2026-05-19T00:00:00Z",
        updated_at="2026-05-19T00:00:00Z",
    )


def test_temporal_results_export_formats_are_generated_under_project_exports(tmp_path: Path) -> None:
    settings = _save_project(tmp_path)
    project_id = "temporal-export-formats-test"

    paths = {
        export_format: build_temporal_results_export_file(project_id, export_format, settings=settings)
        for export_format in ("xlsx", "kml", "geojson", "topojson", "json", "tsv", "shapefile")
    }

    assert all(path.parent == settings.temporal_projects_dir / project_id / "exports" for path in paths.values())
    assert paths["xlsx"].name == "results.xlsx"
    assert paths["kml"].read_bytes().startswith(b"<?xml")

    geojson = json.loads(paths["geojson"].read_text())
    assert geojson["type"] == "FeatureCollection"
    assert geojson["features"]
    first_props = geojson["features"][0]["properties"]
    assert first_props["project_id"] == project_id
    assert first_props["release_identifier"] == "WB_2024_R03"
    assert first_props["layer_type"] in {"additions", "cumulative_growth", "buffer_10m", "buffer_15m", "buffer_20m", "diagnostics"}
    assert "area_m2" in first_props

    summary = json.loads(paths["json"].read_text())
    assert summary["project"]["project_id"] == project_id
    assert "features" not in json.dumps(summary)

    tsv = paths["tsv"].read_text()
    assert tsv.splitlines()[0].split("\t") == [
        "project_id",
        "project_name",
        "date",
        "layer_type",
        "area_m2",
        "centroid_lon",
        "centroid_lat",
    ]
    assert "\"geometry\"" not in tsv

    topojson = json.loads(paths["topojson"].read_text())
    assert topojson["type"] == "Topology"
    assert "results" in topojson["objects"]
    assert topojson["bbox"]
    assert topojson["transform"]
    assert topojson["arcs"]
    assert not (paths["topojson"].parent / "results_full.topojson").exists()

    with zipfile.ZipFile(paths["shapefile"]) as archive:
        names = archive.namelist()
        top_level_folders = {name.split("/", 1)[0] for name in names if "/" in name}
        assert top_level_folders == {
            "tous_les_nouveaux_batiments",
            "batiments_ajoutes_par_date",
            "buffer_10m",
            "buffer_15m",
            "buffer_20m",
            "rasters",
        }
        assert len([name for name in names if name.endswith(".qgz") and "/" not in name]) == 1
        assert "batiments_ajoutes_par_date/batiments_ajoutes_2024_Q1.shp" in names
        assert "buffer_10m/buffer_10m_2024_Q1.shp" in names
        assert "buffer_10m/buffer_10m_2022_Q1_2024_Q1.shp" in names
        assert not any("2022_Q1.shp" in name and "2022_Q1_2024_Q1" not in name for name in names)
        for shp_name in [name for name in names if name.endswith(".shp")]:
            stem = shp_name.removesuffix(".shp")
            for suffix in (".shx", ".dbf", ".prj", ".cpg"):
                assert f"{stem}{suffix}" in names
        qgz_name = next(name for name in names if name.endswith(".qgz"))
        qgz_path = tmp_path / "results.qgz"
        qgz_path.write_bytes(archive.read(qgz_name))
        with zipfile.ZipFile(qgz_path) as qgz:
            qgs_name = next(name for name in qgz.namelist() if name.endswith(".qgs"))
            qgs = qgz.read(qgs_name).decode("utf-8")
        assert "./batiments_ajoutes_par_date/batiments_ajoutes_2024_Q1.shp" in qgs
        assert "./buffer_10m/buffer_10m_2022_Q1_2024_Q1.shp" in qgs
        assert "./rasters/WB_2022_R03/reference_imagery_cog.tif" in qgs
        assert "./rasters/WB_2024_R03/reference_imagery_cog.tif" in qgs
        assert 'name="Bâtiments ajoutés par date"' in qgs
        assert 'mutually-exclusive="1"' in qgs
        assert "OpenStreetMap" not in qgs
        assert "Fond de carte en ligne" not in qgs
        assert 'name="Synthèse"' in qgs


def test_temporal_results_xlsx_schema_excludes_internal_fields_and_uses_150_zoom(tmp_path: Path) -> None:
    settings = _save_project(tmp_path)
    workbook_path = build_temporal_results_export_file("temporal-export-formats-test", "xlsx", settings=settings)

    workbook = load_workbook(workbook_path, data_only=True)

    assert workbook.sheetnames == ["Synthèse", "Jalons", "Détails blocs"]
    assert all(sheet.sheet_view.zoomScale == 150 for sheet in workbook.worksheets if sheet.sheet_state == "visible")
    assert all(sheet.sheet_view.zoomScaleNormal == 150 for sheet in workbook.worksheets if sheet.sheet_state == "visible")

    summary_values = [
        value
        for row in workbook["Synthèse"].iter_rows(values_only=True)
        for value in row
        if value is not None
    ]
    assert "Backend utilisé" not in summary_values
    assert "bandon_mps" not in summary_values
    assert "Identifiant du projet" in summary_values
    assert "Nom du projet" in summary_values
    assert "Date d'export" in summary_values
    assert "Nombre de jalons" in summary_values
    assert "Système de coordonnées utilisé pour les surfaces" in summary_values

    milestones = workbook["Jalons"]
    milestone_headers = [cell.value for cell in milestones[1]]
    assert milestone_headers[0] == "Date d'archive"
    assert "Jalon" not in milestone_headers
    assert milestones.max_column == len(milestone_headers)

    blocks = workbook["Détails blocs"]
    block_headers = [cell.value for cell in blocks[1]]
    assert block_headers == [
        "Date d'archive",
        "Identifiant bloc",
        "Surface (m²)",
        "Type géométrie",
        "Longitude centroïde",
        "Latitude centroïde",
    ]
    assert blocks.max_column == len(block_headers)


def test_temporal_results_tsv_powerbi_schema_ordering_and_row_specific_areas(tmp_path: Path) -> None:
    settings = _settings(tmp_path)
    project = _three_milestone_project()
    _save_project_payload(settings, project)

    path = build_temporal_results_export_file(project.project_id, "tsv", settings=settings)
    lines = path.read_text(encoding="utf-8").splitlines()
    header = lines[0].split("\t")

    assert header == [
        "project_id",
        "project_name",
        "date",
        "layer_type",
        "area_m2",
        "centroid_lon",
        "centroid_lat",
    ]
    removed_columns = {
        "run_id",
        "release_identifier",
        "growth_label",
        "feature_count",
        "added_surface_m2",
        "current_footprint_m2",
    }
    assert not (set(header) & removed_columns)
    assert all(len(line.split("\t")) == 7 for line in lines)

    rows = [dict(zip(header, line.split("\t", maxsplit=6))) for line in lines[1:]]
    assert [row["date"] for row in rows] == sorted(row["date"] for row in rows)
    assert all(re.fullmatch(r"\d{4}-\d{2}-\d{2}", row["date"]) for row in rows)

    known_layer_order = {"additions": 0, "buffer_10m": 1, "buffer_15m": 2, "buffer_20m": 3}
    for date_value in sorted({row["date"] for row in rows}):
        layers_for_date = [row["layer_type"] for row in rows if row["date"] == date_value]
        assert layers_for_date == sorted(layers_for_date, key=lambda layer: known_layer_order[layer])

    for row in rows:
        assert row["area_m2"]
        float(row["area_m2"])
        float(row["centroid_lon"])
        float(row["centroid_lat"])

    rows_by_date_layer = {(row["date"], row["layer_type"]): row for row in rows}
    for date_value in {"2020-01-15", "2022-01-15"}:
        additions_area = float(rows_by_date_layer[(date_value, "additions")]["area_m2"])
        for layer_type in ("buffer_10m", "buffer_15m", "buffer_20m"):
            buffer_area = float(rows_by_date_layer[(date_value, layer_type)]["area_m2"])
            assert buffer_area != pytest.approx(additions_area)
            assert buffer_area > additions_area


def test_temporal_shapefile_labels_baseline_exclusion_and_layer_order() -> None:
    assert _filesystem_safe_label("2017 Q1") == "2017_Q1"
    assert _filesystem_safe_label("2026 Q2") == "2026_Q2"
    assert _filesystem_safe_label("Bâtiments ajoutés / test") == "Batiments_ajoutes_test"

    fc = _feature_collection()
    project = TemporalProject(
        project_id="temporal-order",
        name="Temporal order",
        milestones=[
            TemporalMilestone(release_identifier="WB_2017_R05", release_date="2017-03-01", status="complete"),
            TemporalMilestone(
                release_identifier="WB_2019_R03",
                release_date="2019-03-01",
                status="complete",
                additions_geojson=fc,
                buffer_layers_geojson={"10m": fc, "15m": fc, "20m": fc},
                cumulative_union_geojson=fc,
            ),
            TemporalMilestone(
                release_identifier="WB_2021_R04",
                release_date="2021-03-01",
                status="complete",
                additions_geojson=fc,
                buffer_layers_geojson={"10m": fc, "15m": fc, "20m": fc},
                cumulative_union_geojson=fc,
            ),
        ],
        created_at="2026-01-01T00:00:00Z",
        updated_at="2026-01-01T00:00:00Z",
    )

    layers = _temporal_shapefile_export_layers(project)
    additions = [layer.filename for layer in layers if layer.group_key == "batiments_ajoutes_par_date"]
    buffer_10m = [layer.filename for layer in layers if layer.group_key == "buffer_10m"]

    assert additions == ["batiments_ajoutes_2021_Q1", "batiments_ajoutes_2019_Q1"]
    assert buffer_10m == [
        "buffer_10m_2021_Q1",
        "buffer_10m_2019_Q1",
        "buffer_10m_2017_Q1_2021_Q1",
    ]
    assert not any(filename.endswith("2017_Q1") for filename in additions + buffer_10m)


def test_qgis_temporal_styles_match_frontend_chronological_palette() -> None:
    fc = _feature_collection()
    project = TemporalProject(
        project_id="temporal-colors",
        name="Temporal colors",
        milestones=[
            TemporalMilestone(release_identifier="baseline", release_date="2017-03-01", status="complete"),
            *[
                TemporalMilestone(
                    release_identifier=release,
                    release_date=date_value,
                    status="complete",
                    additions_geojson=fc,
                    buffer_layers_geojson={"10m": fc},
                    cumulative_union_geojson=fc,
                )
                for release, date_value in (
                    ("2019", "2019-03-01"),
                    ("2021", "2021-03-01"),
                    ("2023", "2023-03-01"),
                    ("2025", "2025-03-01"),
                    ("2026", "2026-06-01"),
                )
            ],
        ],
        created_at="2026-01-01T00:00:00Z",
        updated_at="2026-01-01T00:00:00Z",
    )
    assert _temporal_milestone_color_map(project) == {
        "2019": "#00B050",
        "2021": "#FFD700",
        "2023": "#0066FF",
        "2025": "#E31A1C",
        "2026": "#00C8C8",
    }
    layers = _temporal_shapefile_export_layers(project)
    dated = {layer.release_identifier: _qgis_layer_style(project, layer)[0] for layer in layers if not layer.is_global}
    assert dated["2019"] in {"0,176,80,72", "0,176,80,150"}
    assert dated["2021"] in {"255,215,0,72", "255,215,0,150"}
    assert dated["2023"] in {"0,102,255,72", "0,102,255,150"}
    assert dated["2025"] in {"227,26,28,72", "227,26,28,150"}
    assert dated["2026"] in {"0,200,200,72", "0,200,200,150"}


def test_temporal_results_export_reuses_valid_cache(tmp_path: Path) -> None:
    settings = _save_project(tmp_path)
    first = build_temporal_results_export_file("temporal-export-formats-test", "geojson", settings=settings)
    first_mtime = first.stat().st_mtime_ns
    second = build_temporal_results_export_file("temporal-export-formats-test", "geojson", settings=settings)
    assert second == first
    assert second.stat().st_mtime_ns == first_mtime


def test_temporal_results_exports_resolve_file_backed_artifacts_for_all_formats_and_custom_scope(
    tmp_path: Path,
    caplog: pytest.LogCaptureFixture,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr("src.services.temporal_projects.TEMPORAL_VECTOR_TILE_METADATA_THRESHOLD_BYTES", 1)
    settings = _settings(tmp_path)
    project = TemporalProject(
        project_id="temporal-file-backed-export",
        name="File Backed Export",
        aoi_geojson=_feature_collection_from_bounds((-7.001, 32.999, -6.998, 33.002))["features"][0]["geometry"],
        milestones=[
            TemporalMilestone(
                release_identifier="WB_2018_R01",
                release_date="2018-01-15",
                status="complete",
                metrics=TemporalMilestoneMetrics(total_area_m2=1000.0),
            ),
            TemporalMilestone(
                release_identifier="WB_2020_R01",
                release_date="2020-01-15",
                status="complete",
                additions_geojson=None,
                buffer_layers_geojson={},
                metrics=TemporalMilestoneMetrics(added_area_m2=100.0, total_area_m2=1100.0, additions_feature_count=1),
            ),
            TemporalMilestone(
                release_identifier="WB_2022_R01",
                release_date="2022-01-15",
                status="complete",
                additions_geojson=None,
                buffer_layers_geojson={},
                metrics=TemporalMilestoneMetrics(added_area_m2=100.0, total_area_m2=1200.0, additions_feature_count=1),
            ),
        ],
        created_at="2026-05-19T00:00:00Z",
        updated_at="2026-05-19T00:00:00Z",
    )
    _save_project_payload(settings, project)

    artifact_payloads = {
        "WB_2020_R01": _feature_collection_from_bounds((-7.0000, 33.0000, -6.9998, 33.0002)),
        "WB_2022_R01": _feature_collection_from_bounds((-6.9996, 33.0000, -6.9994, 33.0002)),
    }
    buffer_payloads = {
        "WB_2020_R01": _feature_collection_from_bounds((-7.00000, 33.00000, -6.99970, 33.00030)),
        "WB_2022_R01": _feature_collection_from_bounds((-6.99960, 33.00000, -6.99930, 33.00030)),
    }
    for release, additions in artifact_payloads.items():
        milestone_dir = settings.temporal_projects_dir / project.project_id / "milestones" / release
        milestone_dir.mkdir(parents=True, exist_ok=True)
        (milestone_dir / "additions.geojson").write_text(json.dumps(additions), encoding="utf-8")
        for distance in ("10m", "15m", "20m"):
            (milestone_dir / f"building_change_buffer_{distance}.geojson").write_text(
                json.dumps(buffer_payloads[release]),
                encoding="utf-8",
            )

    caplog.set_level(logging.INFO)
    paths = {
        export_format: build_temporal_results_export_file(project.project_id, export_format, settings=settings)
        for export_format in ("xlsx", "kml", "geojson", "topojson", "json", "tsv")
    }
    assert all(path.is_file() and path.stat().st_size > 0 for path in paths.values())
    assert len(json.loads(paths["geojson"].read_text())["features"]) == 8
    assert json.loads(paths["json"].read_text())["milestones"][1]["layer_feature_counts"] == {
        "additions": 1,
        "buffer_10m": 1,
        "buffer_15m": 1,
        "buffer_20m": 1,
    }
    assert len(paths["tsv"].read_text().splitlines()) == 9

    shapefile_path = build_temporal_results_export_file(project.project_id, "shapefile", settings=settings)
    log_text = "\n".join(record.getMessage() for record in caplog.records)
    assert "EXPORT_FILE_BACKED_LAYER_SELECTED" in log_text
    assert "EXPORT_LAYER_MISSING" not in log_text
    with zipfile.ZipFile(shapefile_path) as archive:
        names = archive.namelist()
        assert "batiments_ajoutes_par_date/batiments_ajoutes_2022_Q1.shp" in names
        assert "batiments_ajoutes_par_date/batiments_ajoutes_2020_Q1.shp" in names
        assert "buffer_10m/buffer_10m_2022_Q1.shp" in names
        assert "buffer_10m/buffer_10m_2020_Q1.shp" in names

    perimeter = {
        "mode": "custom_geometry",
        "source": "imported",
        "geometry": {
            "type": "Polygon",
            "coordinates": [[
                [-7.0002, 32.9999],
                [-6.9997, 32.9999],
                [-6.9997, 33.0004],
                [-7.0002, 33.0004],
                [-7.0002, 32.9999],
            ]],
        },
    }
    first_custom = build_temporal_results_export_file(project.project_id, "geojson", settings=settings, perimeter=perimeter)
    first_custom_mtime = first_custom.stat().st_mtime_ns
    second_custom = build_temporal_results_export_file(project.project_id, "geojson", settings=settings, perimeter=perimeter)
    assert second_custom == first_custom
    assert second_custom.stat().st_mtime_ns == first_custom_mtime
    custom_payload = json.loads(second_custom.read_text())
    assert custom_payload["export_metadata"]["perimeter_source"] == "imported"
    assert custom_payload["features"]


def test_temporal_results_shapefile_invalidates_legacy_cache_without_version_metadata(tmp_path: Path) -> None:
    settings = _save_project(tmp_path)
    export_dir = settings.temporal_projects_dir / "temporal-export-formats-test" / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    legacy = export_dir / "results_shapefile.zip"
    legacy.write_bytes(b"legacy-three-layer-export")

    regenerated = build_temporal_results_export_file("temporal-export-formats-test", "shapefile", settings=settings)

    assert regenerated.read_bytes().startswith(b"PK")
    metadata = json.loads((export_dir / "results_shapefile.zip.metadata.json").read_text())
    assert metadata["version"] == "zone-clipped-mutually-exclusive-qgz-v13"


def test_temporal_results_qgz_has_valid_ids_extents_paths_groups_and_visibility(tmp_path: Path) -> None:
    settings = _save_project(tmp_path)
    path = build_temporal_results_export_file("temporal-export-formats-test", "shapefile", settings=settings)
    extract_root = tmp_path / "extracted"
    with zipfile.ZipFile(path) as archive:
        archive.extractall(extract_root)
    qgz_path = next(extract_root.glob("*.qgz"))
    assert qgz_path.stat().st_size > 1024
    with zipfile.ZipFile(qgz_path) as archive:
        qgs_names = [name for name in archive.namelist() if name.endswith(".qgs")]
        assert len(qgs_names) == 1
        root = ET.fromstring(archive.read(qgs_names[0]))

    maplayers = root.findall(".//projectlayers/maplayer")
    tree_layers = root.findall(".//layer-tree-layer")
    maplayer_ids = [layer.findtext("id") for layer in maplayers]
    tree_ids = [layer.attrib["id"] for layer in tree_layers]
    assert len(maplayer_ids) == len(set(maplayer_ids))
    assert set(maplayer_ids) == set(tree_ids)

    extents: dict[str, tuple[float, float, float, float]] = {}
    for layer in maplayers:
        layer_id = layer.findtext("id")
        datasource = layer.findtext("datasource") or ""
        assert datasource.startswith("./")
        assert ".." not in Path(datasource).parts
        assert not datasource.startswith("/")
        expected_crs = "EPSG:3857" if layer.findtext("provider") == "gdal" else "EPSG:4326"
        assert layer.findtext(".//authid") == expected_crs
        extent = layer.find("extent")
        assert extent is not None
        bounds = tuple(float(extent.findtext(key)) for key in ("xmin", "ymin", "xmax", "ymax"))
        assert all(math.isfinite(value) for value in bounds)
        assert bounds[0] <= bounds[2]
        assert bounds[1] <= bounds[3]
        extents[layer_id] = bounds

    canvas_extent = root.find(".//mapcanvas/extent")
    assert canvas_extent is not None
    canvas_bounds = tuple(float(canvas_extent.findtext(key)) for key in ("xmin", "ymin", "xmax", "ymax"))
    assert all(math.isfinite(value) for value in canvas_bounds)
    assert canvas_bounds[0] <= canvas_bounds[2]
    assert canvas_bounds[1] <= canvas_bounds[3]
    assert abs(canvas_bounds[0]) > 180
    assert abs(canvas_bounds[1]) > 90
    assert root.findtext(".//projectCrs/spatialrefsys/authid") == "EPSG:3857"
    root_tags = [child.tag for child in root]
    assert root_tags.index("projectCrs") < root_tags.index("layer-tree-group")
    assert root.findtext(".//mapcanvas/srs/spatialrefsys/authid") == "EPSG:3857"
    assert root.findtext(".//projectViewSettings/defaultViewExtent/srs/spatialrefsys/authid") == "EPSG:3857"
    assert root.findtext("projectionsEnabled") == "1"
    assert root.findtext("./properties/SpatialRefSys/ProjectionsEnabled") == "1"
    expected_project_bounds = _qgis_project_extent((-7.0, 33.0, -6.999, 33.001))
    assert canvas_bounds == pytest.approx(expected_project_bounds)

    for group in root.findall(".//layer-tree-group/layer-tree-group"):
        child_ids = [node.attrib["id"] for node in group.findall("layer-tree-layer")]
        if not child_ids:
            continue
        child_bounds = [extents[layer_id] for layer_id in child_ids if layer_id in extents]
        if not child_bounds:
            continue
        union = (
            min(bounds[0] for bounds in child_bounds),
            min(bounds[1] for bounds in child_bounds),
            max(bounds[2] for bounds in child_bounds),
            max(bounds[3] for bounds in child_bounds),
        )
        assert all(math.isfinite(value) for value in union)

    layer_order = [node.attrib["id"] for node in root.findall("./layerorder/layer")]
    raster_ids = [layer.findtext("id") for layer in maplayers if layer.findtext("provider") == "gdal"]
    vector_ids = [layer.findtext("id") for layer in maplayers if layer.findtext("provider") == "ogr"]
    assert layer_order == vector_ids + list(reversed(raster_ids))
    assert root.find(".//renderer-v2[@type='categorizedSymbol'][@attr='release_id']") is not None
    date_root = root.find(".//layer-tree-group[@name='Bâtiments ajoutés par date']")
    assert date_root is not None
    assert date_root.attrib["mutually-exclusive"] == "1"
    assert date_root.attrib["mutually-exclusive-child"] == "0"
    date_groups = {group.attrib["name"]: group for group in date_root.findall("layer-tree-group")}
    assert date_groups["2024 Q1"].attrib["checked"] == "Qt::Checked"
    assert date_groups["2022 Q1"].attrib["checked"] == "Qt::Unchecked"
    latest_children = date_groups["2024 Q1"].findall("layer-tree-layer")
    assert latest_children[0].attrib["name"] == "Bâtiments ajoutés 2024 Q1"
    assert "Imagerie de référence – 2024 Q1" not in {node.attrib["name"] for node in latest_children}
    assert {node.attrib["name"] for node in latest_children if node.attrib["checked"] == "Qt::Checked"} == {
        "Bâtiments ajoutés 2024 Q1",
        "Buffer 10m 2024 Q1",
    }
    synthesis = root.find(".//layer-tree-group[@name='Synthèse']")
    assert synthesis is not None
    synthesis_children = synthesis.findall("layer-tree-layer")
    synthesis_names = [node.attrib["name"] for node in synthesis_children]
    assert synthesis_names == [
        "Tous les nouveaux bâtiments 2022 Q1 → 2024 Q1",
        "Buffer 10m 2022 Q1 → 2024 Q1",
        "Buffer 15m 2022 Q1 → 2024 Q1",
        "Buffer 20m 2022 Q1 → 2024 Q1",
        "Imagerie de référence – 2024 Q1",
    ]
    assert {node.attrib["name"]: node.attrib["checked"] for node in synthesis_children} == {
        "Tous les nouveaux bâtiments 2022 Q1 → 2024 Q1": "Qt::Checked",
        "Buffer 10m 2022 Q1 → 2024 Q1": "Qt::Checked",
        "Buffer 15m 2022 Q1 → 2024 Q1": "Qt::Unchecked",
        "Buffer 20m 2022 Q1 → 2024 Q1": "Qt::Unchecked",
        "Imagerie de référence – 2024 Q1": "Qt::Checked",
    }


def test_temporal_results_qgz_synthesis_visibility_imagery_order_colors_and_dissolved_buffers(tmp_path: Path) -> None:
    settings = _settings(tmp_path)
    project = _three_milestone_project()
    _save_project_payload(settings, project)

    layers = _temporal_shapefile_export_layers(project)
    buffer_10m_layers = [layer for layer in layers if layer.group_key == "buffer_10m"]
    assert {layer.display_name: len(layer.feature_collection["features"]) for layer in buffer_10m_layers} == {
        "Buffer 10m 2022 Q1": 1,
        "Buffer 10m 2020 Q1": 1,
        "Buffer 10m 2018 Q1 → 2022 Q1": 2,
    }

    path = build_temporal_results_export_file(project.project_id, "shapefile", settings=settings)
    extract_root = tmp_path / "three-extracted"
    with zipfile.ZipFile(path) as archive:
        archive.extractall(extract_root)
    qgz_path = next(extract_root.glob("*.qgz"))
    with zipfile.ZipFile(qgz_path) as archive:
        root = ET.fromstring(archive.read(next(name for name in archive.namelist() if name.endswith(".qgs"))))

    synthesis = root.find(".//layer-tree-group[@name='Synthèse']")
    assert synthesis is not None
    synthesis_children = synthesis.findall("layer-tree-layer")
    synthesis_names = [node.attrib["name"] for node in synthesis_children]
    assert synthesis_names == [
        "Tous les nouveaux bâtiments 2018 Q1 → 2022 Q1",
        "Buffer 10m 2018 Q1 → 2022 Q1",
        "Buffer 15m 2018 Q1 → 2022 Q1",
        "Buffer 20m 2018 Q1 → 2022 Q1",
        "Imagerie de référence – 2022 Q1",
    ]
    assert {node.attrib["name"]: node.attrib["checked"] for node in synthesis_children} == {
        "Tous les nouveaux bâtiments 2018 Q1 → 2022 Q1": "Qt::Checked",
        "Buffer 10m 2018 Q1 → 2022 Q1": "Qt::Checked",
        "Buffer 15m 2018 Q1 → 2022 Q1": "Qt::Unchecked",
        "Buffer 20m 2018 Q1 → 2022 Q1": "Qt::Unchecked",
        "Imagerie de référence – 2022 Q1": "Qt::Checked",
    }
    assert root.find(".//layer-tree-group[@name='2022 Q1']/layer-tree-layer[@name='Imagerie de référence – 2022 Q1']") is None

    maplayers_by_name = {layer.attrib["name"]: layer for layer in root.findall(".//projectlayers/maplayer")}
    buffer_renderer = maplayers_by_name["Buffer 10m 2018 Q1 → 2022 Q1"].find("renderer-v2")
    assert buffer_renderer is not None
    assert buffer_renderer.attrib["type"] == "categorizedSymbol"
    assert buffer_renderer.attrib["attr"] == "release_id"
    buffer_qgs = ET.tostring(maplayers_by_name["Buffer 10m 2018 Q1 → 2022 Q1"], encoding="unicode")
    assert "0,176,80,72" in buffer_qgs
    assert "255,215,0,72" in buffer_qgs
    assert "245,158,11,112" not in buffer_qgs

    additions_qgs = ET.tostring(maplayers_by_name["Tous les nouveaux bâtiments 2018 Q1 → 2022 Q1"], encoding="unicode")
    assert "0,176,80,150" in additions_qgs
    assert "255,215,0,150" in additions_qgs

    per_milestone_2022 = gpd.read_file(extract_root / "buffer_10m" / "buffer_10m_2022_Q1.shp", engine="pyogrio")
    per_milestone_2020 = gpd.read_file(extract_root / "buffer_10m" / "buffer_10m_2020_Q1.shp", engine="pyogrio")
    synthesis_buffer = gpd.read_file(extract_root / "buffer_10m" / "buffer_10m_2018_Q1_2022_Q1.shp", engine="pyogrio")
    assert len(per_milestone_2022) == 1
    assert len(per_milestone_2020) == 1
    assert len(synthesis_buffer) == 2
    assert set(synthesis_buffer["release_id"]) == {"WB_2020_R01", "WB_2022_R01"}


def test_custom_perimeter_clips_geojson_and_adds_styled_qgis_export_zone(tmp_path: Path) -> None:
    settings = _save_project(tmp_path)
    custom_geometry = {
        "type": "Polygon",
        "coordinates": [[
            [-7.0005, 32.9995],
            [-6.9995, 32.9995],
            [-6.9995, 33.0005],
            [-7.0005, 33.0005],
            [-7.0005, 32.9995],
        ]],
    }
    perimeter = {"mode": "custom_geometry", "source": "drawn", "geometry": custom_geometry}

    geojson_path = build_temporal_results_export_file(
        "temporal-export-formats-test",
        "geojson",
        settings=settings,
        perimeter=perimeter,
    )
    geojson = json.loads(geojson_path.read_text())
    assert geojson["export_metadata"]["perimeter_mode"] == "custom_geometry"
    assert geojson["export_metadata"]["perimeter_source"] == "drawn"
    assert geojson["export_metadata"]["was_clipped_to_project_aoi"] is True
    assert geojson["features"]
    assert all(shape(feature["geometry"]).within(shape(_project().aoi_geojson)) for feature in geojson["features"])

    shapefile_path = build_temporal_results_export_file(
        "temporal-export-formats-test",
        "shapefile",
        settings=settings,
        perimeter=perimeter,
    )
    with zipfile.ZipFile(shapefile_path) as archive:
        assert "zone_export/zone_export.shp" in archive.namelist()
        qgz_name = next(name for name in archive.namelist() if name.endswith(".qgz"))
        qgz_path = tmp_path / "custom-results.qgz"
        qgz_path.write_bytes(archive.read(qgz_name))
    with zipfile.ZipFile(qgz_path) as qgz:
        qgs = qgz.read(next(name for name in qgz.namelist() if name.endswith(".qgs"))).decode("utf-8")
    root = ET.fromstring(qgs)
    tree_names = [node.attrib["name"] for node in root.findall(".//layer-tree-layer")]
    assert "OpenStreetMap" not in tree_names
    assert "Google Satellite" not in tree_names
    assert root.find(".//layer-tree-group[@name='Fond de carte en ligne']") is None
    assert "Zone d’export" in tree_names
    assert "0,176,80,150" in qgs
    assert 'type="categorizedSymbol" attr="release_id"' in qgs
    assert "245,158,11,36" in qgs
    assert "0,176,80,72" in qgs
    canvas_extent = root.find(".//mapcanvas/extent")
    assert canvas_extent is not None
    canvas_bounds = tuple(float(canvas_extent.findtext(key)) for key in ("xmin", "ymin", "xmax", "ymax"))
    expected_custom_bounds = _qgis_project_extent(tuple(shape(custom_geometry).intersection(shape(_project().aoi_geojson)).bounds))
    assert canvas_bounds == pytest.approx(expected_custom_bounds)
    assert abs(canvas_bounds[0]) > 180
    assert abs(canvas_bounds[1]) > 90
    assert root.findtext(".//projectCrs/spatialrefsys/authid") == "EPSG:3857"
    with zipfile.ZipFile(shapefile_path) as archive:
        raster_names = [name for name in archive.namelist() if name.endswith("reference_imagery_export_zone.tif")]
        assert len(raster_names) == 2
        for raster_name in raster_names:
            raster_path = tmp_path / Path(raster_name).name
            raster_path.write_bytes(archive.read(raster_name))
            with rasterio.open(raster_path) as dataset:
                assert dataset.width < 16
                assert dataset.height < 16
                assert dataset.crs.to_string() == "EPSG:3857"
                zone_bounds = transform_bounds("EPSG:4326", dataset.crs, *shape(custom_geometry).bounds)
                assert dataset.bounds.left <= zone_bounds[2]
                assert dataset.bounds.right >= zone_bounds[0]
                assert dataset.bounds.bottom <= zone_bounds[3]
                assert dataset.bounds.top >= zone_bounds[1]


def test_custom_perimeter_outside_project_aoi_is_rejected(tmp_path: Path) -> None:
    settings = _save_project(tmp_path)
    perimeter = {
        "mode": "custom_geometry",
        "source": "imported",
        "geometry": {
            "type": "Polygon",
            "coordinates": [[[1, 1], [2, 1], [2, 2], [1, 2], [1, 1]]],
        },
    }
    with pytest.raises(ValueError, match="hors de l’AOI"):
        build_temporal_results_export_file(
            "temporal-export-formats-test",
            "geojson",
            settings=settings,
            perimeter=perimeter,
        )


def test_temporal_results_topojson_is_clean_quantized_default(tmp_path: Path) -> None:
    settings = _save_project(tmp_path)
    path = build_temporal_results_export_file("temporal-export-formats-test", "topojson", settings=settings)
    payload = json.loads(path.read_text())

    assert payload["type"] == "Topology"
    assert isinstance(payload["bbox"], list)
    assert len(payload["bbox"]) == 4
    assert set(payload["transform"]) == {"scale", "translate"}
    assert payload["objects"]["results"]["type"] == "GeometryCollection"
    assert payload["arcs"]

    def assert_integer_arcs(value: object) -> None:
        if isinstance(value, list):
            for item in value:
                assert_integer_arcs(item)
        else:
            assert isinstance(value, int)

    assert_integer_arcs(payload["arcs"])

    geometries = payload["objects"]["results"]["geometries"]
    assert geometries
    allowed_property_keys = {"id", "project", "date", "year", "period", "layer", "area_m2", "area_ha"}
    removed_keys = {
        "run_id",
        "release_identifier",
        "release_t1",
        "release_t2",
        "src_date_t1",
        "src_date_t2",
        "source_backend",
        "feature_index",
        "buffer_id",
        "buffer_part_index",
        "source_change_block_id",
        "source_change_count",
        "block_gap_m",
        "cluster_gap_m",
        "kind",
        "release_date",
        "source_building_count",
        "confidence",
        "status",
        "score",
    }
    layers = {geometry["properties"]["layer"] for geometry in geometries}
    assert layers == {"additions", "buffer_10m"}
    assert "diagnostics" not in layers
    assert "buffer_15m" not in layers
    assert "buffer_20m" not in layers

    for geometry in geometries:
        properties = geometry["properties"]
        assert set(properties) == allowed_property_keys
        assert not (set(properties) & removed_keys)
        assert re.fullmatch(r"^[0-9]{4}-[a-z0-9-]+-[0-9]{6}$", properties["id"])
        assert properties["project"] == "Export Formats"
        assert properties["date"] == "2024-03-28"
        assert properties["year"] == 2024
        assert properties["layer"] in {"additions", "buffer_10m", "cumulative_growth"}
        assert properties["area_ha"] == round(properties["area_m2"] / 10000, 4)

    metadata_path = path.with_name("results.topojson.metadata.json")
    assert metadata_path.is_file()
    assert json.loads(metadata_path.read_text())["version"] == "clean-quantized-v3"


def test_temporal_results_export_route_headers_and_unsupported_format(tmp_path: Path) -> None:
    settings = _save_project(tmp_path)
    app.dependency_overrides[get_app_settings] = lambda: settings
    try:
        client = TestClient(app)
        response = client.get("/api/temporal-projects/temporal-export-formats-test/exports/results.geojson")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("application/geo+json")
        assert "resultats_temporal-export-formats-test.geojson" in response.headers["content-disposition"]
        assert "Content-Disposition" in response.headers["access-control-expose-headers"]
        assert int(response.headers["content-length"]) == len(response.content)

        tsv = client.get("/api/temporal-projects/temporal-export-formats-test/exports/results.tsv")
        assert tsv.status_code == 200
        assert tsv.headers["content-type"].startswith("text/tab-separated-values")

        shapefile = client.get("/api/temporal-projects/temporal-export-formats-test/exports/results_shapefile.zip")
        assert shapefile.status_code == 200
        assert shapefile.headers["content-type"].startswith("application/zip")
        assert "attachment" in shapefile.headers["content-disposition"]
        assert int(shapefile.headers["content-length"]) == len(shapefile.content)
        shapefile_metadata = json.loads(
            (
                settings.temporal_projects_dir
                / "temporal-export-formats-test"
                / "exports"
                / "results_shapefile.zip.metadata.json"
            ).read_text(encoding="utf-8")
        )
        assert shapefile_metadata["output"]["size_bytes"] == len(shapefile.content)

        xlsx = client.get("/api/temporal-projects/temporal-export-formats-test/exports/results.xlsx")
        assert xlsx.status_code == 200
        assert xlsx.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        workbook = load_workbook(BytesIO(xlsx.content), data_only=True)
        assert workbook.sheetnames == ["Synthèse", "Jalons", "Détails blocs"]
        assert all(sheet.sheet_view.zoomScale == 150 for sheet in workbook.worksheets)

        unsupported = client.get("/api/temporal-projects/temporal-export-formats-test/exports/results.parquet")
        assert unsupported.status_code == 400
        assert unsupported.json()["detail"]["code"] == "unsupported_export_format"

        full_topojson = client.get("/api/temporal-projects/temporal-export-formats-test/exports/results.topojson?mode=full")
        assert full_topojson.status_code == 400
        assert full_topojson.json()["detail"]["code"] == "unsupported_export_mode"

        custom = client.post(
            "/api/temporal-projects/temporal-export-formats-test/exports/results",
            json={
                "format": "geojson",
                "perimeter": {
                    "mode": "custom_geometry",
                    "source": "imported",
                    "geometry": {
                        "type": "Polygon",
                        "coordinates": [[
                            [-7.0005, 32.9995],
                            [-6.9995, 32.9995],
                            [-6.9995, 33.0005],
                            [-7.0005, 33.0005],
                            [-7.0005, 32.9995],
                        ]],
                    },
                },
            },
        )
        assert custom.status_code == 200
        assert custom.json()["export_metadata"]["perimeter_source"] == "imported"

        outside = client.post(
            "/api/temporal-projects/temporal-export-formats-test/exports/results",
            json={
                "format": "geojson",
                "perimeter": {
                    "mode": "custom_geometry",
                    "source": "drawn",
                    "geometry": {
                        "type": "Polygon",
                        "coordinates": [[[1, 1], [2, 1], [2, 2], [1, 2], [1, 1]]],
                    },
                },
            },
        )
        assert outside.status_code == 400
        assert outside.json()["detail"]["code"] == "invalid_export_perimeter"
    finally:
        app.dependency_overrides.clear()


def test_temporal_results_cached_export_fast_path_skips_project_hydration(monkeypatch, tmp_path: Path, caplog) -> None:
    caplog.set_level(logging.INFO, logger="src.services.temporal_exports")
    settings = _save_project(tmp_path)
    first_path = build_temporal_results_export_file(
        "temporal-export-formats-test",
        "shapefile",
        settings=settings,
    )

    monkeypatch.setattr(
        "src.services.temporal_exports._load_project",
        lambda *args, **kwargs: pytest.fail("valid cached project-AOI export should not hydrate project"),
    )
    second_path = build_temporal_results_export_file(
        "temporal-export-formats-test",
        "shapefile",
        settings=settings,
    )

    assert second_path == first_path
    assert "EXPORT_FAST_CACHE_HIT" in caplog.text
    assert "EXPORT_DOWNLOAD_TOTAL_MS" in caplog.text
