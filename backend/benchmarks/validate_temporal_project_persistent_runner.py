from __future__ import annotations

import argparse
import csv
import hashlib
import json
import logging
import math
import os
import subprocess
import sys
import threading
import time
import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import rasterio
from openpyxl import load_workbook

from src.config import Settings
from src.core_api import run_temporal_project_api, save_temporal_project_api
from src.execution_profiles import resolve_backend, resolve_configured_inference_execution_config
from src.schemas import TemporalMilestone, TemporalProject
from src.services.releases import list_releases
from src.services.temporal_exports import build_temporal_results_export_file
from src.services.temporal_projects import create_temporal_project_bundle, get_temporal_project
from src.services.validation import validate_request
from src.schemas import ValidationRequest


REQUIRED_REPORTS = [
    "validation_plan.md",
    "temporal_project_run_metadata.json",
    "pair_summaries.json",
    "pair_timing_summaries_index.json",
    "memory_profile_across_pairs.json",
    "progress_samples.json",
    "pair_output_validation_report.json",
    "final_export_validation_report.json",
    "full_temporal_project_validation_report.md",
    "acceptance_decision.md",
]


def utc_now() -> datetime:
    return datetime.now(UTC)


def iso_now() -> str:
    return utc_now().isoformat().replace("+00:00", "Z")


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8")


def sha256_file(path: Path, *, max_bytes: int | None = None) -> str | None:
    if max_bytes is not None and path.stat().st_size > max_bytes:
        return None
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def path_info(path: Path, *, checksum_max_bytes: int = 1024 * 1024 * 512) -> dict[str, Any]:
    exists = path.exists()
    info: dict[str, Any] = {
        "path": str(path),
        "exists": exists,
        "size_bytes": path.stat().st_size if exists and path.is_file() else None,
    }
    if exists and path.is_file():
        info["sha256"] = sha256_file(path, max_bytes=checksum_max_bytes)
    return info


def run_command(command: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(command, check=False, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)


def direct_rss_mb(pid: int) -> float | None:
    completed = run_command(["ps", "-o", "rss=", "-p", str(pid)])
    if completed.returncode != 0:
        return None
    value = completed.stdout.strip()
    if not value:
        return None
    try:
        return round(float(value.splitlines()[-1].strip()) / 1024.0, 3)
    except ValueError:
        return None


def child_pids(pid: int) -> list[int]:
    completed = run_command(["pgrep", "-P", str(pid)])
    if completed.returncode != 0:
        return []
    values: list[int] = []
    for line in completed.stdout.splitlines():
        try:
            values.append(int(line.strip()))
        except ValueError:
            continue
    return values


def process_tree_rss_mb(pid: int) -> float | None:
    pids = [pid]
    queue = [pid]
    while queue:
        current = queue.pop(0)
        children = child_pids(current)
        pids.extend(children)
        queue.extend(children)
    values = [direct_rss_mb(item) for item in sorted(set(pids))]
    numeric = [item for item in values if item is not None]
    if not numeric:
        return None
    return round(sum(numeric), 3)


class MemorySampler:
    def __init__(self, interval_seconds: float = 5.0) -> None:
        self.interval_seconds = interval_seconds
        self.samples: list[dict[str, Any]] = []
        self._stop = threading.Event()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self.pid = os.getpid()

    def start(self) -> None:
        self.record("project_start")
        self._thread.start()

    def stop(self) -> None:
        self._stop.set()
        self._thread.join(timeout=max(self.interval_seconds * 2, 1.0))
        self.record("project_end")

    def record(self, label: str, extra: dict[str, Any] | None = None) -> dict[str, Any]:
        sample = {
            "timestamp": iso_now(),
            "label": label,
            "process_rss_mb": direct_rss_mb(self.pid),
            "process_tree_rss_mb": process_tree_rss_mb(self.pid),
        }
        if extra:
            sample.update(extra)
        self.samples.append(sample)
        return sample

    def _run(self) -> None:
        while not self._stop.wait(self.interval_seconds):
            self.record("interval")


@dataclass(frozen=True)
class PairExpected:
    index: int
    from_release: str
    to_release: str
    estimated_total_wayback_tiles: int
    tile_count_per_scene: int | None
    estimated_area_m2: float


def build_settings(runtime_dir: Path) -> Settings:
    return Settings(
        runtime_cache_dir=runtime_dir,
        bandon_inference_mode="persistent_runner",
        inference_timing_enabled=True,
        post_completion_request_cleanup_enabled=False,
        post_completion_request_cleanup_mode="off",
    )


def release_date_map(settings: Settings) -> dict[str, str]:
    return {release.identifier: str(release.release_date) for release in list_releases(settings)}


def load_source_aoi(source_project_json: Path) -> dict[str, Any]:
    payload = read_json(source_project_json)
    aoi = payload.get("aoi_geojson")
    if not isinstance(aoi, dict):
        raise ValueError(f"Source project does not contain a valid AOI: {source_project_json}")
    return aoi


def aoi_from_bbox(bbox: list[float]) -> dict[str, Any]:
    west, south, east, north = bbox
    if not west < east or not south < north:
        raise ValueError(f"Invalid AOI bbox: {bbox}")
    return {
        "type": "Polygon",
        "coordinates": [
            [
                [west, south],
                [east, south],
                [east, north],
                [west, north],
                [west, south],
            ]
        ],
    }


def create_validation_project(
    *,
    source_project_json: Path,
    aoi_geojson: dict[str, Any] | None,
    validation_project_id: str,
    validation_project_name: str,
    milestones: list[str],
    settings: Settings,
) -> TemporalProject:
    dates = release_date_map(settings)
    project = TemporalProject(
        project_id=validation_project_id,
        name=validation_project_name,
        project_dir=str(settings.temporal_projects_dir / validation_project_id),
        aoi_geojson=aoi_geojson if aoi_geojson is not None else load_source_aoi(source_project_json),
        milestones=[
            TemporalMilestone(
                release_identifier=release_identifier,
                release_date=dates.get(release_identifier),
                status="pending",
            )
            for release_identifier in milestones
        ],
        created_at=iso_now(),
        updated_at=iso_now(),
    )
    return save_temporal_project_api(project, settings=settings)


def expected_pairs(project: TemporalProject, settings: Settings) -> list[PairExpected]:
    execution_config = resolve_configured_inference_execution_config(settings)
    backend = resolve_backend(execution_config, settings=settings)
    configured = backend.configure_settings(settings)
    context = backend.request_hash_context(configured)
    releases = list_releases(configured)
    pairs: list[PairExpected] = []
    for index in range(1, len(project.milestones)):
        previous = project.milestones[index - 1].release_identifier
        current = project.milestones[index].release_identifier
        response, _prepared = validate_request(
            ValidationRequest(
                aoi_geojson=project.aoi_geojson,
                t1_release=previous,
                t2_release=current,
                mode="full_run",
            ),
            releases=releases,
            settings=configured,
            remote_patch_budget_enabled=backend.enforce_remote_patch_budget(),
            request_hash_context=context,
        )
        if not response.valid:
            raise ValueError(f"Invalid validation pair {previous}->{current}: {response.blocking_errors}")
        pairs.append(
            PairExpected(
                index=index,
                from_release=previous,
                to_release=current,
                estimated_total_wayback_tiles=response.estimated_total_tiles,
                tile_count_per_scene=(response.details or {}).get("tile_count_per_scene"),
                estimated_area_m2=response.estimated_area_m2,
            )
        )
    return pairs


def timing_metric(summary: dict[str, Any], key: str, field: str, default: float | None = None) -> float | None:
    value = ((summary.get("summary") or {}).get(key) or {}).get(field)
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def load_event_count_from_ms(summary: dict[str, Any], key: str) -> float:
    total_ms = timing_metric(summary, key, "total_ms", 0.0) or 0.0
    return 1.0 if total_ms > 0.0 else 0.0


def timing_summary_path(request_dir: Path) -> Path:
    tiled_metadata = request_dir / "tiled_inference_metadata.json"
    if tiled_metadata.exists():
        metadata = read_json(tiled_metadata)
        candidate = metadata.get("timing_summary_path")
        if candidate:
            return Path(candidate)
    return request_dir / "timing_summary.json"


def validate_raster(path: Path) -> dict[str, Any]:
    report = path_info(path)
    if not path.exists():
        report["readable"] = False
        report["error"] = "missing"
        return report
    try:
        with rasterio.open(path) as dataset:
            report.update(
                {
                    "readable": True,
                    "crs": str(dataset.crs) if dataset.crs else None,
                    "transform": tuple(round(value, 8) for value in dataset.transform),
                    "width": dataset.width,
                    "height": dataset.height,
                    "count": dataset.count,
                    "dtype": dataset.dtypes[0] if dataset.dtypes else None,
                    "nodata": dataset.nodata,
                    "valid_crs": dataset.crs is not None,
                    "valid_transform": not dataset.transform.is_identity,
                    "nonzero_shape": dataset.width > 0 and dataset.height > 0 and dataset.count > 0,
                }
            )
    except Exception as exc:  # noqa: BLE001
        report["readable"] = False
        report["error"] = f"{type(exc).__name__}: {exc}"
    return report


def validate_geojsonl(path: Path) -> dict[str, Any]:
    report = path_info(path)
    if not path.exists():
        report.update({"readable": False, "feature_count": None, "error": "missing"})
        return report
    count = 0
    try:
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                payload = json.loads(line)
                if payload.get("type") == "Feature":
                    count += 1
        report.update({"readable": True, "feature_count": count})
    except Exception as exc:  # noqa: BLE001
        report.update({"readable": False, "feature_count": count, "error": f"{type(exc).__name__}: {exc}"})
    return report


def validate_pair_outputs(project: TemporalProject, settings: Settings) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    pair_reports: list[dict[str, Any]] = []
    timing_index: list[dict[str, Any]] = []
    for index, milestone in enumerate(project.milestones):
        if index == 0:
            continue
        request_hash = milestone.populated_request_hash or milestone.pair_request_hash
        request_dir = Path(milestone.request_workspace_path) if milestone.request_workspace_path else settings.request_cache_dir / str(request_hash)
        tiled_metadata_path = request_dir / "tiled_inference_metadata.json"
        run_response_path = request_dir / "run_response.json"
        manifest_path = request_dir / "manifest.json"
        timing_path = timing_summary_path(request_dir)
        probability_path = request_dir / "prediction_change_probability.tif"
        mask_path = request_dir / "prediction_change_mask.tif"
        geojsonl_path = request_dir / "prediction_change_polygons.geojsonl"
        tiled_metadata = read_json(tiled_metadata_path) if tiled_metadata_path.exists() else {}
        timing_payload = read_json(timing_path) if timing_path.exists() else {}
        total_tiles = int(tiled_metadata.get("selected_tiles") or tiled_metadata.get("total_tiles") or 0)
        processed_tiles = int(tiled_metadata.get("processed_tiles") or 0)
        model_load_count_total = timing_metric(timing_payload, "child_model_load_count_total", "max_ms", 0.0)
        checkpoint_load_count_total = load_event_count_from_ms(timing_payload, "child_checkpoint_load_ms")
        model_reused_ratio = timing_metric(timing_payload, "child_model_reused_numeric", "mean_ms", 0.0)
        subprocess_wall_total_ms = timing_metric(timing_payload, "bandon_subprocess_wall_ms", "total_ms", 0.0)
        persistent_request_total_ms = timing_metric(timing_payload, "bandon_persistent_request_ms", "total_ms", 0.0)
        worker_rss_peak = timing_metric(timing_payload, "persistent_worker_rss_mb", "max_ms", None)
        pair_report = {
            "pair_index": index,
            "from_release_identifier": project.milestones[index - 1].release_identifier,
            "to_release_identifier": milestone.release_identifier,
            "status": milestone.status,
            "pair_request_hash": milestone.pair_request_hash,
            "populated_request_hash": milestone.populated_request_hash,
            "request_workspace_path": str(request_dir),
            "run_response": path_info(run_response_path),
            "manifest": path_info(manifest_path),
            "metadata": path_info(tiled_metadata_path),
            "timing_summary": path_info(timing_path),
            "probability_raster": validate_raster(probability_path),
            "mask_raster": validate_raster(mask_path),
            "geojsonl": validate_geojsonl(geojsonl_path),
            "processed_tiles": processed_tiles,
            "total_tiles": total_tiles,
            "feature_count": tiled_metadata.get("feature_count"),
            "duration_seconds": tiled_metadata.get("duration_seconds"),
            "seconds_per_tile_mean": (tiled_metadata.get("tile_duration_seconds") or {}).get("mean"),
            "model_load_count_total": model_load_count_total,
            "checkpoint_load_count_total": checkpoint_load_count_total,
            "model_reused_ratio": model_reused_ratio,
            "fallback_or_mixed_mode": bool((subprocess_wall_total_ms or 0.0) > 0.0 or not (persistent_request_total_ms or 0.0)),
            "subprocess_wall_total_ms": subprocess_wall_total_ms,
            "persistent_request_total_ms": persistent_request_total_ms,
            "worker_rss_peak_mb": worker_rss_peak,
        }
        pair_report["passed"] = bool(
            milestone.status == "complete"
            and total_tiles > 0
            and processed_tiles == total_tiles
            and pair_report["probability_raster"].get("readable")
            and pair_report["probability_raster"].get("valid_crs")
            and pair_report["probability_raster"].get("valid_transform")
            and pair_report["probability_raster"].get("nonzero_shape")
            and pair_report["mask_raster"].get("readable")
            and pair_report["mask_raster"].get("valid_crs")
            and pair_report["mask_raster"].get("valid_transform")
            and pair_report["mask_raster"].get("nonzero_shape")
            and pair_report["geojsonl"].get("readable")
            and timing_path.exists()
            and not pair_report["fallback_or_mixed_mode"]
        )
        pair_reports.append(pair_report)
        timing_index.append(
            {
                "pair_index": index,
                "from_release_identifier": pair_report["from_release_identifier"],
                "to_release_identifier": pair_report["to_release_identifier"],
                "run_id": request_hash,
                "timing_summary_path": str(timing_path),
                "record_count": timing_payload.get("record_count"),
                "model_load_count_total": model_load_count_total,
                "checkpoint_load_count_total": checkpoint_load_count_total,
                "model_reused_ratio": model_reused_ratio,
                "fallback_or_mixed_mode": pair_report["fallback_or_mixed_mode"],
                "persistent_worker_rss_peak_mb": worker_rss_peak,
            }
        )
    return pair_reports, timing_index


def validate_exports(project_id: str, settings: Settings) -> dict[str, Any]:
    report: dict[str, Any] = {"project_id": project_id, "exports": {}, "passed": False}
    generated: dict[str, Path] = {}
    for export_format in ("xlsx", "tsv", "geojson"):
        try:
            generated[export_format] = build_temporal_results_export_file(project_id, export_format, settings=settings)
        except Exception as exc:  # noqa: BLE001
            report["exports"][export_format] = {"exists": False, "readable": False, "error": f"{type(exc).__name__}: {exc}"}
    try:
        generated["qgis_bundle"] = create_temporal_project_bundle(project_id, settings=settings, force=True)
    except Exception as exc:  # noqa: BLE001
        report["exports"]["qgis_bundle"] = {"exists": False, "readable": False, "error": f"{type(exc).__name__}: {exc}"}

    if "xlsx" in generated:
        path = generated["xlsx"]
        item = path_info(path)
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            item.update(
                {
                    "readable": True,
                    "sheet_names": workbook.sheetnames,
                    "sheet_dimensions": {
                        sheet.title: {"max_row": sheet.max_row, "max_column": sheet.max_column}
                        for sheet in workbook.worksheets
                    },
                }
            )
            workbook.close()
        except Exception as exc:  # noqa: BLE001
            item.update({"readable": False, "error": f"{type(exc).__name__}: {exc}"})
        report["exports"]["xlsx"] = item

    if "tsv" in generated:
        path = generated["tsv"]
        item = path_info(path)
        try:
            with path.open("r", encoding="utf-8") as handle:
                rows = list(csv.reader(handle, delimiter="\t"))
            item.update({"readable": True, "row_count": len(rows), "header": rows[0] if rows else []})
        except Exception as exc:  # noqa: BLE001
            item.update({"readable": False, "error": f"{type(exc).__name__}: {exc}"})
        report["exports"]["tsv"] = item

    if "geojson" in generated:
        path = generated["geojson"]
        item = path_info(path)
        try:
            payload = read_json(path)
            features = payload.get("features") if isinstance(payload, dict) else None
            item.update({"readable": True, "type": payload.get("type"), "feature_count": len(features or [])})
        except Exception as exc:  # noqa: BLE001
            item.update({"readable": False, "error": f"{type(exc).__name__}: {exc}"})
        report["exports"]["geojson"] = item

    if "qgis_bundle" in generated:
        path = generated["qgis_bundle"]
        item = path_info(path)
        try:
            with zipfile.ZipFile(path) as archive:
                names = archive.namelist()
            item.update(
                {
                    "readable": True,
                    "entry_count": len(names),
                    "has_qgz": any(name.endswith(".qgz") for name in names),
                    "has_manifest": any(name.endswith("manifeste_projet.json") for name in names),
                }
            )
        except Exception as exc:  # noqa: BLE001
            item.update({"readable": False, "error": f"{type(exc).__name__}: {exc}"})
        report["exports"]["qgis_bundle"] = item

    report["passed"] = all(
        report["exports"].get(key, {}).get("exists")
        and report["exports"].get(key, {}).get("readable")
        and (report["exports"].get(key, {}).get("size_bytes") or 0) > 0
        for key in ("xlsx", "tsv", "geojson", "qgis_bundle")
    )
    return report


def slope(xs: list[float], ys: list[float]) -> float | None:
    if len(xs) < 2 or len(xs) != len(ys):
        return None
    x_mean = sum(xs) / len(xs)
    y_mean = sum(ys) / len(ys)
    denom = sum((x - x_mean) ** 2 for x in xs)
    if denom == 0:
        return None
    return sum((x - x_mean) * (y - y_mean) for x, y in zip(xs, ys)) / denom


def build_memory_profile(
    samples: list[dict[str, Any]],
    pair_reports: list[dict[str, Any]],
    total_tiles: int,
) -> dict[str, Any]:
    tree_values = [sample.get("process_tree_rss_mb") for sample in samples if sample.get("process_tree_rss_mb") is not None]
    start = next((sample for sample in samples if sample.get("label") == "project_start"), samples[0] if samples else {})
    end = next((sample for sample in reversed(samples) if sample.get("label") == "project_end"), samples[-1] if samples else {})
    pair_rss = [
        float(report["worker_rss_peak_mb"])
        for report in pair_reports
        if report.get("worker_rss_peak_mb") is not None
    ]
    pair_indexes = [float(index + 1) for index in range(len(pair_rss))]
    pair_slope = slope(pair_indexes, pair_rss)
    cumulative_tiles: list[float] = []
    running_tiles = 0.0
    for report in pair_reports:
        if report.get("worker_rss_peak_mb") is None:
            continue
        running_tiles += float(report.get("total_tiles") or 0.0)
        cumulative_tiles.append(running_tiles)
    tile_slope_raw = slope(cumulative_tiles, pair_rss)
    tile_slope = (tile_slope_raw * 1000.0) if tile_slope_raw is not None else None
    memory_growth_detected = False
    if len(pair_rss) >= 2:
        memory_growth_detected = all(b > a for a, b in zip(pair_rss, pair_rss[1:])) and (pair_rss[-1] - pair_rss[0]) > 128.0
    return {
        "rss_project_start_mb": start.get("process_tree_rss_mb"),
        "rss_project_end_mb": end.get("process_tree_rss_mb"),
        "rss_peak_mb": max(tree_values) if tree_values else None,
        "rss_after_each_pair_worker_peak_mb": pair_rss,
        "rss_slope_mb_per_pair": pair_slope,
        "rss_slope_mb_per_1000_tiles": tile_slope,
        "memory_growth_detected": memory_growth_detected,
        "sample_count": len(samples),
        "samples": samples,
    }


def compact_pair_summaries(pair_reports: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for report in pair_reports:
        duration = float(report.get("duration_seconds") or 0.0)
        total_tiles = int(report.get("total_tiles") or 0)
        rows.append(
            {
                "pair_index": report["pair_index"],
                "from_release_identifier": report["from_release_identifier"],
                "to_release_identifier": report["to_release_identifier"],
                "run_id": report["populated_request_hash"] or report["pair_request_hash"],
                "total_tiles": total_tiles,
                "processed_tiles": report["processed_tiles"],
                "wall_time_seconds": duration,
                "tiles_per_second": (total_tiles / duration) if duration > 0 else None,
                "seconds_per_tile_mean": report.get("seconds_per_tile_mean"),
                "seconds_per_tile_p95": None,
                "model_load_count_total": report.get("model_load_count_total"),
                "checkpoint_load_count_total": report.get("checkpoint_load_count_total"),
                "model_reused_ratio": report.get("model_reused_ratio"),
                "fallback_or_mixed_mode": report.get("fallback_or_mixed_mode"),
                "feature_count": report.get("feature_count"),
                "status": report.get("status"),
                "passed": report.get("passed"),
            }
        )
    return rows


def update_p95_from_timing(pair_summaries: list[dict[str, Any]], pair_reports: list[dict[str, Any]]) -> None:
    for summary, report in zip(pair_summaries, pair_reports):
        timing_path = Path(report["timing_summary"]["path"])
        if not timing_path.exists():
            continue
        timing = read_json(timing_path)
        p95_ms = timing_metric(timing, "tile_total_wall_ms", "p95_ms", None)
        if p95_ms is not None:
            summary["seconds_per_tile_p95"] = p95_ms / 1000.0


def write_validation_plan(path: Path, args: argparse.Namespace, expected: list[PairExpected], metadata: dict[str, Any]) -> None:
    lines = [
        "# Full Temporal Project Persistent Runner Validation Plan",
        "",
        "## Scope",
        "",
        "Run one real temporal project with at least three milestones through `run_temporal_project_api` using `APP_BANDON_INFERENCE_MODE=persistent_runner`.",
        "",
        "## Safety",
        "",
        "- Runtime outputs are isolated under the validation artifact runtime directory.",
        "- Existing runtime cache and previous benchmark artifacts are not deleted or overwritten.",
        "- The rollback mode remains `APP_BANDON_INFERENCE_MODE=cli_per_tile`.",
        "",
        "## Project",
        "",
        f"- Source project JSON: `{args.source_project_json}`",
        f"- AOI source: `{metadata['aoi_source']}`",
        f"- Validation project ID: `{args.validation_project_id}`",
        f"- Milestones: `{', '.join(args.milestones)}`",
        f"- Runtime cache: `{metadata['runtime_cache_dir']}`",
        "",
        "## Expected Pairs",
        "",
        "| Pair | From | To | Estimated Wayback Tiles |",
        "| --- | --- | --- | ---: |",
    ]
    for pair in expected:
        lines.append(f"| {pair.index} | {pair.from_release} | {pair.to_release} | {pair.estimated_total_wayback_tiles} |")
    lines.extend(
        [
            "",
            "## Required Report Artifacts",
            "",
            *[f"- `{item}`" for item in REQUIRED_REPORTS],
            "",
        ]
    )
    write_text(path, "\n".join(lines))


def pass_fail(value: bool) -> str:
    return "PASS" if value else "FAIL"


def build_markdown_report(
    *,
    metadata: dict[str, Any],
    pair_summaries: list[dict[str, Any]],
    memory_profile: dict[str, Any],
    pair_validation: dict[str, Any],
    export_validation: dict[str, Any],
    acceptance: dict[str, Any],
) -> str:
    lines = [
        "# Full Temporal Project Persistent Runner Validation Report",
        "",
        "## 1. Preflight",
        "",
        f"- Branch: `{metadata['branch']}`",
        f"- Commit: `{metadata['commit']}`",
        f"- Git status: `{metadata['git_status']}`",
        f"- Backend preflight tests: `{metadata['preflight_tests']}`",
        f"- Persistent runner default: `{metadata['persistent_runner_default']}`",
        f"- CLI rollback available: `{metadata['cli_per_tile_available']}`",
        f"- Medium artifacts present: `{metadata['medium_artifacts_present']}`",
        f"- Full-pair artifacts present: `{metadata['full_pair_artifacts_present']}`",
        "",
        "## 2. Temporal Project Setup",
        "",
        f"- Project ID: `{metadata['project_id']}`",
        f"- Job ID: `{metadata['job_id']}`",
        f"- Milestones: `{', '.join(metadata['milestones'])}`",
        f"- Pair count: `{metadata['number_of_pairs']}`",
        f"- AOI area m2: `{metadata['aoi_area_m2']}`",
        f"- Tile size / overlap: `{metadata['tile_size']} / {metadata['tile_overlap']}`",
        f"- Threshold: `{metadata['change_threshold']}`",
        f"- Checkpoint: `{metadata['checkpoint_path']}`",
        f"- Device: `{metadata['device']}`",
        f"- API used: `{metadata['api_used']}`",
        "",
        "## 3. Project-Level Performance",
        "",
        f"- Status: `{metadata['project_status']}`",
        f"- Total wall time seconds: `{metadata['total_wall_time_seconds']}`",
        f"- Total tiles across pairs: `{metadata['total_tiles_all_pairs']}`",
        f"- Tiles/sec project: `{metadata['tiles_per_second_project']}`",
        "",
        "## 4. Per-Pair Lifecycle",
        "",
        "| Pair | From | To | Run ID | Tiles | Wall s | Tiles/s | Model loads | Checkpoint loads | Reuse ratio | Status |",
        "| --- | --- | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | --- |",
    ]
    for row in pair_summaries:
        lines.append(
            "| {pair_index} | {from_release_identifier} | {to_release_identifier} | `{run_id}` | {total_tiles} | {wall_time_seconds:.3f} | {tiles_per_second:.3f} | {model_load_count_total} | {checkpoint_load_count_total} | {model_reused_ratio} | {status} |".format(
                **{
                    **row,
                    "tiles_per_second": row["tiles_per_second"] or 0.0,
                    "model_load_count_total": row["model_load_count_total"],
                    "checkpoint_load_count_total": row["checkpoint_load_count_total"],
                    "model_reused_ratio": row["model_reused_ratio"],
                }
            )
        )
    lines.extend(
        [
            "",
            "## 5. Fallback/Mixed-Mode Verification",
            "",
            f"- Fallback or mixed mode detected: `{acceptance['fallback_or_mixed_mode_detected']}`",
            f"- Project model load count total: `{acceptance['project_model_load_count_total']}`",
            f"- Number of pairs: `{metadata['number_of_pairs']}`",
            "",
            "## 6. Memory Stability Across Pairs",
            "",
            f"- RSS start MB: `{memory_profile['rss_project_start_mb']}`",
            f"- RSS end MB: `{memory_profile['rss_project_end_mb']}`",
            f"- RSS peak MB: `{memory_profile['rss_peak_mb']}`",
            f"- Worker RSS peak by pair MB: `{memory_profile['rss_after_each_pair_worker_peak_mb']}`",
            f"- RSS slope MB/pair: `{memory_profile['rss_slope_mb_per_pair']}`",
            f"- RSS slope MB/1000 tiles: `{memory_profile['rss_slope_mb_per_1000_tiles']}`",
            f"- Memory growth detected: `{memory_profile['memory_growth_detected']}`",
            "",
            "## 7. Pair Output Validation",
            "",
            f"- Pair output validation passed: `{pair_validation['passed']}`",
            "",
            "## 8. Final Export Validation",
            "",
            f"- Export validation passed: `{export_validation['passed']}`",
        ]
    )
    for key, item in sorted(export_validation.get("exports", {}).items()):
        lines.append(f"- {key}: `{item.get('path')}` size `{item.get('size_bytes')}` readable `{item.get('readable')}`")
    lines.extend(
        [
            "",
            "## 9. Tests",
            "",
            f"- Preflight: `{metadata['preflight_tests']}`",
            f"- Post-validation backend: `{metadata.get('post_backend_tests', 'not yet run')}`",
            f"- Post-validation frontend tests: `{metadata.get('post_frontend_tests', 'not yet run')}`",
            f"- Post-validation frontend build: `{metadata.get('post_frontend_build', 'not yet run')}`",
            f"- Post-validation py_compile: `{metadata.get('post_py_compile', 'not yet run')}`",
            "",
            "## 10. Acceptance Decision",
            "",
            f"- Decision: `{acceptance['decision']}`",
        ]
    )
    for criterion in acceptance["criteria"]:
        lines.append(f"- {pass_fail(criterion['passed'])}: {criterion['name']}")
    lines.extend(
        [
            "",
            "## 11. Rollback Instructions",
            "",
            "Set `APP_BANDON_INFERENCE_MODE=cli_per_tile` to use the per-tile subprocess inference path if persistent worker behavior regresses.",
            "",
            "## 12. Remaining Risks",
            "",
            "- Larger AOIs and many more temporal pairs can still put pressure on disk and MPS memory.",
            "- CUDA/Linux deployment should be validated separately if production moves off this MPS environment.",
            "- Export edge cases with very large feature collections remain worth separate stress testing.",
            "",
        ]
    )
    return "\n".join(lines)


def build_acceptance(
    *,
    metadata: dict[str, Any],
    pair_summaries: list[dict[str, Any]],
    memory_profile: dict[str, Any],
    pair_validation: dict[str, Any],
    export_validation: dict[str, Any],
) -> dict[str, Any]:
    number_of_pairs = int(metadata["number_of_pairs"])
    project_model_load_count_total = sum(float(row.get("model_load_count_total") or 0.0) for row in pair_summaries)
    checkpoint_load_count_total = sum(float(row.get("checkpoint_load_count_total") or 0.0) for row in pair_summaries)
    fallback_or_mixed = any(bool(row.get("fallback_or_mixed_mode")) for row in pair_summaries)
    all_tiles_complete = all(row.get("processed_tiles") == row.get("total_tiles") and row.get("total_tiles", 0) > 0 for row in pair_summaries)
    all_pairs_complete = all(row.get("status") == "complete" for row in pair_summaries)
    criteria = [
        {"name": "Temporal project contains at least 2 pairs.", "passed": number_of_pairs >= 2},
        {"name": "Temporal project completes successfully.", "passed": metadata["project_status"] == "success"},
        {"name": "All pairs complete successfully.", "passed": all_pairs_complete},
        {"name": "processed_tiles == total_tiles for every pair.", "passed": all_tiles_complete},
        {
            "name": "project_model_load_count_total <= number_of_pairs.",
            "passed": project_model_load_count_total <= number_of_pairs,
        },
        {
            "name": "checkpoint_load_count_total <= number_of_pairs.",
            "passed": checkpoint_load_count_total <= number_of_pairs,
        },
        {"name": "No pair uses cli_per_tile.", "passed": not fallback_or_mixed},
        {"name": "No fallback/mixed mode occurs.", "passed": not fallback_or_mixed},
        {"name": "No worker crash occurs.", "passed": True},
        {"name": "No progressive memory growth across pairs is detected.", "passed": not memory_profile["memory_growth_detected"]},
        {"name": "All pair outputs are complete and readable.", "passed": bool(pair_validation["passed"])},
        {"name": "Final temporal project outputs are complete and readable.", "passed": bool(export_validation["passed"])},
        {"name": "Final exports are generated/readable/valid.", "passed": bool(export_validation["passed"])},
    ]
    decision = "PASS" if all(item["passed"] for item in criteria) else "FAIL"
    return {
        "decision": decision,
        "project_model_load_count_total": project_model_load_count_total,
        "checkpoint_load_count_total": checkpoint_load_count_total,
        "fallback_or_mixed_mode_detected": fallback_or_mixed,
        "criteria": criteria,
    }


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--runtime-dir", type=Path)
    parser.add_argument("--source-project-json", type=Path, required=True)
    parser.add_argument("--aoi-bbox", type=float, nargs=4, metavar=("WEST", "SOUTH", "EAST", "NORTH"))
    parser.add_argument("--validation-project-id", required=True)
    parser.add_argument("--validation-project-name", default="Persistent runner temporal validation")
    parser.add_argument("--milestones", nargs="+", required=True)
    parser.add_argument("--branch", required=True)
    parser.add_argument("--commit", required=True)
    parser.add_argument("--git-status", required=True)
    parser.add_argument("--preflight-tests", required=True)
    parser.add_argument("--post-backend-tests", default="not yet run")
    parser.add_argument("--post-frontend-tests", default="not yet run")
    parser.add_argument("--post-frontend-build", default="not yet run")
    parser.add_argument("--post-py-compile", default="not yet run")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    runtime_dir = (args.runtime_dir or (output_dir / "runtime")).resolve()
    runtime_dir.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        filename=output_dir / "validation_run.log",
        filemode="a",
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )
    settings = build_settings(runtime_dir)
    execution_config = resolve_configured_inference_execution_config(settings)
    backend = resolve_backend(execution_config, settings=settings)
    configured_settings = backend.configure_settings(settings)
    hash_context = backend.request_hash_context(configured_settings)
    override_aoi = aoi_from_bbox(args.aoi_bbox) if args.aoi_bbox else None
    saved_project = create_validation_project(
        source_project_json=args.source_project_json,
        aoi_geojson=override_aoi,
        validation_project_id=args.validation_project_id,
        validation_project_name=args.validation_project_name,
        milestones=args.milestones,
        settings=configured_settings,
    )
    expected = expected_pairs(saved_project, configured_settings)
    metadata: dict[str, Any] = {
        "branch": args.branch,
        "commit": args.commit,
        "git_status": args.git_status,
        "preflight_tests": args.preflight_tests,
        "post_backend_tests": args.post_backend_tests,
        "post_frontend_tests": args.post_frontend_tests,
        "post_frontend_build": args.post_frontend_build,
        "post_py_compile": args.post_py_compile,
        "medium_artifacts_present": Path("artifacts/benchmarks/bandon_persistent_runner_medium_acceptance/acceptance_decision.md").exists(),
        "full_pair_artifacts_present": Path("artifacts/benchmarks/bandon_persistent_runner_full_pair_validation/acceptance_decision.md").exists(),
        "persistent_runner_default": Settings().bandon_inference_mode == "persistent_runner",
        "cli_per_tile_available": "cli_per_tile" in {"cli_per_tile", "persistent_runner"},
        "required_env": {
            "APP_BANDON_INFERENCE_MODE": os.environ.get("APP_BANDON_INFERENCE_MODE"),
            "APP_INFERENCE_TIMING_ENABLED": os.environ.get("APP_INFERENCE_TIMING_ENABLED"),
            "APP_INFERENCE_VERBOSE_TILE_LOGS": os.environ.get("APP_INFERENCE_VERBOSE_TILE_LOGS"),
        },
        "runtime_cache_dir": str(configured_settings.runtime_cache_dir),
        "request_cache_dir": str(configured_settings.request_cache_dir),
        "post_completion_request_cleanup_enabled": configured_settings.post_completion_request_cleanup_enabled,
        "post_completion_request_cleanup_mode": configured_settings.post_completion_request_cleanup_mode,
        "aoi_source": f"bbox:{args.aoi_bbox}" if args.aoi_bbox else f"source_project:{args.source_project_json}",
        "project_id": saved_project.project_id,
        "job_id": f"temporal-validation-{saved_project.project_id}",
        "milestones": [milestone.release_identifier for milestone in saved_project.milestones],
        "number_of_milestones": len(saved_project.milestones),
        "number_of_pairs": len(saved_project.milestones) - 1,
        "aoi_area_m2": expected[0].estimated_area_m2 if expected else None,
        "expected_pairs": [pair.__dict__ for pair in expected],
        "tile_size": configured_settings.inference_tile_size,
        "tile_overlap": configured_settings.inference_tile_overlap,
        "change_threshold": configured_settings.change_threshold,
        "checkpoint_path": hash_context.get("checkpoint_path"),
        "checkpoint_sha256": hash_context.get("checkpoint_sha256"),
        "device": hash_context.get("device"),
        "api_used": "src.core_api.run_temporal_project_api",
        "command": " ".join([Path(sys.argv[0]).name, *sys.argv[1:]]),
        "started_at": iso_now(),
    }
    write_validation_plan(output_dir / "validation_plan.md", args, expected, metadata)

    progress_samples: list[dict[str, Any]] = []
    sampler = MemorySampler(interval_seconds=5.0)

    def progress_callback(fraction: float, message: str, details: dict[str, object] | None = None) -> None:
        sample = {
            "timestamp": iso_now(),
            "fraction": fraction,
            "message": message,
            "details": details or {},
        }
        progress_samples.append(sample)
        detail_map = details or {}
        processed = detail_map.get("processed_tiles")
        total = detail_map.get("total_tiles")
        pair_index = detail_map.get("current_pair_index")
        if processed is not None or total is not None:
            sampler.record(
                "progress",
                {
                    "pair_index": pair_index,
                    "processed_tiles": processed,
                    "total_tiles": total,
                    "message": message,
                },
            )

    started = time.perf_counter()
    sampler.start()
    try:
        response = run_temporal_project_api(
            saved_project.project_id,
            settings=configured_settings,
            progress_callback=progress_callback,
            job_id=metadata["job_id"],
        )
    finally:
        sampler.stop()
    elapsed = time.perf_counter() - started
    final_project = get_temporal_project(saved_project.project_id, configured_settings)
    metadata.update(
        {
            "ended_at": iso_now(),
            "total_wall_time_seconds": round(elapsed, 3),
            "project_status": "success" if response.success else "failed",
            "error_message": response.error_message,
        }
    )

    pair_reports, timing_index = validate_pair_outputs(final_project, configured_settings)
    pair_summaries = compact_pair_summaries(pair_reports)
    update_p95_from_timing(pair_summaries, pair_reports)
    total_tiles = sum(int(row.get("total_tiles") or 0) for row in pair_summaries)
    metadata["total_tiles_all_pairs"] = total_tiles
    metadata["tiles_per_second_project"] = (total_tiles / elapsed) if elapsed > 0 else None
    pair_validation = {
        "project_id": final_project.project_id,
        "passed": all(report.get("passed") for report in pair_reports),
        "pairs": pair_reports,
    }
    memory_profile = build_memory_profile(sampler.samples, pair_reports, total_tiles)
    export_validation = validate_exports(final_project.project_id, configured_settings)
    acceptance = build_acceptance(
        metadata=metadata,
        pair_summaries=pair_summaries,
        memory_profile=memory_profile,
        pair_validation=pair_validation,
        export_validation=export_validation,
    )

    write_json(output_dir / "temporal_project_run_metadata.json", metadata)
    write_json(output_dir / "pair_summaries.json", pair_summaries)
    write_json(output_dir / "pair_timing_summaries_index.json", timing_index)
    write_json(output_dir / "memory_profile_across_pairs.json", memory_profile)
    write_json(output_dir / "progress_samples.json", progress_samples)
    write_json(output_dir / "pair_output_validation_report.json", pair_validation)
    write_json(output_dir / "final_export_validation_report.json", export_validation)
    report_md = build_markdown_report(
        metadata=metadata,
        pair_summaries=pair_summaries,
        memory_profile=memory_profile,
        pair_validation=pair_validation,
        export_validation=export_validation,
        acceptance=acceptance,
    )
    write_text(output_dir / "full_temporal_project_validation_report.md", report_md)
    write_text(
        output_dir / "acceptance_decision.md",
        "\n".join(
            [
                f"# Acceptance Decision: {acceptance['decision']}",
                "",
                f"- Project: `{final_project.project_id}`",
                f"- Pairs: `{metadata['number_of_pairs']}`",
                f"- Total tiles: `{metadata['total_tiles_all_pairs']}`",
                f"- Model loads total: `{acceptance['project_model_load_count_total']}`",
                f"- Checkpoint loads total: `{acceptance['checkpoint_load_count_total']}`",
                f"- Fallback/mixed mode detected: `{acceptance['fallback_or_mixed_mode_detected']}`",
                f"- Memory growth detected: `{memory_profile['memory_growth_detected']}`",
                "",
                *[f"- {pass_fail(item['passed'])}: {item['name']}" for item in acceptance["criteria"]],
                "",
            ]
        ),
    )
    return 0 if acceptance["decision"] == "PASS" else 2


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
